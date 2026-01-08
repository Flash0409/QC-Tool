""" Modern Production Tool - Complete Integration with Manager Dashboard
Integrates handover system + visual navigation + manager status updates
AUTO-OPENS PRODUCTION MODE when cabinet is loaded from queue
"""
import tkinter as tk
from tkinter import messagebox, simpledialog, Menu
from PIL import Image, ImageTk, ImageDraw, ImageFont
import fitz  # PyMuPDF
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime
import os
import sys
import json
import getpass
import re
import sqlite3

from handover_database import HandoverDB
from database_manager import DatabaseManager


def get_app_base_dir():
    """Returns the directory where the app is running from"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


class ManagerDB:
    """Manager database integration for status tracking"""
    
    def __init__(self, db_path):
        self.db_path = db_path
        self.init_database()
    
    def init_database(self):
        """Initialize tables if they don't exist"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS cabinets (
            cabinet_id TEXT PRIMARY KEY,
            project_name TEXT,
            sales_order_no TEXT,
            total_pages INTEGER DEFAULT 0,
            annotated_pages INTEGER DEFAULT 0,
            total_punches INTEGER DEFAULT 0,
            open_punches INTEGER DEFAULT 0,
            implemented_punches INTEGER DEFAULT 0,
            closed_punches INTEGER DEFAULT 0,
            status TEXT DEFAULT 'quality_inspection',
            created_date TEXT,
            last_updated TEXT,
            storage_location TEXT,
            excel_path TEXT
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS category_occurrences (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cabinet_id TEXT,
            project_name TEXT,
            category TEXT,
            subcategory TEXT,
            occurrence_date TEXT
        )''')
        
        # Add columns if they don't exist (migration for existing databases)
        try:
            cursor.execute('ALTER TABLE cabinets ADD COLUMN storage_location TEXT')
        except sqlite3.OperationalError:
            pass  # Column already exists
        
        try:
            cursor.execute('ALTER TABLE cabinets ADD COLUMN excel_path TEXT')
        except sqlite3.OperationalError:
            pass  # Column already exists
        
        conn.commit()
        conn.close()
    
    def update_cabinet(self, cabinet_id, project_name, sales_order_no, total_pages, annotated_pages,
                      total_punches, open_punches, implemented_punches, closed_punches, status,
                      storage_location=None, excel_path=None):
        """Update cabinet statistics WITH excel_path and storage_location"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT OR REPLACE INTO cabinets
                (cabinet_id, project_name, sales_order_no, total_pages, annotated_pages,
                 total_punches, open_punches, implemented_punches, closed_punches, status,
                 storage_location, excel_path, created_date, last_updated)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                        COALESCE((SELECT created_date FROM cabinets WHERE cabinet_id = ?), ?),
                        ?)
            ''', (cabinet_id, project_name, sales_order_no, total_pages, annotated_pages,
                  total_punches, open_punches, implemented_punches, closed_punches, status,
                  storage_location, excel_path,
                  cabinet_id, datetime.now().isoformat(),
                  datetime.now().isoformat()))
            
            conn.commit()
            conn.close()
            print(f"✓ Manager DB: Updated {cabinet_id} - Status: {status}")
            return True
        except Exception as e:
            print(f"Manager DB update error: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def update_status(self, cabinet_id, status):
        """Update cabinet status only"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                UPDATE cabinets
                SET status = ?, last_updated = ?
                WHERE cabinet_id = ?
            ''', (status, datetime.now().isoformat(), cabinet_id))
            
            conn.commit()
            conn.close()
            print(f"✓ Manager DB: Status updated for {cabinet_id} → {status}")
            return True
        except Exception as e:
            print(f"Status update error: {e}")
            return False


class ProductionTool:
    def __init__(self, root):
        self.root = root
        self.root.title("Production Tool")
        self.root.geometry("1400x900")
        
        # Data / files
        self.pdf_document = None
        self.current_pdf_path = None
        self.current_page = 0
        self.project_name = ""
        self.sales_order_no = ""
        self.cabinet_id = ""
        self.storage_location = ""
        self.annotations = []
        
        base = get_app_base_dir()
        
        # Initialize databases
        self.handover_db = HandoverDB(os.path.join(base, "inspection_tool.db")))
        self.db = DatabaseManager(os.path.join(base, "inspection_tool.db"))
        self.manager_db = ManagerDB(os.path.join(base, "manager.db"))
        
        self.excel_file = None
        self.working_excel_path = None
        self.zoom_level = 1.0
        self.current_sr_no = 1
        self.current_page_image = None
        self.session_refs = set()
        
        # Visual navigation for production mode
        self.production_arrow_id = None
        self.production_highlight_id = None
        self.production_dialog_open = False
        
        # Column mapping
        self.punch_sheet_name = 'Punch Sheet'
        self.punch_cols = {
            'sr_no': 'A',
            'ref_no': 'B',
            'desc': 'C',
            'category': 'D',
            'implemented_name': 'G',
            'implemented_date': 'H',
            'closed_name': 'I',
            'closed_date': 'J'
        }
        
        self.interphase_sheet_name = 'Interphase'
        self.interphase_cols = {
            'ref_no': 'B',
            'description': 'C',
            'status': 'D',
        }
        
        self.header_cells = {
            "Interphase": {
                "project_name": "C4",
                "sales_order": "C6",
                "cabinet_id": "F6"
            },
            "Punch Sheet": {
                "project_name": "C2",
                "sales_order": "C4",
                "cabinet_id": "H4"
            }
        }
        
        # Drawing state
        self.drawing = False
        self.drawing_type = None
        self.rect_start_x = None
        self.rect_start_y = None
        self.temp_rect_id = None
        self.selected_annotation = None
        
        self.setup_ui()
        self.current_sr_no = self.get_next_sr_no()

    # ================================================================
    # MANAGER SYNC - PRODUCTION SPECIFIC
    # ================================================================
    
    def sync_manager_stats(self):
        """Sync current cabinet statistics to manager database"""
        if not self.cabinet_id:
            return
        
        try:
            # Count from Excel - start from row 9
            implemented_punches = 0
            closed_punches = 0
            total_punches = 0
            
            if self.excel_file and os.path.exists(self.excel_file):
                try:
                    wb = load_workbook(self.excel_file, data_only=True)
                    ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
                    
                    row = 9  # Start from row 9
                    while row <= ws.max_row + 5:
                        # Check if row has a punch (has checked_name in column E)
                        checked = self.read_cell(ws, row, 'E')
                        if not checked:
                            row += 1
                            if row > 2000:  # Safety limit
                                break
                            continue
                        
                        total_punches += 1
                        implemented = self.read_cell(ws, row, self.punch_cols['implemented_name'])
                        closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
                        
                        if closed:
                            closed_punches += 1
                        elif implemented:
                            implemented_punches += 1
                        
                        row += 1
                        if row > 2000:  # Safety limit
                            break
                    
                    wb.close()
                except Exception as e:
                    print(f"Excel read error: {e}")
            
            # Calculate open punches
            open_punches = total_punches - implemented_punches - closed_punches
            
            # Update manager database with production status
            self.manager_db.update_cabinet(
                self.cabinet_id,
                self.project_name,
                self.sales_order_no,
                0,  # total_pages (not relevant in production)
                0,  # annotated_pages
                total_punches,
                open_punches,
                implemented_punches,
                closed_punches,
                'in_progress',  # Production status
                storage_location=getattr(self, 'storage_location', None),
                excel_path=self.excel_file
            )
        
        except Exception as e:
            print(f"Manager sync error: {e}")
            import traceback
            traceback.print_exc()

    # ================================================================
    # CELL HELPERS
    # ================================================================
    
    def split_cell(self, cell_ref):
        m = re.match(r"([A-Z]+)(\d+)", cell_ref)
        if not m:
            raise ValueError(f"Invalid cell reference: {cell_ref}")
        col, row = m.groups()
        return int(row), col
    
    def _resolve_merged_target(self, ws, row, col_idx):
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col_idx <= merged.max_col:
                return merged.min_row, merged.min_col
        return row, col_idx
    
    def write_cell(self, ws, row, col, value):
        if isinstance(col, str):
            col_idx = column_index_from_string(col)
        else:
            col_idx = int(col)
        target_row, target_col = self._resolve_merged_target(ws, int(row), col_idx)
        ws.cell(row=target_row, column=target_col).value = value
    
    def read_cell(self, ws, row, col):
        if isinstance(col, str):
            col_idx = column_index_from_string(col)
        else:
            col_idx = int(col)
        target_row, target_col = self._resolve_merged_target(ws, int(row), col_idx)
        return ws.cell(row=target_row, column=target_col).value

    # ================================================================
    # MODERN UI SETUP
    # ================================================================
    
    def setup_ui(self):
        """Setup modern professional UI"""
        # Main toolbar with modern styling
        toolbar = tk.Frame(self.root, bg='#1e293b', height=70)
        toolbar.pack(side=tk.TOP, fill=tk.X)
        
        # Enhanced Menu Bar
        menubar = Menu(self.root, bg='#1e293b', fg='white', activebackground='#3b82f6')
        self.root.config(menu=menubar)
        
        # File Menu
        file_menu = Menu(menubar, tearoff=0, bg='#1e293b', fg='white', activebackground='#3b82f6')
        menubar.add_cascade(label="📁 File", menu=file_menu)
        file_menu.add_command(label="Load from Production Queue", command=self.load_from_handover_queue, accelerator="Ctrl+O")
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)
        
        # Tools Menu
        tools_menu = Menu(menubar, tearoff=0, bg='#1e293b', fg='white', activebackground='#3b82f6')
        menubar.add_cascade(label="🛠️ Tools", menu=tools_menu)
        tools_menu.add_command(label="🏭 Production Mode", command=self.production_mode, accelerator="Ctrl+P")
        tools_menu.add_separator()
        tools_menu.add_command(label="✅ Complete & Handback", command=self.complete_rework_handback, accelerator="Ctrl+H")
        
        # View Menu
        view_menu = Menu(menubar, tearoff=0, bg='#1e293b', fg='white', activebackground='#3b82f6')
        menubar.add_cascade(label="👁️ View", menu=view_menu)
        view_menu.add_command(label="Zoom In", command=self.zoom_in, accelerator="Ctrl++")
        view_menu.add_command(label="Zoom Out", command=self.zoom_out, accelerator="Ctrl+-")
        view_menu.add_command(label="Reset Zoom", command=lambda: setattr(self, 'zoom_level', 1.0) or self.display_page())
        
        # Keyboard shortcuts
        self.root.bind_all("<Control-o>", lambda e: self.load_from_handover_queue())
        self.root.bind_all("<Control-p>", lambda e: self.production_mode())
        self.root.bind_all("<Control-h>", lambda e: self.complete_rework_handback())
        self.root.bind_all("<Control-plus>", lambda e: self.zoom_in())
        self.root.bind_all("<Control-minus>", lambda e: self.zoom_out())
        
        # Modern button style
        btn_style = {
            'bg': '#3b82f6',
            'fg': 'white',
            'padx': 12,
            'pady': 10,
            'font': ('Segoe UI', 9, 'bold'),
            'relief': tk.FLAT,
            'borderwidth': 0,
            'cursor': 'hand2'
        }
        
        # Left section - Load operations
        left_frame = tk.Frame(toolbar, bg='#1e293b')
        left_frame.pack(side=tk.LEFT, padx=10, pady=10)
        
        tk.Button(left_frame, text="📦 Load from Queue", command=self.load_from_handover_queue,
                 bg='#8b5cf6', fg='white', padx=15, pady=10,
                 font=('Segoe UI', 10, 'bold'), relief=tk.FLAT, borderwidth=0,
                 cursor='hand2').pack(side=tk.LEFT, padx=3)
        
        # Center section - Navigation
        center_frame = tk.Frame(toolbar, bg='#1e293b')
        center_frame.pack(side=tk.LEFT, padx=20)
        
        self.page_label = tk.Label(center_frame, text="Page: 0/0", bg='#1e293b', fg='white',
                                   font=('Segoe UI', 10, 'bold'))
        self.page_label.pack(side=tk.LEFT, padx=10)
        
        nav_btn_style = btn_style.copy()
        nav_btn_style['bg'] = '#64748b'
        
        tk.Button(center_frame, text="◀", command=self.prev_page, width=3,
                 **nav_btn_style).pack(side=tk.LEFT, padx=2)
        tk.Button(center_frame, text="▶", command=self.next_page, width=3,
                 **nav_btn_style).pack(side=tk.LEFT, padx=2)
        
        # Zoom controls
        zoom_frame = tk.Frame(center_frame, bg='#1e293b')
        zoom_frame.pack(side=tk.LEFT, padx=15)
        
        zoom_btn_style = btn_style.copy()
        zoom_btn_style['bg'] = '#10b981'
        
        tk.Button(zoom_frame, text="🔍+", command=self.zoom_in, width=4,
                 **zoom_btn_style).pack(side=tk.LEFT, padx=2)
        tk.Button(zoom_frame, text="🔍−", command=self.zoom_out, width=4,
                 **zoom_btn_style).pack(side=tk.LEFT, padx=2)
        
        # Right section - Action buttons
        right_frame = tk.Frame(toolbar, bg='#1e293b')
        right_frame.pack(side=tk.RIGHT, padx=10, pady=10)
        
        tk.Button(right_frame, text="🏭 Production Mode", command=self.production_mode,
                 bg='#f59e0b', fg='white', padx=15, pady=10,
                 font=('Segoe UI', 9, 'bold'), relief=tk.FLAT, borderwidth=0,
                 cursor='hand2').pack(side=tk.LEFT, padx=3)
        
        tk.Button(right_frame, text="✅ Handback to Quality", command=self.complete_rework_handback,
                 bg='#10b981', fg='white', padx=15, pady=10,
                 font=('Segoe UI', 9, 'bold'), relief=tk.FLAT, borderwidth=0,
                 cursor='hand2').pack(side=tk.LEFT, padx=3)
        
        # Canvas with scrollbars
        canvas_frame = tk.Frame(self.root, bg='#f1f5f9')
        canvas_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        v_scrollbar = tk.Scrollbar(canvas_frame, orient=tk.VERTICAL)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        h_scrollbar = tk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.canvas = tk.Canvas(canvas_frame, bg='#f8fafc',
                               yscrollcommand=v_scrollbar.set,
                               xscrollcommand=h_scrollbar.set,
                               highlightthickness=0)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        v_scrollbar.config(command=self.canvas.yview)
        h_scrollbar.config(command=self.canvas.xview)
        
        # Bind mouse events
        self.canvas.bind("<Double-Button-1>", self.on_double_left_zoom)
        self.canvas.bind("<Double-Button-3>", self.on_double_right_zoom)
        
        # Modern status bar
        status_bar = tk.Frame(self.root, bg='#334155', height=40)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        instructions_text = "🏭 Production Tool | Load items from queue → Production Mode (auto-opens) → Mark as done → Handback to Quality"
        tk.Label(status_bar, text=instructions_text, bg='#334155', fg='#e2e8f0',
                font=('Segoe UI', 9), pady=10).pack()

    # ================================================================
    # LOAD FROM HANDOVER QUEUE - WITH AUTO-OPEN PRODUCTION MODE
    # ================================================================
    
    def load_from_handover_queue(self):
        """Load item from production handover queue with modern UI"""
        pending_items = self.handover_db.get_pending_production_items()
        
        if not pending_items:
            messagebox.showinfo("No Items", 
                              "✓ No items in production queue.\n"
                              "All items have been processed!", 
                              icon='info')
            return
        
        # Create modern selection dialog
        dlg = tk.Toplevel(self.root)
        dlg.title("Production Queue")
        dlg.geometry("1000x600")
        dlg.configure(bg='#f8fafc')
        dlg.transient(self.root)
        dlg.grab_set()
        
        # Header
        header_frame = tk.Frame(dlg, bg='#8b5cf6', height=60)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="📦 Production Queue - Select Item",
                bg='#8b5cf6', fg='white',
                font=('Segoe UI', 14, 'bold')).pack(pady=15)
        
        # Info bar
        info_frame = tk.Frame(dlg, bg='#eff6ff')
        info_frame.pack(fill=tk.X, padx=20, pady=(15, 5))
        
        tk.Label(info_frame, text=f"Total items in queue: {len(pending_items)}",
                bg='#eff6ff', fg='#1e40af',
                font=('Segoe UI', 10, 'bold')).pack(pady=8)
        
        # Listbox frame
        list_frame = tk.Frame(dlg, bg='white')
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        tk.Label(list_frame, text="Select item to load:",
                font=('Segoe UI', 10, 'bold'), bg='white', fg='#1e293b').pack(anchor='w', pady=(0, 10))
        
        # Scrollbar and Listbox
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        listbox = tk.Listbox(list_frame, font=('Consolas', 9),
                            yscrollcommand=scrollbar.set,
                            bg='#f8fafc', relief=tk.FLAT,
                            selectmode=tk.SINGLE, height=18)
        listbox.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=listbox.yview)
        
        # Populate listbox
        for item in pending_items:
            status_icon = "⚙️" if item['status'] == 'in_progress' else "📦"
            display = (
                f"{status_icon} {item['cabinet_id']:20} | {item['project_name']:30} | "
                f"Punches: {item['open_punches']:3} | By: {item['handed_over_by']:15} | "
                f"{item['handed_over_date'][:10]}"
            )
            listbox.insert(tk.END, display)
        
        def load_selected():
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select an item first.")
                return
            
            item = pending_items[selection[0]]
            dlg.destroy()
            self.load_handover_item(item)
        
        # Buttons
        btn_frame = tk.Frame(dlg, bg='#f8fafc')
        btn_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        btn_style = {
            'font': ('Segoe UI', 10, 'bold'),
            'relief': tk.FLAT,
            'cursor': 'hand2',
            'padx': 20,
            'pady': 12
        }
        
        tk.Button(btn_frame, text="📂 Load Selected", command=load_selected,
                 bg='#3b82f6', fg='white', **btn_style).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Cancel", command=dlg.destroy,
                 bg='#64748b', fg='white', **btn_style).pack(side=tk.RIGHT, padx=5)
        
        listbox.bind('<Double-Button-1>', lambda e: load_selected())
    
    def load_handover_item(self, item):
        """Load a specific handover item - WITH AUTO-OPEN PRODUCTION MODE"""
        try:
            # Verify files exist
            if not os.path.exists(item['pdf_path']):
                messagebox.showerror("File Not Found", 
                                   f"PDF not found:\n{item['pdf_path']}")
                return
            
            if not os.path.exists(item['excel_path']):
                messagebox.showerror("File Not Found", 
                                   f"Excel not found:\n{item['excel_path']}")
                return
            
            # Get project from database to get storage location
            project_data = self.db.get_project(item['cabinet_id'])
            if not project_data:
                messagebox.showerror("Error", "Project not found in database")
                return
            
            # Load PDF
            self.pdf_document = fitz.open(item['pdf_path'])
            self.current_pdf_path = item['pdf_path']
            self.current_page = 0
            self.zoom_level = 1.0
            
            # Set project details
            self.cabinet_id = item['cabinet_id']
            self.project_name = item['project_name']
            self.sales_order_no = item['sales_order_no']
            self.storage_location = project_data['storage_location']
            
            # Set Excel
            self.excel_file = item['excel_path']
            self.working_excel_path = item['excel_path']
            
            # Load session if available
            if item.get('session_path') and os.path.exists(item['session_path']):
                self.load_session_from_path(item['session_path'])
            else:
                self.annotations = []
                self.session_refs.clear()
            
            # Mark as in progress in handover DB
            try:
                username = os.getlogin()
            except:
                username = getpass.getuser()
            
            self.handover_db.update_production_status(
                item['cabinet_id'],
                status='in_progress',
                user=username
            )
            
            # UPDATE MANAGER STATUS TO "IN_PROGRESS"
            self.manager_db.update_status(self.cabinet_id, 'in_progress')
            
            # SYNC INITIAL STATS TO MANAGER
            self.sync_manager_stats()
            
            self.display_page()
            
            # Show brief notification
            messagebox.showinfo(
                "Item Loaded",
                f"✓ Loaded from Production Queue:\n\n"
                f"Cabinet: {self.cabinet_id}\n"
                f"Project: {self.project_name}\n"
                f"Open Punches: {item['open_punches']}\n\n"
                f"Production Mode will open automatically...",
                icon='info'
            )
            
            # AUTO-OPEN PRODUCTION MODE after brief delay
            self.root.after(500, self.production_mode)
        
        except Exception as e:
            messagebox.showerror("Load Error", f"Failed to load item:\n{e}")
            import traceback
            traceback.print_exc()

    # ================================================================
    # UPDATED: COMPLETE REWORK & HANDBACK - CHECK IMPLEMENTED COLUMN
    # ================================================================
    
    def complete_rework_handback(self):
        """Complete rework and handback to Quality - CHECK IMPLEMENTED COLUMN"""
        if not self.pdf_document or not self.excel_file:
            messagebox.showwarning("No Item Loaded", 
                                 "Please load an item from the production queue first.")
            return
        
        # Check if item is from handover
        item = self.handover_db.get_item_by_cabinet_id(self.cabinet_id, "quality_to_production")
        if not item:
            messagebox.showwarning("Not from Queue", 
                                 "This item was not loaded from the production queue.\n"
                                 "Only items from the handover queue can be handed back.")
            return
        
        # NEW: Check for punches without "Implemented By"
        not_implemented = self.get_punches_without_implementation()
        if not_implemented:
            # Show dialog with list of not-implemented punches
            self.show_not_implemented_dialog(not_implemented)
            return
        
        # Get remarks
        remarks = simpledialog.askstring(
            "Production Remarks",
            "Enter any remarks or notes for Quality team:\n\n"
            "(Optional - Leave blank if none)",
            parent=self.root
        )
        
        try:
            username = os.getlogin()
        except:
            username = getpass.getuser()
        
        handback_data = {
            "cabinet_id": self.cabinet_id,
            "project_name": self.project_name,
            "sales_order_no": self.sales_order_no,
            "pdf_path": self.current_pdf_path,
            "excel_path": self.excel_file,
            "session_path": self.get_session_path_for_pdf(),
            "rework_completed_by": username,
            "rework_completed_date": datetime.now().isoformat(),
            "production_remarks": remarks or "No remarks"
        }
        
        success = self.handover_db.add_production_handback(handback_data)
        
        if success:
            # SYNC FINAL STATS TO MANAGER
            self.sync_manager_stats()
            
            # UPDATE MANAGER STATUS TO "BEING_CLOSED_BY_QUALITY"
            self.manager_db.update_status(self.cabinet_id, 'being_closed_by_quality')
            
            messagebox.showinfo(
                "Handback Complete",
                f"✓ Successfully handed back to Quality:\n\n"
                f"Cabinet: {self.cabinet_id}\n"
                f"Project: {self.project_name}\n\n"
                f"Quality team will verify and close this item.",
                icon='info'
            )
            
            # Clear current work
            self.pdf_document = None
            self.current_pdf_path = None
            self.excel_file = None
            self.annotations = []
            self.canvas.delete("all")
            self.page_label.config(text="Page: 0/0")
            self.root.title("Production Tool")
        else:
            messagebox.showerror("Error", "Failed to handback item to Quality.")
    
    def get_punches_without_implementation(self):
        """Get list of punches that don't have 'Implemented By' filled"""
        not_implemented = []
        
        try:
            if not self.excel_file or not os.path.exists(self.excel_file):
                return not_implemented
            
            wb = load_workbook(self.excel_file, data_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
            
            row = 9  # Start from row 9
            while row <= ws.max_row + 5:
                # Check if row has a punch (column E has value)
                checked = self.read_cell(ws, row, 'E')
                if not checked:
                    row += 1
                    if row > 2000:  # Safety limit
                        break
                    continue
                
                # Check if already closed
                closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
                if closed:
                    row += 1
                    continue
                
                # Check if implemented
                implemented = self.read_cell(ws, row, self.punch_cols['implemented_name'])
                if not implemented:
                    # This punch needs implementation
                    sr_no = self.read_cell(ws, row, self.punch_cols['sr_no'])
                    ref_no = self.read_cell(ws, row, self.punch_cols['ref_no'])
                    desc = self.read_cell(ws, row, self.punch_cols['desc'])
                    category = self.read_cell(ws, row, self.punch_cols['category'])
                    
                    not_implemented.append({
                        'row': row,
                        'sr_no': sr_no,
                        'ref_no': ref_no,
                        'description': desc,
                        'category': category
                    })
                
                row += 1
                if row > 2000:  # Safety limit
                    break
            
            wb.close()
            return not_implemented
        
        except Exception as e:
            print(f"Error checking implementation: {e}")
            return []
    
    def show_not_implemented_dialog(self, not_implemented):
        """Show dialog listing punches without implementation"""
        dlg = tk.Toplevel(self.root)
        dlg.title("⚠️ Implementation Required")
        dlg.geometry("800x600")
        dlg.configure(bg='#fef3c7')
        dlg.transient(self.root)
        dlg.grab_set()
        
        # Header
        header_frame = tk.Frame(dlg, bg='#f59e0b', height=60)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="⚠️ IMPLEMENTATION REQUIRED",
                bg='#f59e0b', fg='white',
                font=('Segoe UI', 14, 'bold')).pack(pady=15)
        
        # Info
        info_frame = tk.Frame(dlg, bg='#fef3c7')
        info_frame.pack(fill=tk.X, padx=20, pady=15)
        
        tk.Label(info_frame, 
                text=f"The following {len(not_implemented)} punch(es) have not been marked as 'Implemented'.\n"
                     "Please complete implementation before handing back to Quality.",
                font=('Segoe UI', 11), bg='#fef3c7', fg='#78350f',
                justify='left').pack(anchor='w')
        
        # List frame
        list_frame = tk.Frame(dlg, bg='white')
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        tk.Label(list_frame, text="Punches requiring implementation:",
                font=('Segoe UI', 10, 'bold'), bg='white', fg='#1e293b').pack(anchor='w', padx=10, pady=(10, 5))
        
        # Scrollbar and Text widget
        scroll_frame = tk.Frame(list_frame, bg='white')
        scroll_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        scrollbar = tk.Scrollbar(scroll_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        text_widget = tk.Text(scroll_frame, wrap=tk.WORD, font=('Courier New', 9),
                            yscrollcommand=scrollbar.set, bg='#f8fafc', relief=tk.FLAT,
                            padx=10, pady=10)
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=text_widget.yview)
        
        # Populate list
        for idx, punch in enumerate(not_implemented, 1):
            text_widget.insert(tk.END, f"\n{'='*70}\n")
            text_widget.insert(tk.END, f"#{idx} - SR No: {punch['sr_no']} | Ref: {punch['ref_no']}\n")
            text_widget.insert(tk.END, f"Category: {punch['category']}\n")
            text_widget.insert(tk.END, f"\nDescription:\n{punch['description']}\n")
        
        text_widget.config(state=tk.DISABLED)
        
        # Button
        tk.Button(dlg, text="OK - I'll Complete Implementation First",
                 command=dlg.destroy, bg='#f59e0b', fg='white',
                 font=('Segoe UI', 10, 'bold'), padx=20, pady=12,
                 relief=tk.FLAT, cursor='hand2').pack(pady=20)
    
    def count_open_punches(self):
        """Count open punches in current Excel"""
        try:
            if not self.excel_file or not os.path.exists(self.excel_file):
                return 0
            
            wb = load_workbook(self.excel_file, data_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
            
            open_count = 0
            row = 9  # Start from row 9
            while row <= ws.max_row + 5:
                # Check if row has a punch
                checked = self.read_cell(ws, row, 'E')
                if not checked:
                    row += 1
                    if row > 2000:
                        break
                    continue
                
                closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
                if not closed:
                    open_count += 1
                
                row += 1
                if row > 2000:
                    break
            
            wb.close()
            return open_count
        
        except Exception as e:
            print(f"Error counting open punches: {e}")
            return 0

    # ================================================================
    # ENHANCED PRODUCTION MODE WITH VISUAL NAVIGATION - MODERN UI
    # ================================================================
    
    def production_mode(self):
        """Enhanced production mode with visual navigation and modern UI"""
        if not self.pdf_document or not self.excel_file:
            messagebox.showwarning("No Item", 
                                 "Please load an item from the production queue first.")
            return
        
        punches = self.read_open_punches_from_excel()
        
        if not punches:
            messagebox.showinfo("No Punches", 
                              "✓ All punches are closed!\n"
                              "You can now handback to Quality.", 
                              icon='info')
            return
        
        punches.sort(key=lambda p: (p['implemented'], p['sr_no']))
        
        # Modern dialog
        dlg = tk.Toplevel(self.root)
        dlg.title("Production Mode")
        dlg.geometry("900x550")
        dlg.configure(bg='#f8fafc')
        dlg.transient(self.root)
        dlg.grab_set()
        
        self.production_dialog_open = True
        
        # Header
        header_frame = tk.Frame(dlg, bg='#f59e0b', height=60)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="🏭 PRODUCTION MODE",
                bg='#f59e0b', fg='white',
                font=('Segoe UI', 14, 'bold')).pack(pady=15)
        
        # Progress
        progress_frame = tk.Frame(dlg, bg='#f8fafc')
        progress_frame.pack(fill=tk.X, padx=20, pady=(15, 5))
        
        idx_label = tk.Label(progress_frame, text="",
                           font=('Segoe UI', 11, 'bold'),
                           bg='#f8fafc', fg='#1e293b')
        idx_label.pack()
        
        # Info cards
        info_frame = tk.Frame(dlg, bg='#f8fafc')
        info_frame.pack(fill=tk.X, padx=20, pady=10)
        
        # SR Number card
        sr_card = tk.Frame(info_frame, bg='#dbeafe', relief=tk.FLAT)
        sr_card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        tk.Label(sr_card, text="SR No.", font=('Segoe UI', 8),
                bg='#dbeafe', fg='#1e40af').pack(anchor='w', padx=10, pady=(8, 2))
        
        sr_label = tk.Label(sr_card, text="", font=('Segoe UI', 12, 'bold'),
                          bg='#dbeafe', fg='#1e293b')
        sr_label.pack(anchor='w', padx=10, pady=(0, 8))
        
        # Reference card
        ref_card = tk.Frame(info_frame, bg='#e0e7ff', relief=tk.FLAT)
        ref_card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        tk.Label(ref_card, text="Reference", font=('Segoe UI', 8),
                bg='#e0e7ff', fg='#4338ca').pack(anchor='w', padx=10, pady=(8, 2))
        
        ref_label = tk.Label(ref_card, text="", font=('Segoe UI', 12, 'bold'),
                           bg='#e0e7ff', fg='#1e293b')
        ref_label.pack(anchor='w', padx=10, pady=(0, 8))
        
        # Status card
        status_card = tk.Frame(info_frame, bg='#fef3c7', relief=tk.FLAT)
        status_card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
        
        tk.Label(status_card, text="Status", font=('Segoe UI', 8),
                bg='#fef3c7', fg='#92400e').pack(anchor='w', padx=10, pady=(8, 2))
        
        impl_label = tk.Label(status_card, text="", font=('Segoe UI', 12, 'bold'),
                            bg='#fef3c7', fg='#1e293b')
        impl_label.pack(anchor='w', padx=10, pady=(0, 8))
        
        # Content
        content_frame = tk.Frame(dlg, bg='white', relief=tk.FLAT)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        tk.Label(content_frame, text="Punch Description:",
                font=('Segoe UI', 9, 'bold'), bg='white', fg='#64748b',
                anchor='w').pack(fill=tk.X, padx=15, pady=(10, 5))
        
        text_widget = tk.Text(content_frame, wrap=tk.WORD, height=12,
                            font=('Segoe UI', 10), bg='#f8fafc', relief=tk.FLAT,
                            padx=10, pady=10)
        text_widget.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))
        text_widget.config(state=tk.DISABLED)
        
        pos = [0]
        
        def show_item():
            p = punches[pos[0]]
            
            # Update progress
            progress_text = f"Item {pos[0]+1} of {len(punches)}"
            progress_pct = f"({int((pos[0]+1)/len(punches)*100)}% complete)"
            idx_label.config(text=f"{progress_text} {progress_pct}")
            
            # Update info cards
            sr_label.config(text=str(p['sr_no']))
            ref_label.config(text=str(p['ref_no']))
            
            impl_status = "✓ Implemented" if p['implemented'] else "⚠ Not Implemented"
            impl_color = '#10b981' if p['implemented'] else '#f59e0b'
            impl_label.config(text=impl_status, fg=impl_color)
            
            # Update description
            text_widget.config(state=tk.NORMAL)
            text_widget.delete("1.0", tk.END)
            text_widget.insert(tk.END, p['punch_text'])
            text_widget.insert(tk.END, f"\n\n───────────────────────\n")
            text_widget.insert(tk.END, f"Category: {p['category']}\n")
            text_widget.insert(tk.END, f"Implementation: {'YES' if p['implemented'] else 'NO'}\n")
            
            ann = next((a for a in self.annotations if a.get('excel_row') == p['row']), None)
            if ann and ann.get('implementation_remark'):
                text_widget.insert(tk.END, f"\n───────────────────────\n")
                text_widget.insert(tk.END, "Previous Remarks:\n")
                text_widget.insert(tk.END, ann['implementation_remark'])
            
            text_widget.config(state=tk.DISABLED)
            
            # Navigate to annotation on PDF
            self.navigate_to_punch(p['sr_no'], p['punch_text'])
        
        show_item()
        
        def mark_implemented():
            p = punches[pos[0]]
            
            try:
                default_user = os.getlogin()
            except:
                default_user = getpass.getuser()
            
            name = simpledialog.askstring("Implemented By",
                                        "Enter your name:",
                                        initialvalue=default_user,
                                        parent=dlg)
            if not name:
                return
            
            remark = simpledialog.askstring("Remarks (optional)",
                                          "Add remarks about the implementation (optional):",
                                          parent=dlg)
            
            try:
                wb = load_workbook(self.excel_file)
                ws = wb[self.punch_sheet_name]
                
                self.write_cell(ws, p['row'], self.punch_cols['implemented_name'], name)
                self.write_cell(ws, p['row'], self.punch_cols['implemented_date'],
                              datetime.now().strftime("%Y-%m-%d"))
                
                wb.save(self.excel_file)
                wb.close()
                
                # SYNC TO MANAGER AFTER MARKING IMPLEMENTED
                self.sync_manager_stats()
            
            except PermissionError:
                messagebox.showerror("File Locked",
                                   "⚠️ Please close the Excel file and try again.",
                                   parent=dlg)
                return
            except Exception as e:
                messagebox.showerror("Excel Error", str(e), parent=dlg)
                return
            
            # Update annotation
            ann = next((a for a in self.annotations if a.get('sr_no') == p['sr_no']), None)
            if ann:
                ann['implemented'] = True
                ann['implemented_name'] = name
                ann['implemented_date'] = datetime.now().isoformat()
                ann['implementation_remark'] = remark
            
            if pos[0] < len(punches) - 1:
                pos[0] += 1
                show_item()
            else:
                messagebox.showinfo("Complete",
                                  "✓ All punches reviewed!\n"
                                  "You can now handback to Quality.",
                                  icon='info', parent=dlg)
                self.clear_production_visuals()
                self.production_dialog_open = False
                dlg.destroy()
        
        def next_item():
            if pos[0] < len(punches) - 1:
                pos[0] += 1
                show_item()
        
        def prev_item():
            if pos[0] > 0:
                pos[0] -= 1
                show_item()
        
        def on_close():
            self.clear_production_visuals()
            self.production_dialog_open = False
            dlg.destroy()
        
        dlg.protocol("WM_DELETE_WINDOW", on_close)
        
        # Modern button frame
        btn_frame = tk.Frame(dlg, bg='#f8fafc')
        btn_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        btn_style = {
            'font': ('Segoe UI', 10, 'bold'),
            'relief': tk.FLAT,
            'borderwidth': 0,
            'cursor': 'hand2',
            'padx': 20,
            'pady': 12
        }
        
        tk.Button(btn_frame, text="◀ Previous", command=prev_item,
                 bg='#94a3b8', fg='white', width=12, **btn_style).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="✓ MARK DONE", command=mark_implemented,
                 bg='#10b981', fg='white', width=16, **btn_style).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Next ▶", command=next_item,
                 bg='#94a3b8', fg='white', width=12, **btn_style).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Close", command=on_close,
                 bg='#64748b', fg='white', width=10, **btn_style).pack(side=tk.RIGHT, padx=5)
    
    def navigate_to_punch(self, sr_no, punch_text):
        """Navigate to annotation and highlight it visually"""
        # Find matching annotation
        target_ann = None
        
        # First try SR No match
        for ann in self.annotations:
            if ann.get('sr_no') == sr_no and ann.get('type') == 'error':
                target_ann = ann
                break
        
        # If not found, try fuzzy text match
        if not target_ann:
            best_match = None
            best_score = 0
            
            for ann in self.annotations:
                if ann.get('type') == 'error' and ann.get('punch_text'):
                    # Simple text similarity
                    ann_text = str(ann['punch_text']).lower()
                    search_text = str(punch_text).lower()
                    
                    if search_text in ann_text or ann_text in search_text:
                        score = len(set(search_text.split()) & set(ann_text.split()))
                        if score > best_score:
                            best_score = score
                            best_match = ann
            
            if best_match:
                target_ann = best_match
        
        # Clear previous visuals
        self.clear_production_visuals()
        
        if target_ann:
            # Navigate to page
            if target_ann.get('page') is not None:
                self.current_page = target_ann['page']
                self.display_page()
            
            # Draw arrow and highlight
            if 'bbox_page' in target_ann:
                self.highlight_annotation(target_ann)
                self._last_highlighted_ann = target_ann
    
    def highlight_annotation(self, annotation):
        """Draw visual indicators for the annotation"""
        bbox_display = self.bbox_page_to_display(annotation['bbox_page'])
        x1, y1, x2, y2 = bbox_display
        
        # Calculate center
        cx = (x1 + x2) / 2
        cy = (y1 + y2) / 2
        
        # Draw pulsing highlight
        padding = 20
        self.production_highlight_id = self.canvas.create_rectangle(
            x1 - padding, y1 - padding,
            x2 + padding, y2 + padding,
            outline='#ef4444', width=5, dash=(10, 5)
        )
        
        # Draw arrow pointing to it
        arrow_start_x = cx - 100
        arrow_start_y = cy - 100
        
        self.production_arrow_id = self.canvas.create_line(
            arrow_start_x, arrow_start_y,
            cx - 15, cy - 15,
            arrow=tk.LAST, fill='#ef4444', width=4
        )
        
        # Add text label
        self.canvas.create_text(
            arrow_start_x - 10, arrow_start_y - 10,
            text=f"SR {annotation.get('sr_no', '?')}",
            fill='#ef4444',
            font=('Segoe UI', 12, 'bold'),
            anchor='se'
        )
        
        # Scroll to make it visible
        bbox_all = self.canvas.bbox("all")
        if bbox_all:
            self.canvas.yview_moveto(max(0, (y1 - 100) / max(1, bbox_all[3])))
            self.canvas.xview_moveto(max(0, (x1 - 100) / max(1, bbox_all[2])))
    
    def clear_production_visuals(self):
        """Clear production mode visual indicators"""
        if self.production_arrow_id:
            try:
                self.canvas.delete(self.production_arrow_id)
            except:
                pass
            self.production_arrow_id = None
        
        if self.production_highlight_id:
            try:
                self.canvas.delete(self.production_highlight_id)
            except:
                pass
            self.production_highlight_id = None
    
    def read_open_punches_from_excel(self):
        """Read open punches from Excel - start from row 9"""
        punches = []
        
        if not self.excel_file or not os.path.exists(self.excel_file):
            return punches
        
        wb = load_workbook(self.excel_file, data_only=True)
        ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
        
        row = 9  # Start from row 9
        while True:
            # Check if row has a punch
            checked = self.read_cell(ws, row, 'E')
            if not checked:
                row += 1
                if row > 2000:
                    break
                continue
            
            closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
            if closed:
                row += 1
                continue
            
            implemented = bool(self.read_cell(ws, row, self.punch_cols['implemented_name']))
            sr = self.read_cell(ws, row, self.punch_cols['sr_no'])
            
            punches.append({
                'sr_no': sr,
                'row': row,
                'ref_no': self.read_cell(ws, row, self.punch_cols['ref_no']),
                'punch_text': self.read_cell(ws, row, self.punch_cols['desc']),
                'category': self.read_cell(ws, row, self.punch_cols['category']),
                'implemented': implemented
            })
            
            row += 1
            if row > 2000:
                break
        
        wb.close()
        return punches

    # ================================================================
    # PDF DISPLAY HELPERS
    # ================================================================
    
    def get_next_sr_no(self):
        try:
            if not self.excel_file or not os.path.exists(self.excel_file):
                return 1
            
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
            
            last_sr_no = 0
            row_num = 9
            
            while row_num <= ws.max_row + 5:
                val = self.read_cell(ws, row_num, self.punch_cols['sr_no'])
                if val is None:
                    break
                try:
                    last_sr_no = int(val)
                except:
                    pass
                row_num += 1
            
            wb.close()
            return last_sr_no + 1
        except Exception:
            return 1
    
    def page_to_display_scale(self):
        return 2.0 * self.zoom_level
    
    def page_to_display_coords(self, pts):
        """Convert page-space coordinates to display-space coordinates.
        Handles:
        - Single point: (x, y) -> (x*scale, y*scale)
        - List of points: [(x1,y1), ...] -> [(x1*scale, y1*scale), ...]
        """
        scale = self.page_to_display_scale()
        
        # Handle single point tuple
        if isinstance(pts, tuple) and len(pts) == 2:
            if not isinstance(pts[0], (list, tuple)):
                return (pts[0] * scale, pts[1] * scale)
        
        # Handle list of points
        return [(x * scale, y * scale) for x, y in pts]
    
    def bbox_page_to_display(self, bbox_page):
        scale = self.page_to_display_scale()
        x1, y1, x2, y2 = bbox_page
        return (x1 * scale, y1 * scale, x2 * scale, y2 * scale)
    
    def bbox_display_to_page(self, bbox_display):
        scale = self.page_to_display_scale()
        x1, y1, x2, y2 = bbox_display
        return (x1 / scale, y1 / scale, x2 / scale, y2 / scale)
    
    def display_page(self):
        if not self.pdf_document:
            self.canvas.delete("all")
            self.page_label.config(text="Page: 0/0")
            return
        
        try:
            page = self.pdf_document[self.current_page]
            mat = fitz.Matrix(self.page_to_display_scale(), self.page_to_display_scale())
            pix = page.get_pixmap(matrix=mat)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            
            self.current_page_image = None
            draw = ImageDraw.Draw(img, 'RGBA')
            
            # Try to load a font for text annotations
            try:
                font_size = max(12, int(14 * self.zoom_level))
                font = ImageFont.truetype("arial.ttf", font_size)
            except:
                font = ImageFont.load_default()
            
            for ann in self.annotations:
                if ann.get('page') != self.current_page:
                    continue
                
                ann_type = ann.get('type')
                
                # -------- RECTANGLE ANNOTATIONS (ok/error) --------
                if ann_type in ('ok', 'error') and 'bbox_page' in ann:
                    x1d, y1d, x2d, y2d = self.bbox_page_to_display(ann['bbox_page'])
                    is_selected = (self.selected_annotation is ann)
                    w = int(5 * self.zoom_level) if is_selected else int(3 * self.zoom_level)
                    
                    if ann_type == 'ok':
                        draw.rectangle([x1d, y1d, x2d, y2d],
                                     fill=(0, 255, 0, 80),
                                     outline='blue' if is_selected else 'green',
                                     width=w)
                    else:
                        draw.rectangle([x1d, y1d, x2d, y2d],
                                     fill=(255, 165, 0, 120),
                                     outline='blue' if is_selected else 'orange',
                                     width=w)
                    
                    if ann.get('closed_by'):
                        cx = x1d + 8
                        cy = y1d + 8
                        draw.ellipse([cx - 6, cy - 6, cx + 6, cy + 6],
                                   fill=(0, 128, 0, 200))
                
                # -------- PEN STROKES --------
                elif ann_type == 'pen' and 'points' in ann:
                    points_page = ann['points']
                    if len(points_page) >= 2:
                        points_display = self.page_to_display_coords(points_page)
                        stroke_width = max(2, int(3 * self.zoom_level))
                        
                        for i in range(len(points_display) - 1):
                            x1, y1 = points_display[i]
                            x2, y2 = points_display[i + 1]
                            draw.line([x1, y1, x2, y2], fill='red', width=stroke_width)
                
                # -------- TEXT ANNOTATIONS --------
                elif ann_type == 'text' and 'pos_page' in ann:
                    pos_page = ann['pos_page']
                    pos_display = self.page_to_display_coords(pos_page)
                    text = ann.get('text', '')
                    
                    if text:
                        # Draw text background for visibility
                        try:
                            bbox = draw.textbbox(pos_display, text, font=font)
                            padding = 2
                            draw.rectangle(
                                [bbox[0] - padding, bbox[1] - padding,
                                 bbox[2] + padding, bbox[3] + padding],
                                fill=(255, 255, 200, 200)
                            )
                        except:
                            pass
                        
                        draw.text(pos_display, text, fill='red', font=font)
            
            self.photo = ImageTk.PhotoImage(img)
            self.canvas.delete("all")
            self.canvas.create_image(0, 0, anchor=tk.NW, image=self.photo)
            self.canvas.config(scrollregion=self.canvas.bbox(tk.ALL))
            
            self.page_label.config(text=f"Page: {self.current_page + 1}/{len(self.pdf_document)}")
            
            # Restore production visuals if dialog is open
            if self.production_dialog_open and hasattr(self, '_last_highlighted_ann'):
                self.highlight_annotation(self._last_highlighted_ann)
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to display page: {e}")
    
    def zoom_at_point(self, canvas_x, canvas_y, zoom_delta):
        if not self.pdf_document:
            return
        
        old_zoom = self.zoom_level
        new_zoom = max(0.5, min(3.0, old_zoom + zoom_delta))
        
        if new_zoom == old_zoom:
            return
        
        self.zoom_level = new_zoom
        self.display_page()
        
        scale = new_zoom / old_zoom
        bbox = self.canvas.bbox("all")
        if not bbox:
            return
        
        self.canvas.xview_moveto((canvas_x * scale) / max(1, bbox[2]))
        self.canvas.yview_moveto((canvas_y * scale) / max(1, bbox[3]))
    
    def on_double_left_zoom(self, event):
        self.drawing = False
        self.temp_rect_id = None
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        self.zoom_at_point(x, y, +0.25)
    
    def on_double_right_zoom(self, event):
        self.drawing = False
        self.temp_rect_id = None
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        self.zoom_at_point(x, y, -0.25)
    
    def prev_page(self):
        if self.pdf_document and self.current_page > 0:
            self.current_page -= 1
            self.display_page()
    
    def next_page(self):
        if self.pdf_document and self.current_page < len(self.pdf_document) - 1:
            self.current_page += 1
            self.display_page()
    
    def zoom_in(self):
        if self.zoom_level < 3.0:
            self.zoom_level += 0.25
            self.display_page()
    
    def zoom_out(self):
        if self.zoom_level > 0.5:
            self.zoom_level -= 0.25
            self.display_page()
    
    def get_session_path_for_pdf(self):
        """Get session path for current PDF"""
        if not self.current_pdf_path or not self.cabinet_id:
            return None
        
        # Construct expected session path from storage location
        if hasattr(self, 'storage_location') and self.storage_location:
            project_folder = os.path.join(
                self.storage_location,
                self.project_name.replace(' ', '_')
            )
            cabinet_root = os.path.join(
                project_folder,
                self.cabinet_id.replace(' ', '_')
            )
            session_path = os.path.join(
                cabinet_root,
                "Sessions",
                f"{self.cabinet_id}_annotations.json"
            )
            
            return session_path if os.path.exists(session_path) else None
        
        return None
    
    def load_session_from_path(self, path):
        """Load annotation session from JSON file"""
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror("Session Load Error", f"Failed to load session:\n{e}")
            return
        
        self.project_name = data.get('project_name', self.project_name)
        self.sales_order_no = data.get('sales_order_no', self.sales_order_no)
        self.cabinet_id = data.get('cabinet_id', getattr(self, "cabinet_id", ""))
        self.current_page = data.get('current_page', 0)
        self.zoom_level = data.get('zoom_level', 1.0)
        self.current_sr_no = data.get('current_sr_no', self.current_sr_no)
        
        self.annotations = []
        self.session_refs.clear()
        
        for entry in data.get('annotations', []):
            ann = entry.copy()
            
            # Deserialize bbox_page (rectangles) - convert list to tuple
            if 'bbox_page' in ann:
                ann['bbox_page'] = tuple(float(x) for x in ann['bbox_page'])
            
            # Deserialize points (pen strokes) - convert lists to tuples
            if 'points' in ann:
                ann['points'] = [(float(p[0]), float(p[1])) for p in ann['points']]
            
            # Deserialize pos_page (text position) - convert list to tuple
            if 'pos_page' in ann:
                pos = ann['pos_page']
                ann['pos_page'] = (float(pos[0]), float(pos[1]))
            
            self.annotations.append(ann)
            
            if ann.get('ref_no'):
                self.session_refs.add(str(ann['ref_no']).strip())
        
        self.display_page()
        print(f"Session loaded: {len(self.annotations)} annotations")
        
        # Log what types were loaded
        types_loaded = {}
        for ann in self.annotations:
            ann_type = ann.get('type', 'unknown')
            types_loaded[ann_type] = types_loaded.get(ann_type, 0) + 1
        print(f"Annotation types loaded: {types_loaded}")


def main():
    root = tk.Tk()
    app = ProductionTool(root)
    root.mainloop()


if __name__ == "__main__":
    main()

