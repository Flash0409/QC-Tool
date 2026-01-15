import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, Menu
from PIL import Image, ImageTk, ImageDraw, ImageFont,ImageEnhance,ImageFilter
import fitz  # PyMuPDF
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime
import shutil
import tempfile
import re
import os
import json
import numpy as np
import getpass
import sys
import subprocess
import sqlite3
import shlex
from difflib import SequenceMatcher
from handover_database import HandoverDB
from database_manager import DatabaseManager
import pytesseract
import os
import cv2


TESSERACT_PATH = r"C:\Users\E1547548\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"

if os.path.exists(TESSERACT_PATH):
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
    
def get_app_base_dir():
    """Returns the directory where the app is running from."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


    
    


class ManagerDB:
    """Manager database integration with storage_location and excel_path support"""
    def __init__(self, db_path):
        self.db_path = db_path
        self.init_database()
    
    def init_database(self):
        """Initialize tables if they don't exist"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS cabinets (
            cabinet_id TEXT PRIMARY KEY,
            project_name TEXT,
            sales_order_no TEXT,
            total_pages INTEGER DEFAULT 0,
            annotated_pages INTEGER DEFAULT 0,
            total_punches INTEGER DEFAULT 0,
            open_punches INTEGER DEFAULT 0,
            implemented_punches INTEGER DEFAULT 0,
            closed_punches INTEGER DEFAULT 0,
            status TEXT DEFAULT 'quality_inspection',
            created_date TEXT,
            last_updated TEXT,
            storage_location TEXT,
            excel_path TEXT
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS category_occurrences (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cabinet_id TEXT,
            project_name TEXT,
            category TEXT,
            subcategory TEXT,
            occurrence_date TEXT
        )''')
        
        # Add columns if they don't exist
        try:
            cursor.execute('ALTER TABLE cabinets ADD COLUMN storage_location TEXT')
        except sqlite3.OperationalError:
            pass
        
        try:
            cursor.execute('ALTER TABLE cabinets ADD COLUMN excel_path TEXT')
        except sqlite3.OperationalError:
            pass
        
        conn.commit()
        conn.close()

    """
    Add these methods to your ManagerDB class in the quality inspection code
    (likely in highlighter.py or wherever ManagerDB is defined)
    """

    def split_cell(self, cell_ref):
        """Splits 'F6' -> (6, 'F')"""
        import re
        m = re.match(r"([A-Z]+)(\d+)", cell_ref)
        if not m:
            raise ValueError(f"Invalid cell reference: {cell_ref}")
        col, row = m.groups()
        return int(row), col

    def _resolve_merged_target(self, ws, row, col_idx):
        """Handle merged cells"""
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col_idx <= merged.max_col:
                return merged.min_row, merged.min_col
        return row, col_idx

    def read_cell(self, ws, row, col):
        """Read cell value handling merged cells"""
        from openpyxl.utils import column_index_from_string
        
        if isinstance(col, str):
            col_idx = column_index_from_string(col)
        else:
            col_idx = int(col)
        target_row, target_col = self._resolve_merged_target(ws, int(row), col_idx)
        return ws.cell(row=target_row, column=target_col).value

    def get_status_from_interphase(self, excel_path):
        """Read Interphase worksheet and determine status based on reference number
        Returns: status string or None if not determined from Interphase
        """
        if not excel_path or not os.path.exists(excel_path):
            return None
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=True)
            
            # Check if Interphase worksheet exists
            if 'Interphase' not in wb.sheetnames:
                wb.close()
                return None
            
            ws = wb['Interphase']
            
            # Find the lowest filled status cell in column D
            lowest_status_row = None
            lowest_ref_no = None
            
            # Start from row 2 (assuming row 1 is header)
            for row in range(2, ws.max_row + 1):
                status_cell = self.read_cell(ws, row, 'D')
                
                # If status cell has content, check the reference number
                if status_cell:
                    ref_no_cell = self.read_cell(ws, row, 'B')
                    
                    if ref_no_cell:
                        lowest_status_row = row
                        lowest_ref_no = str(ref_no_cell).strip()
            
            wb.close()
            
            # If we found a reference number, determine the status
            if lowest_ref_no:
                try:
                    # Handle range formats like "1-2" or single numbers like "5"
                    if '-' in lowest_ref_no:
                        # Get the first number in the range
                        ref_num = int(lowest_ref_no.split('-')[0])
                    else:
                        ref_num = int(lowest_ref_no)
                    
                    # Determine status based on reference number
                    if 1 <= ref_num <= 2:
                        return 'project_info_sheet'
                    elif 3 <= ref_num <= 9:
                        return 'mechanical_assembly'
                    elif 10 <= ref_num <= 18:
                        return 'component_assembly'
                    elif 19 <= ref_num <= 26:
                        return 'final_assembly'
                    elif 27 <= ref_num <= 31:
                        return 'final_documentation'
                
                except (ValueError, IndexError):
                    # If we can't parse the reference number, return None
                    pass
            
            return None
            
        except Exception as e:
            print(f"Error reading Interphase worksheet: {e}")
            return None
    
    def update_cabinet(self, cabinet_id, project_name, sales_order_no, 
                      total_pages, annotated_pages, total_punches, 
                      open_punches, implemented_punches, closed_punches, status,
                      storage_location=None, excel_path=None):
        """Update cabinet statistics"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT OR REPLACE INTO cabinets 
                (cabinet_id, project_name, sales_order_no, total_pages, annotated_pages,
                 total_punches, open_punches, implemented_punches, closed_punches, status,
                 storage_location, excel_path,
                 created_date, last_updated)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                        COALESCE((SELECT created_date FROM cabinets WHERE cabinet_id = ?), ?),
                        ?)
            ''', (cabinet_id, project_name, sales_order_no, total_pages, annotated_pages,
                  total_punches, open_punches, implemented_punches, closed_punches, status,
                  storage_location, excel_path,
                  cabinet_id, datetime.now().isoformat(), datetime.now().isoformat()))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Manager DB update error: {e}")
            return False
    
    def log_category_occurrence(self, cabinet_id, project_name, category, subcategory):
        """Log a category occurrence"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO category_occurrences 
                (cabinet_id, project_name, category, subcategory, occurrence_date)
                VALUES (?, ?, ?, ?, ?)
            ''', (cabinet_id, project_name, category, subcategory, datetime.now().isoformat()))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Category logging error: {e}")
            return False
    
    def update_status(self, cabinet_id, status):
        """Update cabinet status only"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                UPDATE cabinets 
                SET status = ?, last_updated = ?
                WHERE cabinet_id = ?
            ''', (status, datetime.now().isoformat(), cabinet_id))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Status update error: {e}")
            return False
    
    def get_cabinet(self, cabinet_id):
        """Get cabinet information"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT cabinet_id, project_name, sales_order_no, total_pages, annotated_pages,
                       total_punches, open_punches, implemented_punches, closed_punches, status,
                       storage_location, excel_path, created_date, last_updated
                FROM cabinets 
                WHERE cabinet_id = ?
            ''', (cabinet_id,))
            
            row = cursor.fetchone()
            conn.close()
            
            if row:
                return {
                    'cabinet_id': row[0],
                    'project_name': row[1],
                    'sales_order_no': row[2],
                    'total_pages': row[3],
                    'annotated_pages': row[4],
                    'total_punches': row[5],
                    'open_punches': row[6],
                    'implemented_punches': row[7],
                    'closed_punches': row[8],
                    'status': row[9],
                    'storage_location': row[10],
                    'excel_path': row[11],
                    'created_date': row[12],
                    'last_updated': row[13]
                }
            return None
            
        except Exception as e:
            print(f"Error getting cabinet: {e}")
            return None


class CircuitInspector:
    def __init__(self, root):
        self.root = root
        self.root.title("Quality Inspection Tool - Highlighter Edition")
        self.root.geometry("1400x900")

        # Data / files
        self.pdf_document = None
        self.current_pdf_path = None
        self.current_page = 0
        self.project_name = ""
        self.sales_order_no = ""
        self.cabinet_id = ""
        self.annotations = []
        base = get_app_base_dir()
        self.master_excel_file = os.path.join(base, "Emerson.xlsx")

        self.excel_file = None
        self.working_excel_path = None
        self.checklist_file = self.excel_file
        self.zoom_level = 1.0
        self.current_sr_no = 1
        self.current_page_image = None
        self.tool_mode = None  # None, "pen", or "text"
        self.pen_points = []
        self.session_refs = set()
        self.project_dirs = {}

        # Fixed column mapping
        self.punch_sheet_name = 'Punch Sheet'
        self.punch_cols = {
            'sr_no': 'A',
            'ref_no': 'B',
            'desc': 'C',
            'category': 'D',
            'checked_name': 'E',
            'checked_date': 'F',
            'implemented_name': 'G',
            'implemented_date': 'H',
            'closed_name': 'I',
            'closed_date': 'J'
        }
        
        self.interphase_sheet_name = 'Interphase'
        self.interphase_cols = {
            'ref_no': 'B',
            'description': 'C',
            'status': 'D',
            'name':'E',
            'date':'F',
            'remark':'G'
        }

        self.header_cells = {
            "Interphase": {
                "project_name": "C4",
                "sales_order": "C6",
                "cabinet_id": "F6"
            },
            "Punch Sheet": {
                "project_name": "C2",
                "sales_order": "C4",
                "cabinet_id": "H4"
            }
        }

        self.categories = []
        self.category_file = os.path.join(os.path.dirname(get_app_base_dir()), "assets", "categories.json")
        self.load_categories()

        # HIGHLIGHTER STATE - 3 COLORS
        self.active_highlighter = None
        self.highlighter_colors = {
            'green': {'rgb': (0, 255, 0), 'rgba': (0, 255, 0, 100), 'name': ' OK'},
            'orange': {'rgb': (255, 165, 0), 'rgba': (255, 165, 0, 120), 'name': ' Error'},
            'yellow': {'rgb': (255, 255, 0), 'rgba': (255, 255, 0, 80), 'name': 'Wiring '}
        }
        self.current_color_key = 'yellow'  # Default color
        self.highlight_points = []

        # Drawing / tool state
        self.drawing = False
        self.drawing_type = None  # 'highlight', 'pen', 'text'
        self.temp_line_ids = []  # Store temporary drawing line IDs
        self.selected_annotation = None
        self.undo_stack = []  # Stack for undo operations
        self.max_undo = 50    # Maximum undo history
        self.hover_annotation = None  # For hover preview

        self.setup_ui()
        self.current_sr_no = self.get_next_sr_no()
        
        base = get_app_base_dir()
        db_path = os.path.join(base, "inspection_tool.db")
        self.db = DatabaseManager(db_path)
        manager_db_path = os.path.join(base, "manager.db")
        self.manager_db = ManagerDB(manager_db_path)
        self.handover_db = HandoverDB(os.path.join(base, "handover_db.json"))
        self.load_recent_projects_ui()
        self.root.after(300000, self.auto_save_session)

    # ================================================================
    # COORDINATE CONVERSION HELPERS
    # ================================================================
    
    def page_to_display_scale(self):
        return 2.0 * self.zoom_level

    def display_to_page_coords(self, pts):
        """Convert display-space coordinates to page-space coordinates."""
        scale = self.page_to_display_scale()
        
        # Handle single point tuple
        if isinstance(pts, tuple) and len(pts) == 2:
            if not isinstance(pts[0], (list, tuple)):
                return (pts[0] / scale, pts[1] / scale)
        
        # Handle list of points
        return [(x / scale, y / scale) for x, y in pts]

    def page_to_display_coords(self, pts):
        """Convert page-space coordinates to display-space coordinates."""
        scale = self.page_to_display_scale()
        
        # Handle single point tuple
        if isinstance(pts, tuple) and len(pts) == 2:
            if not isinstance(pts[0], (list, tuple)):
                return (pts[0] * scale, pts[1] * scale)
        
        # Handle list of points
        return [(x * scale, y * scale) for x, y in pts]

    def bbox_page_to_display(self, bbox_page):
        scale = self.page_to_display_scale()
        x1, y1, x2, y2 = bbox_page
        return (x1 * scale, y1 * scale, x2 * scale, y2 * scale)

    def bbox_display_to_page(self, bbox_display):
        scale = self.page_to_display_scale()
        x1, y1, x2, y2 = bbox_display
        return (x1 / scale, y1 / scale, x2 / scale, y2 / scale)

    # ================================================================
    # HIGHLIGHTER HELPER - AUTO-STRAIGHTEN
    # ================================================================
    
    def straighten_path(self, points):
        """Convert a freehand path into a straight line from start to end."""
        if len(points) < 2:
            return points
        # Simply return start and end points for a perfectly straight line
        return [points[0], points[-1]]

    # ================================================================
    # MOUSE EVENT HANDLERS - HIGHLIGHTER INTEGRATED
    # ================================================================

    def on_left_press(self, event):
        """Handle left mouse button press"""
        if not self.pdf_document:
            messagebox.showwarning("Warning", "Please load a PDF first")
            return

        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)

        # -------- HIGHLIGHTER MODE --------
        if self.active_highlighter:
            self.drawing = True
            self.drawing_type = "highlight"
            self.highlight_points = [(x, y)]
            self.clear_temp_drawings()
            return

        # -------- PEN TOOL --------
        if self.tool_mode == "pen":
            self.drawing = True
            self.drawing_type = "pen"
            self.pen_points = [(x, y)]
            self.clear_temp_drawings()
            return

        # -------- TEXT TOOL --------
        if self.tool_mode == "text":
            self.drawing = True
            self.drawing_type = "text"
            self.text_pos_x = x
            self.text_pos_y = y
            return

    def on_left_drag(self, event):
        """Handle left mouse button drag"""
        if not self.drawing:
            return

        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)

        # -------- HIGHLIGHTER DRAWING --------
        if self.drawing_type == "highlight":
            if len(self.highlight_points) > 0:
                last_x, last_y = self.highlight_points[-1]
                
                # Get highlighter color
                rgba = self.highlighter_colors[self.active_highlighter]['rgba']
                rgb = self.highlighter_colors[self.active_highlighter]['rgb']
                hex_color = f'#{rgb[0]:02x}{rgb[1]:02x}{rgb[2]:02x}'
                
                # Draw thick line segment
                line_id = self.canvas.create_line(
                    last_x, last_y, x, y,
                    fill=hex_color,
                    width=max(15, int(15 * self.zoom_level)),
                    capstyle=tk.ROUND,
                    smooth=True
                )
                self.temp_line_ids.append(line_id)
            
            self.highlight_points.append((x, y))
            return

        # -------- PEN TOOL DRAWING --------
        if self.drawing_type == "pen":
            if len(self.pen_points) > 0:
                last_x, last_y = self.pen_points[-1]
                line_id = self.canvas.create_line(
                    last_x, last_y, x, y,
                    fill="red", width=3,
                    capstyle=tk.ROUND, smooth=True
                )
                self.temp_line_ids.append(line_id)
            self.pen_points.append((x, y))
            return

    def on_left_release_with_ocr(self, event):
        """Handle left mouse button release with OCR text extraction"""
        if not self.pdf_document or not self.drawing:
            return

        # -------- HIGHLIGHTER FINISH WITH OCR --------
        if self.drawing_type == "highlight":
            if len(self.highlight_points) >= 2:
                # ALWAYS apply straightening for highlighter
                processed_points = self.straighten_path(self.highlight_points)
                
                # Convert to page coordinates
                points_page = self.display_to_page_coords(processed_points)
                
                # Calculate bounding box
                xs = [p[0] for p in points_page]
                ys = [p[1] for p in points_page]
                bbox_page = (min(xs), min(ys), max(xs), max(ys))
                
                # Create annotation
                annotation = {
                    'type': 'highlight',
                    'color': self.active_highlighter,
                    'page': self.current_page,
                    'bbox_page': bbox_page,
                    'points_page': points_page,
                    'timestamp': datetime.now().isoformat()
                }
                
                # ✨ NEW: Extract text from highlighted area if orange highlighter
                if self.active_highlighter == 'orange':
                    print("\n🔍 Extracting text from highlighted area...")
                    extracted_text = self.extract_text_from_highlight_area(annotation)
                    
                    if extracted_text:
                        annotation['extracted_text'] = extracted_text
                        print(f"✓ Text extracted: '{extracted_text[:100]}...'")
                    else:
                        annotation['extracted_text'] = None
                        print("⚠️ No text extracted")
                    
                    # Show action menu with extracted text
                    self.handle_error_highlight_with_ocr(annotation)
                else:
                    # Green/Yellow highlighters - no OCR, just add annotation
                    self.annotations.append(annotation)
                    self.add_to_undo_stack('add_annotation', annotation)
                    self.display_page()
            
            self.highlight_points = []
            self.clear_temp_drawings()
            self.drawing = False
            self.drawing_type = None
            self.update_tool_pane()
            return

        # -------- PEN TOOL FINISH - NO CHANGES --------
        if self.drawing_type == "pen":
            if len(self.pen_points) >= 2:
                points_page = self.display_to_page_coords(self.pen_points)
                annotation = {
                    'type': 'pen',
                    'page': self.current_page,
                    'points': points_page,
                    'timestamp': datetime.now().isoformat()
                }
                self.annotations.append(annotation)
                self.add_to_undo_stack('add_annotation', annotation)
            self.pen_points = []
            self.clear_temp_drawings()
            self.drawing = False
            self.drawing_type = None
            self.display_page()
            self.update_tool_pane()
            return

        # -------- TEXT TOOL FINISH - NO CHANGES --------
        if self.drawing_type == "text":
            txt = simpledialog.askstring("Text", "Enter text:", parent=self.root)
            if txt and txt.strip():
                pos_page = self.display_to_page_coords((self.text_pos_x, self.text_pos_y))
                annotation = {
                    'type': 'text',
                    'page': self.current_page,
                    'pos_page': pos_page,
                    'text': txt.strip(),
                    'timestamp': datetime.now().isoformat()
                }
                self.annotations.append(annotation)
                self.add_to_undo_stack('add_annotation', annotation)
                self.display_page()
            self.drawing = False
            self.drawing_type = None
            self.update_tool_pane()
            return


    # ============================================================================
    # MODIFIED: handle_error_highlight with OCR pre-filled text
    # Replace your existing handle_error_highlight method
    # ============================================================================
    def load_categories(self):
        """Load categories from JSON"""
        try:
            if os.path.exists(self.category_file):
                with open(self.category_file, "r", encoding='utf-8') as f:
                    data = json.load(f)
                    self.categories = data if isinstance(data, list) else []
                    print(f"✓ Loaded categories with automatic reference numbers")
            else:
                print("⚠️ Categories file not found")
                self.categories = []
                
        except Exception as e:
            print(f"❌ Error loading categories: {e}")
            self.categories = []


    """
    Enhanced High-Resolution OCR Extraction
    Captures ANY size highlight, intelligently expands it, sharpens, and extracts text
    """


    def extract_text_from_highlight_area(self, annotation):
        """Extract text from highlighted area with automatic padding and rotation support - OPTIMIZED"""
        if self.current_page_image is None:
            return None
        
        try:
            bbox_page = annotation.get('bbox_page')
            if not bbox_page:
                return None
            
            bbox_display = self.bbox_page_to_display(bbox_page)
            x1, y1, x2, y2 = bbox_display
            
            # ✨ EXPAND BBOX - Add generous padding for OCR
            PADDING_X = 10  # Horizontal padding
            PADDING_Y = 20  # Vertical padding (more because text height matters)
            
            height, width = self.current_page_image.shape[:2]
            
            x1 = max(0, int(x1) - PADDING_X)
            y1 = max(0, int(y1) - PADDING_Y)
            x2 = min(width, int(x2) + PADDING_X)
            y2 = min(height, int(y2) + PADDING_Y)
            
            crop_width = x2 - x1
            crop_height = y2 - y1
            
            if crop_width < 20 or crop_height < 15:
                print("⚠️ WARNING: Highlighted area too small")
                return None
            
            cropped = self.current_page_image[y1:y2, x1:x2]
            
            if cropped.size == 0:
                return None
            
            # Upscale for better OCR (smaller scale for faster processing)
            h, w = cropped.shape[:2]
            upscaled = cv2.resize(cropped, (w*2, h*2), interpolation=cv2.INTER_CUBIC)
            
            # Convert to grayscale and threshold in one go
            gray = cv2.cvtColor(upscaled, cv2.COLOR_RGB2GRAY)
            _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
            # Convert to PIL once (reuse for all rotations)
            pil_img = Image.fromarray(binary)
            
            # Smart rotation strategy: Try 0° first (most common), then others only if needed
            # First try: Normal orientation (0°)
            text, confidence = self._ocr_with_confidence(pil_img)
            
            # If confidence is high enough (>60%), accept immediately
            if confidence > 60 and text:
                cleaned_text = self.clean_ocr_text(text)
                if cleaned_text and len(cleaned_text) > 1:
                    print(f"✅ Extracted (0°, {confidence:.1f}%): '{cleaned_text}'")
                    return cleaned_text
            
            # Otherwise, try other rotations
            best_text = text
            best_confidence = confidence
            best_rotation = 0
            
            rotations = [
                (90, pil_img.rotate(270, expand=True), "90°"),   # PIL rotate is counter-clockwise
                (270, pil_img.rotate(90, expand=True), "270°"),
                (180, pil_img.rotate(180, expand=True), "180°")
            ]
            
            for angle, rotated_img, label in rotations:
                text, conf = self._ocr_with_confidence(rotated_img)
                
                if conf > best_confidence:
                    best_confidence = conf
                    best_text = text
                    best_rotation = angle
                    
                    # Early exit if we find high confidence result
                    if conf > 70:
                        break
            
            if best_text:
                if best_rotation != 0:
                    print(f"✨ Best: {best_rotation}° ({best_confidence:.1f}%)")
                
                cleaned_text = self.clean_ocr_text(best_text)
                
                if cleaned_text and len(cleaned_text) > 1:
                    print(f"✅ Extracted: '{cleaned_text}'")
                    return cleaned_text
            
            print("⚠️ No text found")
            return None
                    
        except Exception as e:
            print(f"❌ OCR Error: {e}")
            return None


    def _ocr_with_confidence(self, pil_image):
        """
        Helper method to run OCR and calculate confidence
        
        Args:
            pil_image: PIL Image object
            
        Returns:
            tuple: (text, average_confidence)
        """
        try:
            # Get OCR data with confidence scores
            ocr_data = pytesseract.image_to_data(
                pil_image, 
                lang='eng', 
                config='--psm 6',  # Assume uniform block of text
                output_type=pytesseract.Output.DICT
            )
            
            text_parts = []
            confidences = []
            
            # Extract words with valid confidence
            for i, conf in enumerate(ocr_data['conf']):
                if conf > 0:  # Valid confidence
                    word = ocr_data['text'][i].strip()
                    if word:
                        text_parts.append(word)
                        confidences.append(conf)
            
            if text_parts:
                text = ' '.join(text_parts)
                avg_confidence = sum(confidences) / len(confidences)
                return text, avg_confidence
            
            return None, 0
            
        except Exception as e:
            print(f"⚠️ OCR processing error: {e}")
            return None, 0


    def clean_ocr_text(self, text):
        """
        Clean OCR output text - optimized version
        
        Args:
            text: Raw OCR text
            
        Returns:
            str: Cleaned text or None
        """
        if not text:
            return None
        
        # Single pass cleaning with string operations
        text = ' '.join(text.split())  # Remove extra whitespace
        
        # Character replacements in one go using translate
        replacements = str.maketrans({
            '|': 'I',
            '`': "'",
            '~': '-'
        })
        text = text.translate(replacements)
        
        # Keep only printable characters
        text = ''.join(char for char in text if char.isprintable())
        
        # Strip and return
        text = text.strip()
        
        return text if len(text) >= 2 else None

        # ============================================================================
        # SIMPLIFIED VERSION - If the above is too complex
        # ============================================================================

    def extract_text_simple(self, annotation):
        """
        Simplified OCR extraction - Just upscale and try
        """
        if self.current_page_image is None:
            print("❌ No image loaded")
            return None
        
        try:
            bbox_page = annotation.get('bbox_page')
            if not bbox_page:
                return None
            
            bbox_display = self.bbox_page_to_display(bbox_page)
            x1, y1, x2, y2 = map(int, bbox_display)
            
            # Get image dimensions
            height, width = self.current_page_image.shape[:2]
            
            # Clip to image bounds
            x1 = max(0, min(x1, width))
            y1 = max(0, min(y1, height))
            x2 = max(0, min(x2, width))
            y2 = max(0, min(y2, height))
            
            # Crop
            cropped = self.current_page_image[y1:y2, x1:x2]
            
            print(f"Cropped area: {cropped.shape}")
            
            if cropped.size == 0:
                print("❌ Empty crop")
                return None
            
            # Upscale 3x for better OCR
            h, w = cropped.shape[:2]
            upscaled = cv2.resize(cropped, (w*3, h*3), interpolation=cv2.INTER_CUBIC)
            
            # Convert to grayscale
            gray = cv2.cvtColor(upscaled, cv2.COLOR_RGB2GRAY)
            
            # Threshold
            _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
            # Save for debugging
            try:
                debug_dir = "ocr_debug"
                os.makedirs(debug_dir, exist_ok=True)
                cv2.imwrite(os.path.join(debug_dir, "crop.png"), binary)
                print(f"💾 Saved to ocr_debug/crop.png")
            except:
                pass
            
            # OCR
            pil_img = Image.fromarray(binary)
            text = pytesseract.image_to_string(pil_img, lang='eng', config='--psm 6')
            
            # Clean
            text = ' '.join(text.split()).strip()
            
            if text and len(text) > 1:
                print(f"✅ Extracted: '{text}'")
                return text
            else:
                print("⚠️ No text found")
                return None
                
        except Exception as e:
            print(f"❌ OCR Error: {e}")
            import traceback
            traceback.print_exc()
            return None
        
    
    def preprocess_for_ocr(self, pil_image):
        """
        Preprocess image for better OCR accuracy
        
        Args:
            pil_image: PIL Image object
            
        Returns:
            PIL Image: Preprocessed image
        """
        # Convert to numpy array
        img_array = np.array(pil_image)
        
        # Convert to grayscale
        if len(img_array.shape) == 3:
            gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
        else:
            gray = img_array
        
        # Increase contrast
        # Apply adaptive thresholding for better text detection
        thresh = cv2.adaptiveThreshold(
            gray, 255, 
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
            cv2.THRESH_BINARY, 
            11, 2
        )
        
        # Denoise
        denoised = cv2.fastNlMeansDenoising(thresh)
        
        # Convert back to PIL
        return Image.fromarray(denoised)
    
    
    def clean_ocr_text(self, text):
        """
        Clean OCR output text
        
        Args:
            text: Raw OCR text
            
        Returns:
            str: Cleaned text
        """
        if not text:
            return None
        
        # Remove extra whitespace
        text = ' '.join(text.split())
        
        # Remove common OCR artifacts
        text = text.replace('|', 'I')
        text = text.replace('`', "'")
        
        # Remove non-printable characters
        text = ''.join(char for char in text if char.isprintable())
        
        # Strip leading/trailing whitespace
        text = text.strip()
        
        return text if text else None

    def handle_error_highlight_with_ocr(self, annotation):
        """Handle orange error highlight with OCR-extracted text pre-filled"""
        
        extracted_text = annotation.get('extracted_text', None)
        
        menu = Menu(self.root, tearoff=0)

        for cat in self.categories:
            mode = cat.get("mode", "parent")
            
            # ========== WIRING SELECTOR MODE ==========
            if mode == "wiring_selector":
                menu.add_command(
                    label=f"🔧 {cat['name']}",
                    command=lambda c=cat, ann=annotation, txt=extracted_text: 
                        self.handle_wiring_selector_with_ocr(c, ann, txt)
                )
            
            # ========== TEMPLATE MODE ==========
            elif mode == "template":
                ref_num = cat.get("ref_number", "")
                if ref_num:
                    label = f"🔧 [{ref_num}] {cat['name']}"
                else:
                    label = f"🔧 {cat['name']}"
                
                menu.add_command(
                    label=label,
                    command=lambda c=cat, ann=annotation, txt=extracted_text: 
                        self.handle_template_category_highlight_with_ocr(c, ann, txt)
                )
            
            # ========== PARENT MODE ==========
            elif mode == "parent":
                cat_menu = Menu(menu, tearoff=0)
                for sub in cat.get("subcategories", []):
                    ref_num = sub.get("ref_number", "??")
                    if ref_num:
                        label = f"[{ref_num}] {sub['name']}"
                    else:
                        label = sub['name']
                    cat_menu.add_command(
                        label=label,
                        command=lambda c=cat, s=sub, ann=annotation, txt=extracted_text: 
                            self.handle_subcategory_highlight_with_ocr(c, s, ann, txt)
                    )
                
                menu.add_cascade(label=f"🔧 {cat['name']}", menu=cat_menu)

        menu.add_separator()
        menu.add_command(
            label="📝 Custom Action Point",
            command=lambda ann=annotation, txt=extracted_text: 
                self.log_custom_error_highlight_with_ocr(ann, txt)
        )

        x = self.root.winfo_pointerx()
        y = self.root.winfo_pointery()
        menu.tk_popup(x, y)


    # ============================================================================
    # MODIFIED: Template running with OCR pre-fill
    # Add this helper method to your class
    # ============================================================================

    def run_template_with_ocr(self, template_def, tag_name=None, prefill_text=None):
        """
        Execute a template definition with OCR text pre-filled
        
        Args:
            template_def: Template configuration
            tag_name: Tag name (optional)
            prefill_text: Text extracted from OCR to pre-fill first input
        
        Returns:
            str: Formatted template text
        """
        values = {}
        if tag_name:
            values["tag"] = tag_name

        inputs = template_def.get("inputs", [])
        
        for i, inp in enumerate(inputs):
            # ✨ Pre-fill first input with OCR text if available
            initial_value = ""
            if i == 0 and prefill_text:
                initial_value = prefill_text
                print(f"✓ Pre-filling with OCR text: '{prefill_text[:50]}...'")
            
            val = simpledialog.askstring(
                "Input Required", 
                inp["label"], 
                parent=self.root,
                initialvalue=initial_value  # ✨ Pre-filled!
            )
            
            if not val:
                return None
            values[inp["name"]] = val.strip()

        try:
            return template_def["template"].format(**values)
        except KeyError as e:
            messagebox.showerror("Template Error", f"Missing placeholder: {e}")
            return None


    # ============================================================================
    # MODIFIED: Handler methods with OCR support
    # Add these to your CircuitInspector class
    # ============================================================================

    def handle_wiring_selector_with_ocr(self, category, annotation, extracted_text):
        """Handle wiring type selection with OCR text"""
        wiring_menu = Menu(self.root, tearoff=0, bg='#1e293b', fg='white',
                          activebackground='#3b82f6', font=('Segoe UI', 10))
        
        wiring_types = category.get("wiring_types", [])
        
        for wiring in wiring_types:
            wiring_type = wiring.get("type", "Unknown")
            ref_num = wiring.get("ref_number", "??")
            
            wiring_menu.add_command(
                label=f"[{ref_num}] {wiring_type}",
                command=lambda c=category, w=wiring, ann=annotation, txt=extracted_text: 
                    self.show_wiring_subcategories_with_ocr(c, w, ann, txt)
            )
        
        special_subs = category.get("special_subcategories", [])
        if special_subs:
            wiring_menu.add_separator()
            for special in special_subs:
                ref_num = special.get("ref_number", "??")
                wiring_menu.add_command(
                    label=f"[{ref_num}] {special['name']} (All types)",
                    command=lambda c=category, s=special, ann=annotation, txt=extracted_text:
                        self.handle_special_subcategory_with_ocr(c, s, ann, txt)
                )
        
        x = self.root.winfo_pointerx()
        y = self.root.winfo_pointery()
        wiring_menu.tk_popup(x, y)


    def show_wiring_subcategories_with_ocr(self, category, wiring_data, annotation, extracted_text):
        """Show sub-subcategories with OCR text"""
        subcategories = wiring_data.get("subcategories", [])
        
        if not subcategories:
            self.handle_wiring_type_selected_with_ocr(category, wiring_data, annotation, extracted_text)
            return
        
        sub_menu = Menu(self.root, tearoff=0, bg='#1e293b', fg='white',
                       activebackground='#3b82f6', font=('Segoe UI', 10))
        
        wiring_type = wiring_data.get("type", "Unknown")
        ref_num = wiring_data.get("ref_number", "??")
        
        for sub in subcategories:
            sub_name = sub.get("name", "Unknown")
            sub_menu.add_command(
                label=f"[{ref_num}] {sub_name}",
                command=lambda c=category, w=wiring_data, s=sub, ann=annotation, txt=extracted_text:
                    self.handle_wiring_subcategory_selected_with_ocr(c, w, s, ann, txt)
            )
        
        x = self.root.winfo_pointerx()
        y = self.root.winfo_pointery()
        sub_menu.tk_popup(x, y)


    def handle_wiring_subcategory_selected_with_ocr(self, category, wiring_data, subcategory, annotation, extracted_text):
        """Handle wiring subcategory with OCR pre-fill"""
        
        punch_text = self.run_template_with_ocr(subcategory, tag_name=None, prefill_text=extracted_text)
        if not punch_text:
            return
        
        ref_number = wiring_data.get("ref_number", "??")
        wiring_type = wiring_data.get("type", "Unknown")
        sub_name = subcategory.get("name", "Unknown")
        
        self.log_error_direct(
            component_type=category["name"],
            error_name=f"{wiring_type} - {sub_name}",
            error_template=punch_text,
            annotation=annotation,
            ref_number=ref_number
        )


    def handle_template_category_highlight_with_ocr(self, category, annotation, extracted_text):
        """Handle template category with OCR pre-fill"""
        punch_text = self.run_template_with_ocr(category, tag_name=None, prefill_text=extracted_text)
        if not punch_text:
            return
        
        ref_number = category.get("ref_number", "")
        
        if ref_number:
            self.log_error_direct(
                component_type=category["name"],
                error_name=None,
                error_template=punch_text,
                annotation=annotation,
                ref_number=ref_number
            )
        else:
            self.log_error_with_popup(
                component_type=category["name"],
                error_name=None,
                error_template=punch_text,
                annotation=annotation
            )


    def handle_subcategory_highlight_with_ocr(self, category, subcategory, annotation, extracted_text):
        """Handle subcategory with OCR pre-fill"""
        punch_text = self.run_template_with_ocr(subcategory, tag_name=None, prefill_text=extracted_text)
        if not punch_text:
            return
        
        ref_number = subcategory.get("ref_number", "")
        
        if ref_number:
            self.log_error_direct(
                component_type=category["name"],
                error_name=subcategory["name"],
                error_template=punch_text,
                annotation=annotation,
                ref_number=ref_number
            )
        else:
            self.log_error_with_popup(
                component_type=category["name"],
                error_name=subcategory["name"],
                error_template=punch_text,
                annotation=annotation
            )


    def log_error_direct(self, component_type, error_name, error_template, annotation, ref_number):
        """Log error DIRECTLY without asking for reference number"""
        punch_text = error_template

        if not punch_text:
            messagebox.showerror("Error", "Punch description is empty.")
            return

        # Use ref_number directly - NO POPUP
        ref_no = str(ref_number).strip()
        self.session_refs.add(ref_no)

        try:
            wb = load_workbook(self.excel_file)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active

            row_num = 8
            while True:
                val = self.read_cell(ws, row_num, self.punch_cols['sr_no'])
                if val is None:
                    break
                row_num += 1

            prev_sr = None
            if row_num > 8:
                prev_sr = self.read_cell(ws, row_num - 1, self.punch_cols['sr_no'])

            try:
                sr_no_assigned = int(prev_sr) + 1 if prev_sr is not None else 1
            except:
                sr_no_assigned = 1

            self.write_cell(ws, row_num, self.punch_cols['sr_no'], sr_no_assigned)
            self.write_cell(ws, row_num, self.punch_cols['ref_no'], ref_no)
            self.write_cell(ws, row_num, self.punch_cols['desc'], punch_text)
            self.write_cell(ws, row_num, self.punch_cols['category'], component_type)

            try:
                uname = os.getlogin()
            except:
                uname = getpass.getuser()

            self.write_cell(ws, row_num, self.punch_cols['checked_name'], uname)
            self.write_cell(ws, row_num, self.punch_cols['checked_date'], datetime.now().strftime("%Y-%m-%d"))

            wb.save(self.excel_file)
            wb.close()

            updated = self.update_interphase_status_for_ref(ref_no, status='NOK')
            if updated:
                print(f"✓ Interphase: marked ref {ref_no} as NOK")

            # Update annotation with all the data
            annotation['component'] = component_type
            annotation['subcategory'] = error_name
            annotation['punch_text'] = punch_text
            annotation['ref_no'] = ref_no
            annotation['excel_row'] = row_num
            annotation['sr_no'] = sr_no_assigned
            annotation['implemented'] = False
            annotation['implemented_name'] = None
            annotation['implemented_date'] = None
            annotation['implementation_remark'] = None

            # Add to annotations list
            self.annotations.append(annotation)
            self.current_sr_no = self.get_next_sr_no()
            
            # Redraw to show the color change from orange to red
            self.display_page()

            print(f"✓ Logged: Ref {ref_no}, SR {sr_no_assigned}")
            self._flash_status(f"✓ Logged Ref {ref_no}", bg='#10b981')
            
            try:
                self.manager_db.log_category_occurrence(
                    self.cabinet_id,
                    self.project_name,
                    component_type,
                    error_name
                )
                self.sync_manager_stats_only()
            except Exception as e:
                print(f"Manager category logging failed: {e}")

        except PermissionError:
            messagebox.showerror("Error", "Close the Excel file before writing to it.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to log punch:\n{e}")
            import traceback
            traceback.print_exc()


    def log_error_with_popup(self, component_type, error_name, error_template, annotation):
        """Log error WITH popup (for Design Error only)"""
        punch_text = error_template

        if not punch_text:
            messagebox.showerror("Error", "Punch description is empty.")
            return

        # ASK FOR REFERENCE NUMBER (Design Error)
        ref_no = simpledialog.askstring(
            "Reference Number", 
            "Enter the reference number:",
            parent=self.root
        )
        
        if not ref_no:
            return

        ref_no = str(ref_no).strip()
        self.session_refs.add(ref_no)

        try:
            wb = load_workbook(self.excel_file)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active

            row_num = 8
            while True:
                val = self.read_cell(ws, row_num, self.punch_cols['sr_no'])
                if val is None:
                    break
                row_num += 1

            prev_sr = None
            if row_num > 8:
                prev_sr = self.read_cell(ws, row_num - 1, self.punch_cols['sr_no'])

            try:
                sr_no_assigned = int(prev_sr) + 1 if prev_sr is not None else 1
            except:
                sr_no_assigned = 1

            self.write_cell(ws, row_num, self.punch_cols['sr_no'], sr_no_assigned)
            self.write_cell(ws, row_num, self.punch_cols['ref_no'], ref_no)
            self.write_cell(ws, row_num, self.punch_cols['desc'], punch_text)
            self.write_cell(ws, row_num, self.punch_cols['category'], component_type)

            try:
                uname = os.getlogin()
            except:
                uname = getpass.getuser()

            self.write_cell(ws, row_num, self.punch_cols['checked_name'], uname)
            self.write_cell(ws, row_num, self.punch_cols['checked_date'], datetime.now().strftime("%Y-%m-%d"))

            wb.save(self.excel_file)
            wb.close()

            updated = self.update_interphase_status_for_ref(ref_no, status='NOK')
            if updated:
                print(f"✓ Interphase: marked ref {ref_no} as NOK")

            annotation['component'] = component_type
            annotation['subcategory'] = error_name
            annotation['punch_text'] = punch_text
            annotation['ref_no'] = ref_no
            annotation['excel_row'] = row_num
            annotation['sr_no'] = sr_no_assigned
            annotation['implemented'] = False
            annotation['implemented_name'] = None
            annotation['implemented_date'] = None
            annotation['implementation_remark'] = None

            self.annotations.append(annotation)
            self.current_sr_no = self.get_next_sr_no()
            self.display_page()

            print(f"✓ Logged: Ref {ref_no}, SR {sr_no_assigned}")
            self._flash_status(f"✓ Logged Ref {ref_no}", bg='#10b981')
            
            try:
                self.manager_db.log_category_occurrence(
                    self.cabinet_id,
                    self.project_name,
                    component_type,
                    error_name
                )
                self.sync_manager_stats_only()
            except Exception as e:
                print(f"Manager category logging failed: {e}")

        except PermissionError:
            messagebox.showerror("Error", "Close the Excel file before writing to it.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to log punch:\n{e}")
            import traceback
            traceback.print_exc()


    def log_custom_error_highlight_with_ocr(self, annotation, extracted_text):
        """Log custom error with OCR pre-fill"""
        try:
            # Pre-fill with OCR text
            custom_action = simpledialog.askstring(
                "Custom Action Point",
                "Enter the action point / punch description:",
                parent=self.root,
                initialvalue=extracted_text if extracted_text else ""
            )
            
            if not custom_action:
                return

            custom_category = simpledialog.askstring(
                "Custom Category",
                "Enter the category:",
                parent=self.root
            )
            if not custom_category:
                return

            ref_no = simpledialog.askstring(
                "Reference Number", 
                "Enter the reference number:",
                parent=self.root
            )
            
            if not ref_no:
                messagebox.showwarning("Reference Required", "Reference No is required.")
                return

            ref_no = str(ref_no).strip()
            self.session_refs.add(ref_no)

            wb = load_workbook(self.excel_file)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active

            row_num = 8
            while True:
                val = self.read_cell(ws, row_num, self.punch_cols['sr_no'])
                if val is None:
                    break
                row_num += 1

            prev_sr = None
            if row_num > 8:
                prev_sr = self.read_cell(ws, row_num - 1, self.punch_cols['sr_no'])

            try:
                sr_no_assigned = int(prev_sr) + 1 if prev_sr is not None else 1
            except:
                sr_no_assigned = 1

            self.write_cell(ws, row_num, self.punch_cols['sr_no'], sr_no_assigned)
            self.write_cell(ws, row_num, self.punch_cols['ref_no'], ref_no)
            self.write_cell(ws, row_num, self.punch_cols['desc'], custom_action)
            self.write_cell(ws, row_num, self.punch_cols['category'], custom_category)

            try:
                uname = os.getlogin()
            except:
                uname = getpass.getuser()

            self.write_cell(ws, row_num, self.punch_cols['checked_name'], uname)
            self.write_cell(ws, row_num, self.punch_cols['checked_date'], datetime.now().strftime("%Y-%m-%d"))

            wb.save(self.excel_file)
            wb.close()

            annotation['component'] = custom_category
            annotation['error'] = 'Custom'
            annotation['punch_text'] = custom_action
            annotation['ref_no'] = ref_no
            annotation['excel_row'] = row_num
            annotation['sr_no'] = sr_no_assigned
            annotation['timestamp'] = datetime.now().isoformat()

            self.annotations.append(annotation)
            self.current_sr_no = self.get_next_sr_no()
            self.display_page()

            print(f"✓ Logged custom: Ref {ref_no}, SR {sr_no_assigned}")
            self._flash_status(f"✓ Custom punch Ref {ref_no}", bg='#8b5cf6')

            try:
                self.manager_db.log_category_occurrence(
                    self.cabinet_id,
                    self.project_name,
                    custom_category,
                    None
                )
                self.sync_manager_stats_only()
            except Exception as e:
                print(f"Manager category logging failed: {e}")

            updated = self.update_interphase_status_for_ref(ref_no, status='NOK')
            if updated:
                print(f"✓ Interphase: marked ref {ref_no} as NOK")

        except PermissionError:
            messagebox.showerror("Error", "Close the Excel file before writing to it.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to log custom error:\n{e}")
            import traceback
            traceback.print_exc()


    # ============================================================================
    # Additional helper methods for OCR
    # ============================================================================

    def handle_wiring_type_selected_with_ocr(self, category, wiring_data, annotation, extracted_text):
        """Handle direct wiring type selection with OCR"""
        punch_text = self.run_template_with_ocr(wiring_data, tag_name=None, prefill_text=extracted_text)
        if not punch_text:
            return
        
        ref_number = wiring_data.get("ref_number", "??")
        wiring_type = wiring_data.get("type", "Unknown")
        
        self.log_error_direct(
            component_type=category["name"],
            error_name=wiring_type,
            error_template=punch_text,
            annotation=annotation,
            ref_number=ref_number
        )


    def handle_special_subcategory_with_ocr(self, category, special_sub, annotation, extracted_text):
        """Handle special subcategories with OCR"""
        punch_text = self.run_template_with_ocr(special_sub, tag_name=None, prefill_text=extracted_text)
        if not punch_text:
            return
        
        ref_number = special_sub.get("ref_number", "??")
        
        self.log_error_direct(
            component_type=category["name"],
            error_name=special_sub["name"],
            error_template=punch_text,
            annotation=annotation,
            ref_number=ref_number
        )


    def clear_temp_drawings(self):
        """Clear temporary drawing elements from canvas"""
        for line_id in self.temp_line_ids:
            try:
                self.canvas.delete(line_id)
            except:
                pass
        self.temp_line_ids.clear()

    # ================================================================
    # DISPLAY PAGE - WITH HIGHLIGHTER, PEN AND TEXT RENDERING
    # ================================================================

    def display_page(self):
        """Render the current PDF page with all annotations"""
        if not self.pdf_document:
            self.canvas.delete("all")
            self.page_label.config(text="Page: 0/0")
            return

        try:
            page = self.pdf_document[self.current_page]
            mat = fitz.Matrix(self.page_to_display_scale(), self.page_to_display_scale())
            pix = page.get_pixmap(matrix=mat)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            self.current_page_image = np.array(img)
            draw = ImageDraw.Draw(img, 'RGBA')

            # Try to load a font for text
            try:
                font_size = max(12, int(14 * self.zoom_level))
                font = ImageFont.truetype("arial.ttf", font_size)
            except:
                font = ImageFont.load_default()

            for ann in self.annotations:

                if ann.get('page') != self.current_page:
                    continue

                ann_type = ann.get('type')

                # -------- HIGHLIGHTER STROKES --------
                if ann_type == 'highlight' and 'points_page' in ann:
                    points_page = ann['points_page']
                    if len(points_page) >= 2:
                        points_display = self.page_to_display_coords(points_page)
                        color_key = ann.get('color', 'yellow')
                        rgba = self.highlighter_colors[color_key]['rgba']
                        
                        # Draw thick semi-transparent strokes
                        stroke_width = max(15, int(15 * self.zoom_level))
                        for i in range(len(points_display) - 1):
                            x1, y1 = points_display[i]
                            x2, y2 = points_display[i + 1]
                            draw.line([x1, y1, x2, y2], fill=rgba, width=stroke_width)
                        
                        # Add closed indicator if applicable
                        if ann.get('closed_by'):
                            bbox_display = self.bbox_page_to_display(ann['bbox_page'])
                            cx = bbox_display[0] + 8
                            cy = bbox_display[1] + 8
                            draw.ellipse([cx - 6, cy - 6, cx + 6, cy + 6], fill=(0, 128, 0, 200))

                # -------- PEN STROKES --------
                elif ann_type == 'pen' and 'points' in ann:
                    points_page = ann['points']
                    if len(points_page) >= 2:
                        points_display = self.page_to_display_coords(points_page)
                        stroke_width = max(2, int(3 * self.zoom_level))
                        for i in range(len(points_display) - 1):
                            x1, y1 = points_display[i]
                            x2, y2 = points_display[i + 1]
                            draw.line([x1, y1, x2, y2], fill='red', width=stroke_width)

                # -------- TEXT ANNOTATIONS --------
                elif ann_type == 'text' and 'pos_page' in ann:
                    pos_page = ann['pos_page']
                    pos_display = self.page_to_display_coords(pos_page)
                    text = ann.get('text', '')
                    if text:
                        # Draw text background for visibility
                        try:
                            bbox = draw.textbbox(pos_display, text, font=font)
                            padding = 2
                            draw.rectangle(
                                [bbox[0] - padding, bbox[1] - padding,
                                 bbox[2] + padding, bbox[3] + padding],
                                fill=(255, 255, 200, 200)
                            )
                        except:
                            pass
                        draw.text(pos_display, text, fill='red', font=font)

            self.photo = ImageTk.PhotoImage(img)
            self.canvas.delete("all")
            self.canvas.create_image(0, 0, anchor=tk.NW, image=self.photo)
            self.canvas.config(scrollregion=self.canvas.bbox(tk.ALL))
            self.page_label.config(text=f"Page: {self.current_page + 1}/{len(self.pdf_document)}")
            self.sync_manager_stats_only()
            self.update_tool_pane()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to display page: {e}")

    # ================================================================
    # SAVE SESSION - WITH HIGHLIGHTER SERIALIZATION
    # ================================================================

    def save_session(self):
        """Save current session to JSON file with all annotation types including highlights"""
        if not self.pdf_document:
            messagebox.showwarning("No PDF", "Load a PDF first before saving a session.")
            return

        if not hasattr(self, 'project_dirs') or not self.project_dirs.get("sessions"):
            messagebox.showerror("Error", "Project directories not set up. Load a PDF first.")
            return

        save_path = os.path.join(
            self.project_dirs["sessions"],
            f"{self.cabinet_id}_annotations.json"
        )

        data = {
            'project_name': self.project_name,
            'sales_order_no': self.sales_order_no,
            'cabinet_id': getattr(self, 'cabinet_id', ''),
            'pdf_path': self.current_pdf_path,
            'current_page': self.current_page,
            'zoom_level': self.zoom_level,
            'current_sr_no': self.current_sr_no,
            'session_refs': list(self.session_refs),
            'annotations': [],
            'undo_stack_size': len(self.undo_stack) if hasattr(self, 'undo_stack') else 0,
            'save_timestamp': datetime.now().isoformat()
        }

        # Process all annotation types
        for ann in self.annotations:
            entry = ann.copy()

            # ===== HIGHLIGHTER ANNOTATIONS - Convert tuples to lists =====
            if 'points_page' in entry:
                entry['points_page'] = [[float(x), float(y)] for x, y in entry['points_page']]
            
            # ===== BBOX for highlights and old rectangles =====
            if 'bbox_page' in entry:
                entry['bbox_page'] = [float(x) for x in entry['bbox_page']]

            # ===== PEN STROKES - Convert tuples to lists =====
            if 'points' in entry:
                entry['points'] = [[float(x), float(y)] for x, y in entry['points']]

            # ===== TEXT ANNOTATIONS - Convert tuple to list =====
            if 'pos_page' in entry:
                pos = entry['pos_page']
                entry['pos_page'] = [float(pos[0]), float(pos[1])]
            
            # Ensure text content is saved
            if 'text' in entry:
                entry['text'] = str(entry['text'])

            data['annotations'].append(entry)
        
        self.sync_manager_stats_only()

        try:
            with open(save_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            
            # Count annotation types for feedback
            highlight_count = len([a for a in self.annotations if a.get('type') == 'highlight'])
            pen_count = len([a for a in self.annotations if a.get('type') == 'pen'])
            text_count = len([a for a in self.annotations if a.get('type') == 'text'])
            
            summary = f"Session saved successfully!\n\n"
            summary += f"Total annotations: {len(self.annotations)}\n"
            if highlight_count > 0:
                summary += f"🖍️ Highlights: {highlight_count}\n"
            if pen_count > 0:
                summary += f"✏️ Pen strokes: {pen_count}\n"
            if text_count > 0:
                summary += f"🅰️ Text annotations: {text_count}\n"
            summary += f"\nSaved to:\n{save_path}"
            
            messagebox.showinfo("Saved", summary)
            self._flash_status(f"✓ Saved {len(self.annotations)} annotations", bg='#10b981')

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save session:\n{e}")
            import traceback
            traceback.print_exc()

    # ================================================================
    # LOAD SESSION - WITH HIGHLIGHTER DESERIALIZATION
    # ================================================================

    def load_session(self):
        """Load session from JSON file via file dialog"""
        path = filedialog.askopenfilename(
            title="Load Session JSON",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if not path:
            return

        self.load_session_from_path(path)
        self.sync_manager_stats_only()

    def load_session_from_path(self, path):
        """Load session from a specific JSON file path with all annotation types"""
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror("Session Load Error", f"Failed to load session:\n{e}")
            return

        # Restore basic state
        self.project_name = data.get('project_name', self.project_name)
        self.sales_order_no = data.get('sales_order_no', self.sales_order_no)
        self.cabinet_id = data.get('cabinet_id', getattr(self, "cabinet_id", ""))
        self.current_page = data.get('current_page', 0)
        self.zoom_level = data.get('zoom_level', 1.0)
        self.current_sr_no = data.get('current_sr_no', self.current_sr_no)

        # Restore session refs
        self.session_refs = set(data.get('session_refs', []))

        # Restore annotations with proper type conversion
        self.annotations = []
        highlight_count = 0
        pen_count = 0
        text_count = 0
        
        for entry in data.get('annotations', []):
            ann = entry.copy()

            # ===== HIGHLIGHTER ANNOTATIONS - Convert lists to tuples =====
            if 'points_page' in ann:
                ann['points_page'] = [(float(p[0]), float(p[1])) for p in ann['points_page']]
                highlight_count += 1
            
            # ===== BBOX - Convert list back to tuple =====
            if 'bbox_page' in ann:
                ann['bbox_page'] = tuple(float(x) for x in ann['bbox_page'])

            # ===== PEN STROKES - Convert lists to tuples =====
            if 'points' in ann:
                ann['points'] = [(float(p[0]), float(p[1])) for p in ann['points']]
                pen_count += 1

            # ===== TEXT ANNOTATIONS - Convert list to tuple =====
            if 'pos_page' in ann:
                pos = ann['pos_page']
                ann['pos_page'] = (float(pos[0]), float(pos[1]))
                text_count += 1
            
            # Ensure text content is restored
            if 'text' in ann:
                ann['text'] = str(ann['text'])

            self.annotations.append(ann)

            # Add ref_no to session refs
            if ann.get('ref_no'):
                self.session_refs.add(str(ann['ref_no']).strip())

        self.display_page()
        
        summary = f"Session loaded successfully!\n\n"
        summary += f"Total annotations: {len(self.annotations)}\n"
        if highlight_count > 0:
            summary += f"🖍️ Highlights: {highlight_count}\n"
        if pen_count > 0:
            summary += f"✏️ Pen strokes: {pen_count}\n"
        if text_count > 0:
            summary += f"🅰️ Text annotations: {text_count}\n"
        summary += f"\nMake sure the same PDF is open."
        
        messagebox.showinfo("Loaded", summary)
        self._flash_status(f"✓ Loaded {len(self.annotations)} annotations", bg='#3b82f6')

    # ================================================================
    # EXPORT ANNOTATED PDF - WITH HIGHLIGHTER SUPPORT
    # ================================================================

    # ============================================================================
    # FIXED TRANSFORMATION METHODS FOR HIGHLIGHTER ANNOTATIONS
    # Replace the existing transform methods with these updated versions
    # ============================================================================

    def transform_bbox_for_rotation(self, rect, page):
        """Transform bbox for page rotation (for old rectangle annotations)"""
        r = page.rotation
        w = page.rect.width
        h = page.rect.height
        x1, y1, x2, y2 = rect

        if r == 0:
            return fitz.Rect(x1, y1, x2, y2)
        if r == 90:
            return fitz.Rect(y1, w - x2, y2, w - x1)
        if r == 180:
            return fitz.Rect(w - x2, h - y2, w - x1, h - y1)
        if r == 270:
            return fitz.Rect(h - y2, x1, h - y1, x2)

        return fitz.Rect(x1, y1, x2, y2)


    def transform_point_for_rotation(self, point, page):
        """Transform a single point (x, y) for page rotation
        
        Used for:
        - Pen stroke points
        - Text annotation positions
        """
        r = page.rotation
        w = page.rect.width
        h = page.rect.height
        x, y = point

        if r == 0:
            return fitz.Point(x, y)
        elif r == 90:
            return fitz.Point(y, w - x)
        elif r == 180:
            return fitz.Point(w - x, h - y)
        elif r == 270:
            return fitz.Point(h - y, x)
        
        return fitz.Point(x, y)


    def transform_highlight_points_for_rotation(self, points, page):
        """Transform highlighter stroke points for page rotation
        
        Highlighters store a list of (x, y) tuples representing the stroke path.
        Each point needs to be individually transformed based on page rotation.
        
        Args:
            points: List of (x, y) tuples representing the highlight stroke
            page: PyMuPDF page object with rotation info
            
        Returns:
            List of fitz.Point objects, transformed for the page rotation
        """
        r = page.rotation
        w = page.rect.width
        h = page.rect.height
        
        transformed_points = []
        
        for point in points:
            x, y = point
            
            if r == 0:
                transformed_points.append(fitz.Point(x, y))
            elif r == 90:
                transformed_points.append(fitz.Point(y, w - x))
            elif r == 180:
                transformed_points.append(fitz.Point(w - x, h - y))
            elif r == 270:
                transformed_points.append(fitz.Point(h - y, x))
            else:
                transformed_points.append(fitz.Point(x, y))
        
        return transformed_points
    def get_text_position_for_highlight(self, rect, page):
        """Get the correct position for text beside a highlight annotation based on page rotation
        
        Args:
            rect: fitz.Rect object (already transformed)
            page: fitz page object with rotation info
        
        Returns:
            fitz.Point object for text position
        """
        r = page.rotation
        offset = 5  # Small gap between highlight and text
        
        if r == 0:
            # Normal orientation - text to the right of highlight
            return fitz.Point(rect.x1 + offset, rect.y0)
        elif r == 90:
            # 90° rotation - text below highlight
            return fitz.Point(rect.x0, rect.y1 + offset)
        elif r == 180:
            # 180° rotation - text to the left of highlight
            return fitz.Point(rect.x0 - offset, rect.y1)
        elif r == 270:
            # 270° rotation - text above highlight
            return fitz.Point(rect.x1, rect.y0 - offset)
        
        # Default fallback
        return fitz.Point(rect.x1 + offset, rect.y0)

    def export_annotated_pdf(self):
        """Export PDF with all annotations including highlighter strokes"""
        if not self.pdf_document:
            messagebox.showwarning("Warning", "Please load a PDF first")
            return

        if not hasattr(self, 'project_dirs') or not self.project_dirs.get("annotated_drawings"):
            messagebox.showerror("Error", "Project directories not set up.")
            return

        try:
            save_path = os.path.join(
                self.project_dirs["annotated_drawings"],
                f"{self.cabinet_id.replace(' ', '_')}_Annotated.pdf"
            )

            # Create output PDF
            out_doc = fitz.open()
            for pnum in range(len(self.pdf_document)):
                out_doc.insert_pdf(self.pdf_document, from_page=pnum, to_page=pnum)

            # Open Excel for SR No lookup
            wb = None
            ws = None
            if self.excel_file and os.path.exists(self.excel_file):
                try:
                    wb = load_workbook(self.excel_file, data_only=True)
                    ws = wb[self.punch_sheet_name]
                except:
                    pass

            # Draw annotations
            for ann in self.annotations:
                p = ann.get('page')
                if p is None or p < 0 or p >= len(out_doc):
                    continue

                target_page = out_doc[p]
                ann_type = ann.get('type')

                # -------- HIGHLIGHTER ANNOTATIONS (NEW) --------
                if ann_type == 'highlight' and 'points_page' in ann:
                    points_page = ann['points_page']
                    if len(points_page) >= 2:
                        color_key = ann.get('color', 'yellow')
                        rgb = self.highlighter_colors[color_key]['rgb']
                        # Normalize RGB to 0-1 range for PyMuPDF
                        color = (rgb[0]/255, rgb[1]/255, rgb[2]/255)
                        
                        # Transform points for page rotation
                        transformed_points = self.transform_highlight_points_for_rotation(
                            points_page, 
                            target_page
                        )
                        
                        # Convert to list of tuples for ink annotation
                        stroke = [(pt.x, pt.y) for pt in transformed_points]
                        
                        if len(stroke) >= 2:
                            ink_list = [stroke]  # Wrap in list for PyMuPDF
                            annot = target_page.add_ink_annot(ink_list)
                            annot.set_colors(stroke=color)
                            annot.set_border(width=15)  # Thick highlighter stroke
                            annot.set_opacity(0.4)  # Semi-transparent
                            annot.update()
                            
                            # Add SR number text if available (for error highlights)
                            if color_key == 'orange' and 'bbox_page' in ann:
                                sr_text = None
                                row = ann.get('excel_row')
                                if row and ws:
                                    try:
                                        sr_val = self.read_cell(ws, row, self.punch_cols['sr_no'])
                                        if sr_val is not None:
                                            sr_text = f"Sr {sr_val}"
                                    except:
                                        pass
                                
                                if sr_text:
                                    # Use bbox for text position
                                    x1, y1, x2, y2 = ann['bbox_page']
                                    bbox_rect = self.transform_bbox_for_rotation(
                                        (x1, y1, x2, y2), 
                                        target_page
                                    )
                                    # Position text beside the highlight
                                    text_pos = self.get_text_position_for_highlight(bbox_rect, target_page)
                                    try:
                                        target_page.insert_text(
                                            text_pos, 
                                            sr_text, 
                                            fontsize=8, 
                                            color=(1, 0, 0)
                                        )
                                    except:
                                        pass


                # -------- PEN STROKES --------
                elif ann_type == 'pen' and 'points' in ann:
                    points = ann['points']
                    if len(points) >= 2:
                        # Transform points for rotation
                        transformed_points = [
                            self.transform_point_for_rotation(pt, target_page) 
                            for pt in points
                        ]
                        
                        # Draw lines between consecutive points
                        for i in range(len(transformed_points) - 1):
                            p1 = transformed_points[i]
                            p2 = transformed_points[i + 1]
                            target_page.draw_line(p1, p2, color=(1, 0, 0), width=2)

                # -------- TEXT ANNOTATIONS --------
                elif ann_type == 'text' and 'pos_page' in ann:
                    pos = ann['pos_page']
                    text = ann.get('text', '')
                    if text:
                        text_point = self.transform_point_for_rotation(pos, target_page)
                        try:
                            target_page.insert_text(
                                text_point, text,
                                fontsize=10, color=(1, 0, 0),
                                rotate=target_page.rotation
                            )
                        except:
                            pass

            if wb:
                wb.close()

            out_doc.save(save_path)
            out_doc.close()

            messagebox.showinfo("Success", f"Annotated PDF saved to:\n{save_path}")
            self.sync_manager_stats_only()

        except PermissionError:
            messagebox.showerror("Error", "Close the target file (if open) and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export annotated PDF:\n{e}")
            import traceback
            traceback.print_exc()


    # ================================================================
    # UI SETUP WITH HIGHLIGHTER CONTROLS
    # ================================================================

    def setup_ui(self):
        """Setup modern UI with highlighter controls"""
        # Main toolbar
        toolbar = tk.Frame(self.root, bg='#1e293b', height=80)
        toolbar.pack(side=tk.TOP, fill=tk.X)
        
        # Enhanced Menu Bar
        menubar = Menu(self.root, bg='#1e293b', fg='white', activebackground='#3b82f6')
        self.root.config(menu=menubar)
        
        # File Menu
        file_menu = Menu(menubar, tearoff=0, bg='#1e293b', fg='white', activebackground='#3b82f6')
        menubar.add_cascade(label="📁 File", menu=file_menu)
        file_menu.add_command(label="Open PDF", command=self.load_pdf, accelerator="Ctrl+O")
        file_menu.add_separator()
        file_menu.add_command(label="Load Session", command=self.load_session, accelerator="Ctrl+L")
        file_menu.add_command(label="Save Session", command=self.save_session, accelerator="Ctrl+S")
        file_menu.add_separator()
        file_menu.add_command(label="Export Annotated PDF", command=self.export_annotated_pdf, accelerator="Ctrl+E")
        file_menu.add_command(label="Save Interphase Excel", command=self.save_interphase_excel)
        file_menu.add_command(label="Open Excel", command=self.open_excel, accelerator="Ctrl+Shift+E")
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)
        
        # Tools Menu
        tools_menu = Menu(menubar, tearoff=0, bg='#1e293b', fg='white', activebackground='#3b82f6')
        menubar.add_cascade(label="🛠️ Tools", menu=tools_menu)
        tools_menu.add_command(label="Review Checklist", command=self.review_checklist_now, accelerator="Ctrl+R")
        tools_menu.add_command(label="Punch Closing Mode", command=self.punch_closing_mode, accelerator="Ctrl+Shift+P")
        tools_menu.add_separator()
        tools_menu.add_command(label="🔍 Verify Rework", command=self.view_production_handbacks, accelerator="Ctrl+Shift+V")
        
        # View Menu
        view_menu = Menu(menubar, tearoff=0, bg='#1e293b', fg='white', activebackground='#3b82f6')
        menubar.add_cascade(label="👁️ View", menu=view_menu)
        view_menu.add_command(label="Zoom In", command=self.zoom_in, accelerator="Ctrl++")
        view_menu.add_command(label="Zoom Out", command=self.zoom_out, accelerator="Ctrl+-")
        view_menu.add_command(label="Reset Zoom", command=lambda: setattr(self, 'zoom_level', 1.0) or self.display_page())
        
        # Keyboard shortcuts
        self.root.bind_all("<Control-o>", lambda e: self.load_pdf())
        self.root.bind_all("<Control-s>", lambda e: self.save_session())
        self.root.bind_all("<Control-l>", lambda e: self.load_session())
        self.root.bind_all("<Control-e>", lambda e: self.export_annotated_pdf())
        self.root.bind_all("<Control-r>", lambda e: self.review_checklist_now())
        self.root.bind_all("<Control-z>", lambda e: self.undo_last_action())
        self.root.bind_all("<Control-P>", lambda e: self.punch_closing_mode())
        self.root.bind_all("<Control-E>", lambda e: self.open_excel())
        self.root.bind_all("<Control-V>", lambda e: self.view_production_handbacks())
        self.root.bind_all("<Control-plus>", lambda e: self.zoom_in())
        self.root.bind_all("<Control-minus>", lambda e: self.zoom_out())
        self.root.bind_all("<Escape>", lambda e: self.deactivate_all())
        
        # Modern button style
        btn_style = {
            'bg': '#3b82f6',
            'fg': 'white',
            'padx': 12,
            'pady': 10,
            'font': ('Segoe UI', 9, 'bold'),
            'relief': tk.FLAT,
            'borderwidth': 0,
            'cursor': 'hand2'
        }
        
        # Left section - File operations
        left_frame = tk.Frame(toolbar, bg='#1e293b')
        left_frame.pack(side=tk.LEFT, padx=10, pady=10)
        
        tk.Button(left_frame, text="📁 Open PDF", command=self.load_pdf, **btn_style).pack(side=tk.LEFT, padx=3)
        
        # Recent Projects Dropdown
        recent_frame = tk.Frame(left_frame, bg='#1e293b')
        recent_frame.pack(side=tk.LEFT, padx=8)
        
        tk.Label(recent_frame, text="Recent:", bg='#1e293b', fg='#94a3b8',
                font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=(0, 5))
        
        self.recent_var = tk.StringVar(value="Select Project...")
        self.recent_dropdown = tk.OptionMenu(recent_frame, self.recent_var,
                                            "Select Project...",
                                            command=self.load_recent_projects_ui)
        self.recent_dropdown.config(bg='#334155', fg='white', font=('Segoe UI', 9),
                                   width=22, relief=tk.FLAT, borderwidth=0)
        self.recent_dropdown.pack(side=tk.LEFT)
        
        # Center - HIGHLIGHTER COLOR PICKER (Circular button) - UPDATED
        highlighter_frame = tk.Frame(toolbar, bg='#1e293b')
        highlighter_frame.pack(side=tk.LEFT, padx=30)
        
        tk.Label(highlighter_frame, text="Highlighter:", bg='#1e293b', fg='#94a3b8',
                font=('Segoe UI', 9, 'bold')).pack(side=tk.LEFT, padx=(0, 10))
        
        # Color picker with circular button
        self.color_picker_frame = tk.Frame(highlighter_frame, bg='#1e293b')
        self.color_picker_frame.pack(side=tk.LEFT)
        
        # Circular color button - NO BOX BORDER
        self.color_canvas = tk.Canvas(
            self.color_picker_frame,
            width=44,
            height=44,
            bg='#1e293b',
            highlightthickness=0,
            borderwidth=0,
            cursor='hand2'
        )
        self.color_canvas.pack(side=tk.LEFT)
        self.update_color_button()
        self.color_canvas.bind("<Button-1>", lambda e: self.toggle_highlighter())
        
        # Dropdown arrow - sleeker design
        self.dropdown_btn = tk.Button(
            self.color_picker_frame,
            text="▼",
            font=('Segoe UI', 8),
            bg='#1e293b',
            fg='#94a3b8',
            activebackground='#334155',
            activeforeground='white',
            relief=tk.FLAT,
            borderwidth=0,
            width=2,
            height=1,
            command=self.show_color_menu,
            cursor='hand2'
        )
        self.dropdown_btn.pack(side=tk.LEFT, padx=(4, 0))
        
        # Navigation
        center_frame = tk.Frame(toolbar, bg='#1e293b')
        center_frame.pack(side=tk.LEFT, padx=20)
        
        self.page_label = tk.Label(center_frame, text="Page: 0/0", bg='#1e293b',
                                   fg='white', font=('Segoe UI', 10, 'bold'))
        self.page_label.pack(side=tk.LEFT, padx=10)
        
        nav_btn_style = btn_style.copy()
        nav_btn_style['bg'] = '#64748b'
        
        tk.Button(center_frame, text="◀", command=self.prev_page, width=3,
                 **nav_btn_style).pack(side=tk.LEFT, padx=2)
        tk.Button(center_frame, text="▶", command=self.next_page, width=3,
                 **nav_btn_style).pack(side=tk.LEFT, padx=2)
        
        # Zoom controls
        zoom_frame = tk.Frame(center_frame, bg='#1e293b')
        zoom_frame.pack(side=tk.LEFT, padx=15)
        
        zoom_btn_style = btn_style.copy()
        zoom_btn_style['bg'] = '#10b981'
        
        tk.Button(zoom_frame, text="🔍+", command=self.zoom_in, width=4,
                 **zoom_btn_style).pack(side=tk.LEFT, padx=2)
        tk.Button(zoom_frame, text="🔍−", command=self.zoom_out, width=4,
                 **zoom_btn_style).pack(side=tk.LEFT, padx=2)
        
        # Tool section
        tool_frame = tk.Frame(toolbar, bg='#1e293b')
        tool_frame.pack(side=tk.LEFT, padx=10)
        
        tk.Label(tool_frame, text="Tools:", bg='#1e293b', fg='#94a3b8',
                font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=(0, 8))
        
        # Load icons or use fallback
        try:
            assets_dir = os.path.join(os.path.dirname(get_app_base_dir()), "assets")
            icon_size = (44, 44)
            
            pen_icon_path = os.path.join(assets_dir, "pen_icon.png")
            pen_img = Image.open(pen_icon_path).resize(icon_size, Image.Resampling.LANCZOS)
            self.pen_icon = ImageTk.PhotoImage(pen_img)
            
            text_icon_path = os.path.join(assets_dir, "text_icon.png")
            text_img = Image.open(text_icon_path).resize(icon_size, Image.Resampling.LANCZOS)
            self.text_icon = ImageTk.PhotoImage(text_img)
            
            undo_icon_path = os.path.join(assets_dir, "undo_icon.png")
            undo_img = Image.open(undo_icon_path).resize(icon_size, Image.Resampling.LANCZOS)
            self.undo_icon = ImageTk.PhotoImage(undo_img)
            
            self.pen_btn = tk.Button(tool_frame, image=self.pen_icon,
                                    command=lambda: self.set_tool_mode("pen"),
                                    bg='#334155', width=48, height=48,
                                    relief=tk.FLAT, cursor='hand2')
            self.pen_btn.pack(side=tk.LEFT, padx=2)
            
            self.text_btn = tk.Button(tool_frame, image=self.text_icon,
                                     command=lambda: self.set_tool_mode("text"),
                                     bg='#334155', width=48, height=48,
                                     relief=tk.FLAT, cursor='hand2')
            self.text_btn.pack(side=tk.LEFT, padx=2)
            
            self.undo_btn = tk.Button(tool_frame, image=self.undo_icon,
                                      command=self.undo_last_action,
                                      bg='#334155', width=48, height=48,
                                      relief=tk.FLAT, cursor='hand2')
            self.undo_btn.pack(side=tk.LEFT, padx=2)
            
        except Exception as e:
            print(f"Could not load tool icons: {e}")
        
        # Right section - Action buttons
        right_frame = tk.Frame(toolbar, bg='#1e293b')
        right_frame.pack(side=tk.RIGHT, padx=10, pady=10)
        
        verify_btn_style = btn_style.copy()
        verify_btn_style['bg'] = '#ec4899'
        
        tk.Button(right_frame, text="🔍 Verify Rework",
                 command=self.view_production_handbacks,
                 **verify_btn_style).pack(side=tk.RIGHT, padx=3)
        
        handover_btn_style = btn_style.copy()
        handover_btn_style['bg'] = '#8b5cf6'
        
        tk.Button(right_frame, text="🚀 Handover to Production",
                 command=self.handover_to_production,
                 **handover_btn_style).pack(side=tk.RIGHT, padx=3)
        
        # Canvas with scrollbars
        canvas_frame = tk.Frame(self.root, bg='#f1f5f9')
        canvas_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        v_scrollbar = tk.Scrollbar(canvas_frame, orient=tk.VERTICAL)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        h_scrollbar = tk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.canvas = tk.Canvas(canvas_frame, bg='#f8fafc',
                               yscrollcommand=v_scrollbar.set,
                               xscrollcommand=h_scrollbar.set,
                               highlightthickness=0)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        v_scrollbar.config(command=self.canvas.yview)
        h_scrollbar.config(command=self.canvas.xview)
        
        # Bind mouse events
        self.canvas.bind("<ButtonPress-1>", self.on_left_press)
        self.canvas.bind("<B1-Motion>", self.on_left_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_left_release_with_ocr)
        
        # Modern status bar
        status_bar = tk.Frame(self.root, bg='#334155', height=40)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        instructions_text = "🖍️ Highlighter: Auto-Straighten ON | ✏️ Pen: Freehand | 🅰️ Text | Esc: Deactivate | Ctrl+Z: Undo"
        tk.Label(status_bar, text=instructions_text, bg='#334155', fg='#e2e8f0',
                font=('Segoe UI', 9), pady=10).pack()


    # ================================================================
    # HIGHLIGHTER UI HELPERS - UPDATED
    # ================================================================

    def update_color_button(self):
        """Update the circular color button display - sharper widget style"""
        self.color_canvas.delete("all")
        
        # Get current color
        rgb = self.highlighter_colors[self.current_color_key]['rgb']
        hex_color = f'#{rgb[0]:02x}{rgb[1]:02x}{rgb[2]:02x}'
        
        # Draw clean circular button with subtle shadow effect
        if self.active_highlighter:
            # Active state - prominent glow ring
            self.color_canvas.create_oval(
                0, 0, 44, 44,
                outline='#3b82f6',
                width=3,
                fill='#1e293b'
            )
            # Main color circle
            self.color_canvas.create_oval(
                6, 6, 38, 38,
                fill=hex_color,
                outline='',
                width=0
            )
        else:
            # Inactive state - subtle border
            self.color_canvas.create_oval(
                2, 2, 42, 42,
                outline='#475569',
                width=1,
                fill='#1e293b'
            )
            # Main color circle
            self.color_canvas.create_oval(
                6, 6, 38, 38,
                fill=hex_color,
                outline='',
                width=0
            )


    def show_color_menu(self):
        """Show color picker dropdown menu"""
        menu = Menu(self.root, tearoff=0, bg='#1e293b', fg='white',
                   activebackground='#3b82f6', activeforeground='white',
                   font=('Segoe UI', 10))
        
        for color_key, color_info in self.highlighter_colors.items():
            rgb = color_info['rgb']
            hex_color = f'#{rgb[0]:02x}{rgb[1]:02x}{rgb[2]:02x}'
            label = f"● {color_info['name']}"
            
            menu.add_command(
                label=label,
                command=lambda ck=color_key: self.change_color(ck),
                foreground=hex_color,
                font=('Arial', 12, 'bold')
            )
        
        x = self.dropdown_btn.winfo_rootx()
        y = self.dropdown_btn.winfo_rooty() + self.dropdown_btn.winfo_height()
        menu.post(x, y)


    def change_color(self, color_key):
        """Change the highlighter color"""
        self.current_color_key = color_key
        self.update_color_button()
        
        if self.active_highlighter:
            self.active_highlighter = color_key
            self.root.config(cursor="pencil")


    def toggle_highlighter(self):
        """Toggle highlighter on/off"""
        if self.active_highlighter:
            self.active_highlighter = None
            self.root.config(cursor="")
            self.update_color_button()
        else:
            self.active_highlighter = self.current_color_key
            self.root.config(cursor="pencil")
            self.update_color_button()
            
            if self.tool_mode:
                self.tool_mode = None
                self.pen_btn.config(bg='#334155', relief=tk.FLAT)
                self.text_btn.config(bg='#334155', relief=tk.FLAT)

    def set_tool_mode(self, mode):
        """Set tool mode (pen or text)"""
        if self.active_highlighter:
            self.toggle_highlighter()
        
        if self.tool_mode == mode:
            self.tool_mode = None
            if mode == "pen":
                self.pen_btn.config(bg='#334155', relief=tk.FLAT)
            else:
                self.text_btn.config(bg='#334155', relief=tk.FLAT)
        else:
            self.tool_mode = mode
            if mode == "pen":
                self.pen_btn.config(bg='#3b82f6', relief=tk.SUNKEN)
                self.text_btn.config(bg='#334155', relief=tk.FLAT)
            else:
                self.text_btn.config(bg='#3b82f6', relief=tk.SUNKEN)
                self.pen_btn.config(bg='#334155', relief=tk.FLAT)

    def deactivate_all(self):
        """Deactivate all tools and highlighters"""
        if self.active_highlighter:
            self.toggle_highlighter()
        if self.tool_mode:
            self.set_tool_mode(self.tool_mode)

    def update_tool_pane(self):
        """Update annotation statistics"""
        # Placeholder - implement if you have a tool pane
        pass

    def _flash_status(self, message, bg='#10b981'):
        """Show a temporary status message"""
        status_label = tk.Label(
            self.root, 
            text=message, 
            bg=bg, 
            fg='white', 
            font=('Segoe UI', 10, 'bold'),
            padx=25, 
            pady=12,
            relief=tk.FLAT
        )
        status_label.place(relx=0.5, rely=0.08, anchor='center')
        self.root.after(1500, lambda: status_label.destroy())

    # ================================================================
    # UNDO FUNCTIONALITY
    # ================================================================

    def add_to_undo_stack(self, action_type, annotation):
        """Add an action to the undo stack"""
        self.undo_stack.append({
            'type': action_type,
            'annotation': annotation.copy()
        })
        
        if len(self.undo_stack) > self.max_undo:
            self.undo_stack.pop(0)

    def undo_last_action(self):
        """Undo the last annotation action"""
        if not self.undo_stack:
            messagebox.showinfo("Nothing to Undo", "No actions to undo.", icon='info')
            return
        
        last_action = self.undo_stack.pop()
        
        if last_action['type'] == 'add_annotation':
            annotation = last_action['annotation']
            if annotation in self.annotations:
                self.annotations.remove(annotation)
                self.display_page()
                self._flash_status("✓ Annotation removed", bg='#10b981')
        
        self.update_tool_pane()

    def clear_all_annotations(self):
        """Clear all annotations with confirmation"""
        if not self.annotations:
            messagebox.showinfo("No Annotations", "There are no annotations to clear.", icon='info')
            return
        
        confirm = messagebox.askyesno(
            "Clear All Annotations",
            f"Are you sure you want to delete all {len(self.annotations)} annotations?\n\n"
            "This action cannot be undone!",
            icon='warning'
        )
        
        if confirm:
            self.annotations.clear()
            self.selected_annotation = None
            self.undo_stack.clear()
            self.display_page()
            self.update_tool_pane()
            messagebox.showinfo("Cleared", "All annotations have been removed.", icon='info')

    # ================================================================
    # NAVIGATION AND ZOOM
    # ================================================================

    def prev_page(self):
        if self.pdf_document and self.current_page > 0:
            self.current_page -= 1
            self.display_page()

    def next_page(self):
        if self.pdf_document and self.current_page < len(self.pdf_document) - 1:
            self.current_page += 1
            self.display_page()

    def zoom_in(self):
        if self.zoom_level < 3.0:
            self.zoom_level += 0.25
            self.display_page()

    def zoom_out(self):
        if self.zoom_level > 0.5:
            self.zoom_level -= 0.25
            self.display_page()

    # ================================================================
    # PLACEHOLDER METHODS - Implement from your original code
    # ================================================================

    def load_pdf(self):
        """Load PDF - implement with your original logic"""
        file_path = filedialog.askopenfilename(
            title="Select Circuit Diagram PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if file_path:
            try:
                self.pdf_document = fitz.open(file_path)
                self.current_pdf_path = file_path
                self.current_page = 0
                self.annotations = []
                self.zoom_level = 1.0
                self.tool_mode = None
                self.active_highlighter = None
                self.update_color_button()
                self.root.config(cursor="")
                self.current_sr_no = self.get_next_sr_no()
                self.display_page()
                messagebox.showinfo("Success", f"Loaded PDF with {len(self.pdf_document)} pages")
                
                self.ask_project_details()
                self.prepare_project_folders()

                try:
                    self.working_excel_path = os.path.join(
                        self.project_dirs["working_excel"],
                        f"{self.cabinet_id.replace(' ', '_')}_Working.xlsx"
                    )

                    if os.path.exists(self.working_excel_path):
                        resume = messagebox.askyesno(
                            "Resume Inspection",
                            f"Existing working Excel found. Resume previous inspection?"
                        )
                        if not resume:
                            shutil.copy2(self.master_excel_file, self.working_excel_path)
                    else:
                        shutil.copy2(self.master_excel_file, self.working_excel_path)

                    self.excel_file = self.working_excel_path

                except Exception as e:
                    messagebox.showerror("Excel Error", f"Failed to prepare working Excel:\n{e}")
                    return

                self.write_project_details_to_excel()

                expected_session_path = os.path.join(
                    self.project_dirs["sessions"],
                    f"{self.cabinet_id}_annotations.json"
                )

                self.db.update_project(self.cabinet_id, {
                    'pdf_path': self.current_pdf_path,
                    'excel_path': self.excel_file,
                    'session_path': expected_session_path if os.path.exists(expected_session_path) else None,
                    'storage_location': self.storage_location,
                    'last_accessed': datetime.now().isoformat()
                })

                if os.path.exists(expected_session_path):
                    resume = messagebox.askyesno(
                        "Resume Session",
                        "Existing session found. Do you want to resume it?"
                    )
                    if resume:
                        self.load_session_from_path(expected_session_path)
                
                self.save_recent_project()

            except Exception as e:
                messagebox.showerror("Error", f"Failed to load PDF: {str(e)}")

    def load_categories(self):
        """Load categories from JSON"""
        try:
            if os.path.exists(self.category_file):
                with open(self.category_file, "r", encoding="utf-8") as f:
                    self.categories = json.load(f)
            else:
                self.categories = []
        except Exception as e:
            print(f"Error loading categories: {e}")
            self.categories = []

    def get_next_sr_no(self):
        """Get next serial number"""
        try:
            if not self.excel_file or not os.path.exists(self.excel_file):
                return 1
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
            last_sr_no = 0
            row_num = 8
            while row_num <= ws.max_row + 5:
                val = self.read_cell(ws, row_num, self.punch_cols['sr_no'])
                if val is None:
                    break
                try:
                    last_sr_no = int(val)
                except:
                    pass
                row_num += 1
            wb.close()
            return last_sr_no + 1
        except Exception:
            return 1

    def run_template(self, template_def, tag_name=None):
        """Execute a template definition"""
        values = {}
        if tag_name:
            values["tag"] = tag_name

        for inp in template_def.get("inputs", []):
            val = simpledialog.askstring("Input Required", inp["label"], parent=self.root)
            if not val:
                return None
            values[inp["name"]] = val.strip()

        try:
            return template_def["template"].format(**values)
        except KeyError as e:
            messagebox.showerror("Template Error", f"Missing placeholder: {e}")
            return None

    # Excel cell helpers
    def split_cell(self, cell_ref):
        m = re.match(r"([A-Z]+)(\d+)", cell_ref)
        if not m:
            raise ValueError(f"Invalid cell reference: {cell_ref}")
        col, row = m.groups()
        return int(row), col

    def _resolve_merged_target(self, ws, row, col_idx):
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col_idx <= merged.max_col:
                return merged.min_row, merged.min_col
        return row, col_idx

    def write_cell(self, ws, row, col, value):
        if isinstance(col, str):
            col_idx = column_index_from_string(col)
        else:
            col_idx = int(col)
        target_row, target_col = self._resolve_merged_target(ws, int(row), col_idx)
        ws.cell(row=target_row, column=target_col).value = value

    def read_cell(self, ws, row, col):
        if isinstance(col, str):
            col_idx = column_index_from_string(col)
        else:
            col_idx = int(col)
        target_row, target_col = self._resolve_merged_target(ws, int(row), col_idx)
        return ws.cell(row=target_row, column=target_col).value

    def update_interphase_status_for_ref(self, ref_no, status='NOK'):
        """Update Interphase status"""
        try:
            wb = load_workbook(self.excel_file)
            if self.interphase_sheet_name not in wb.sheetnames:
                wb.close()
                return False
            ws = wb[self.interphase_sheet_name]
            
            updated_any = False
            current_date = datetime.now().strftime("%Y-%m-%d")
            
            try:
                username = os.getlogin()
            except:
                username = getpass.getuser()
            
            for r in range(1, ws.max_row + 1):
                cell_val = self.read_cell(ws, r, self.interphase_cols['ref_no'])
                if cell_val and str(cell_val).strip() == str(ref_no).strip():
                    self.write_cell(ws, r, self.interphase_cols['status'], status)
                    self.write_cell(ws, r, self.interphase_cols['name'], username)
                    self.write_cell(ws, r, self.interphase_cols['date'], current_date)
                    updated_any = True
            
            if updated_any:
                wb.save(self.excel_file)
            wb.close()
            return updated_any
        except Exception as e:
            print(f"Interphase update error: {e}")
            return False

    # Placeholder methods - implement from original code
    def ask_project_details(self):
        """Ask for project details including storage location"""
        dlg = tk.Toplevel(self.root)
        dlg.title("Project Details")
        dlg.geometry("500x380")
        dlg.transient(self.root)
        dlg.grab_set()

        tk.Label(dlg, text="Cabinet ID", font=('Segoe UI', 10, 'bold')).pack(anchor="w", padx=20, pady=(15, 0))
        cabinet_var = tk.StringVar(value=getattr(self, "cabinet_id", ""))
        tk.Entry(dlg, textvariable=cabinet_var, font=('Segoe UI', 10)).pack(fill="x", padx=20)

        tk.Label(dlg, text="Project Name", font=('Segoe UI', 10, 'bold')).pack(anchor="w", padx=20, pady=(10, 0))
        project_var = tk.StringVar(value=self.project_name)
        project_entry = tk.Entry(dlg, textvariable=project_var, font=('Segoe UI', 10))
        project_entry.pack(fill="x", padx=20)

        tk.Label(dlg, text="Sales Order Number", font=('Segoe UI', 10, 'bold')).pack(anchor="w", padx=20, pady=(10, 0))
        so_var = tk.StringVar(value=self.sales_order_no)
        tk.Entry(dlg, textvariable=so_var, font=('Segoe UI', 10)).pack(fill="x", padx=20)

        # Storage Location Frame
        tk.Label(dlg, text="Storage Location", font=('Segoe UI', 10, 'bold')).pack(anchor="w", padx=20, pady=(15, 0))
        
        location_frame = tk.Frame(dlg)
        location_frame.pack(fill="x", padx=20, pady=5)
        
        location_var = tk.StringVar(value=getattr(self, "storage_location", ""))
        location_entry = tk.Entry(location_frame, textvariable=location_var, 
                                 font=('Segoe UI', 9), state='readonly')
        location_entry.pack(side=tk.LEFT, fill="x", expand=True, padx=(0, 5))
        
        def browse_location():
            folder = filedialog.askdirectory(
                title="Select Project Storage Location",
                mustexist=True
            )
            if folder:
                location_var.set(folder)
        
        browse_btn = tk.Button(location_frame, text="Browse...", command=browse_location,
                              bg='#3b82f6', fg='white', font=('Segoe UI', 9, 'bold'),
                              relief=tk.FLAT, padx=15, pady=5)
        browse_btn.pack(side=tk.RIGHT)

        # Auto-load location when project name changes
        def on_project_name_change(*args):
            project_name = project_var.get().strip()
            if project_name:
                # Check database for existing project location
                existing_location = self.db.get_project_location(project_name)
                if existing_location:
                    location_var.set(existing_location)
                    # Show visual feedback
                    location_entry.config(bg='#dcfce7')  # Light green
                    dlg.after(1000, lambda: location_entry.config(bg='white'))
                else:
                    location_var.set("")
        
        # Bind the trace to project name entry
        project_var.trace('w', on_project_name_change)

        def on_ok():
            cabinet = cabinet_var.get().strip()
            project = project_var.get().strip()
            so = so_var.get().strip()
            location = location_var.get().strip()
            
            if not cabinet or not project:
                messagebox.showerror("Missing Information", 
                                   "Please fill in Cabinet ID and Project Name.")
                return
            
            # Check if this cabinet already exists in database
            if self.db.project_exists(cabinet):
                existing = self.db.get_project(cabinet)
                
                # If it's the same project, use existing location
                if existing['project_name'] == project:
                    location = existing['storage_location']
                    messagebox.showinfo("Existing Cabinet", 
                                      f"Cabinet '{cabinet}' found in project '{project}'.\n"
                                      f"Using existing location:\n{location}")
                else:
                    messagebox.showerror("Error", 
                                       f"Cabinet ID '{cabinet}' already exists in different project:\n"
                                       f"{existing['project_name']}")
                    return
            else:
                # New cabinet - check if project exists with different cabinet
                existing_project_location = self.db.get_project_location(project)
                
                if existing_project_location:
                    # Project exists, use its location
                    location = existing_project_location
                    messagebox.showinfo("Existing Project", 
                                      f"Project '{project}' found.\n"
                                      f"Using existing location:\n{location}")
                else:
                    # Brand new project - must have location
                    if not location:
                        messagebox.showerror("Missing Location", 
                                           "This is a new project. Please select a storage location.")
                        return
            
            self.cabinet_id = cabinet
            self.project_name = project
            self.sales_order_no = so
            self.storage_location = location
            
            # Save to database with all paths
            self.db.add_project({
                'cabinet_id': self.cabinet_id,
                'project_name': self.project_name,
                'sales_order_no': self.sales_order_no,
                'storage_location': self.storage_location,
                'pdf_path': self.current_pdf_path if hasattr(self, 'current_pdf_path') else None,
                'excel_path': self.excel_file if hasattr(self, 'excel_file') else None,
                'created_date': datetime.now().isoformat(),
                'last_accessed': datetime.now().isoformat()
            })
            
            dlg.destroy()

        tk.Button(dlg, text="OK", command=on_ok, 
                 bg="#10b981", fg="white", font=('Segoe UI', 10, 'bold'),
                 relief=tk.FLAT, padx=30, pady=10).pack(pady=20)
        dlg.wait_window()

    def write_project_details_to_excel(self):
        if not self.excel_file or not os.path.exists(self.excel_file):
            return

        try:
            wb = load_workbook(self.excel_file)

            for sheet_name, cells in self.header_cells.items():
                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]

                if getattr(self, "project_name", ""):
                    r, c = self.split_cell(cells["project_name"])
                    self.write_cell(ws, r, c, self.project_name)

                if getattr(self, "sales_order_no", ""):
                    r, c = self.split_cell(cells["sales_order"])
                    self.write_cell(ws, r, c, self.sales_order_no)

                if getattr(self, "cabinet_id", ""):
                    r, c = self.split_cell(cells["cabinet_id"])
                    self.write_cell(ws, r, c, self.cabinet_id)

            wb.save(self.excel_file)
            wb.close()

        except PermissionError:
            messagebox.showerror("Excel Locked", "Please close the Excel file before entering project details.")
        except Exception as e:
            messagebox.showerror("Excel Error", f"Failed to write project details:\n{e}")

    def prepare_project_folders(self):
        """Prepare project folders at user-selected location
        Structure: storage_location/project_name/cabinet_id/...
        """
        if not hasattr(self, 'storage_location') or not self.storage_location:
            messagebox.showerror("Error", "Storage location not set")
            return False
        
        if not self.project_name or not self.cabinet_id:
            messagebox.showerror("Error", "Project name and Cabinet ID required")
            return False
        
        # Create structure: storage_location/project_name/cabinet_id/
        project_folder = os.path.join(
            self.storage_location,
            self.project_name.replace(' ', '_')
        )
        
        cabinet_root = os.path.join(
            project_folder,
            self.cabinet_id.replace(' ', '_')
        )
        
        folders = {
            "root": cabinet_root,
            "working_excel": os.path.join(cabinet_root, "Working_Excel"),
            "interphase_export": os.path.join(cabinet_root, "Interphase_Export"),
            "annotated_drawings": os.path.join(cabinet_root, "Annotated_Drawings"),
            "sessions": os.path.join(cabinet_root, "Sessions")
        }
        
        for p in folders.values():
            os.makedirs(p, exist_ok=True)
        
        self.project_dirs = folders
        return True

    def get_session_path_for_pdf(self):
        if not self.current_pdf_path:
            return None

        session_path = os.path.join(
            self.project_dirs.get("sessions", ""),
            f"{self.cabinet_id}_annotations.json"
        )

        return session_path if os.path.exists(session_path) else None

    # ================================================================
    # CHECKLIST FUNCTIONS
    # ================================================================

    def review_checklist_now(self):
        if not self.excel_file or not os.path.exists(self.excel_file):
            messagebox.showerror("Excel Missing", "Working Excel file not found.")
            return

        self.checklist_file = self.excel_file

        try:
            self.review_checklist_before_save(self.checklist_file, self.session_refs)
        except Exception as e:
            messagebox.showerror("Checklist Error", f"Checklist review failed:\n{e}")

    # ================================================================
    # UPDATED: gather_checklist_matches - Updated for new column structure
    # ================================================================
    def read_open_punches_from_excel(self):
        """Reads punch sheet and returns list of open punches with all details."""
        punches = []

        if not self.excel_file or not os.path.exists(self.excel_file):
            return punches

        try:
            wb = load_workbook(self.excel_file, data_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active

            row = 8
            while True:
                sr = self.read_cell(ws, row, self.punch_cols['sr_no'])
                if sr is None:
                    break

                # Check if punch is closed
                closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
                if closed:
                    row += 1
                    continue

                # Check if punch is implemented
                implemented = bool(self.read_cell(ws, row, self.punch_cols['implemented_name']))

                punches.append({
                    'sr_no': sr,
                    'row': row,
                    'ref_no': self.read_cell(ws, row, self.punch_cols['ref_no']),
                    'punch_text': self.read_cell(ws, row, self.punch_cols['desc']),
                    'category': self.read_cell(ws, row, self.punch_cols['category']),
                    'implemented': implemented,
                    'implemented_name': self.read_cell(ws, row, self.punch_cols['implemented_name']),
                    'implemented_date': self.read_cell(ws, row, self.punch_cols['implemented_date']),
                    'checked_name': self.read_cell(ws, row, self.punch_cols['checked_name']),
                    'checked_date': self.read_cell(ws, row, self.punch_cols['checked_date'])
                })

                row += 1

            wb.close()
            return punches
            
        except Exception as e:
            print(f"Error reading open punches: {e}")
            import traceback
            traceback.print_exc()
            return []


    # ================================================================
    # 3. UPDATED: review_checklist_before_save - With name and date updates
    # ================================================================

    def review_checklist_before_save(self, checklist_path, refs_set):
        """Modern dialog for reviewing and marking checklist items with name and date"""
        try:
            cols, matches = self.gather_checklist_matches(checklist_path, refs_set)
        except Exception as e:
            raise

        if not matches:
            messagebox.showinfo("Checklist Complete", 
                              "✓ No items requiring review.\nAll Interphase items are up to date.",
                              icon='info')
            return

        wb = load_workbook(checklist_path)
        ws = wb[self.interphase_sheet_name]
        
        # Extract all columns
        status_col = cols['status_col']
        date_col = cols['date_col']
        name_col = cols['name_col']
        remark_col = cols['remark_col']

        # Modern dialog window
        dlg = tk.Toplevel(self.root)
        dlg.title("Interphase Checklist Review")
        dlg.geometry("950x520")
        dlg.configure(bg='#f8fafc')
        dlg.transient(self.root)
        dlg.grab_set()
        
        # Header frame
        header_frame = tk.Frame(dlg, bg='#1e293b', height=60)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="📋 Interphase Checklist Review", 
                bg='#1e293b', fg='white', 
                font=('Segoe UI', 14, 'bold')).pack(pady=15)
        
        # Progress bar
        progress_frame = tk.Frame(dlg, bg='#f8fafc')
        progress_frame.pack(fill=tk.X, padx=20, pady=(15, 5))
        
        idx_label = tk.Label(progress_frame, text="", font=('Segoe UI', 11, 'bold'),
                            bg='#f8fafc', fg='#1e293b')
        idx_label.pack()
        
        # Content frame with modern styling
        content_frame = tk.Frame(dlg, bg='white', relief=tk.FLAT, borderwidth=0)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Reference info frame
        ref_frame = tk.Frame(content_frame, bg='#eff6ff', relief=tk.FLAT)
        ref_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        ref_label = tk.Label(ref_frame, text="", font=('Segoe UI', 10),
                            bg='#eff6ff', fg='#1e40af', anchor='w')
        ref_label.pack(fill=tk.X, padx=15, pady=10)
        
        # Description text
        tk.Label(content_frame, text="Description:", font=('Segoe UI', 9, 'bold'),
                bg='white', fg='#64748b', anchor='w').pack(fill=tk.X, padx=15, pady=(5, 2))
        
        text_widget = tk.Text(content_frame, wrap=tk.WORD, height=12, 
                             font=('Segoe UI', 10), bg='#f8fafc',
                             relief=tk.FLAT, borderwidth=1, padx=10, pady=10)
        text_widget.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))
        text_widget.config(state=tk.DISABLED)

        pos = [0]

        def show_item(p):
            r, ref_str, desc = matches[p]
            
            # Update progress
            progress_text = f"Item {p+1} of {len(matches)}"
            progress_pct = f"({int((p+1)/len(matches)*100)}% complete)"
            idx_label.config(text=f"{progress_text} {progress_pct}")
            
            # Update reference info
            ref_label.config(text=f"📌 Reference: {ref_str}  |  Row: {r}")
            
            # Update description
            text_widget.config(state=tk.NORMAL)
            text_widget.delete('1.0', tk.END)
            text_widget.insert(tk.END, desc)
            text_widget.config(state=tk.DISABLED)

        show_item(pos[0])

        def do_action_set_status(status_value):
            r, ref_str, desc = matches[pos[0]]
            current_date = datetime.now().strftime("%Y-%m-%d")
            
            # Get username
            try:
                username = os.getlogin()
            except:
                username = getpass.getuser()

            try:
                # Update status, name, and date
                self.write_cell(ws, r, status_col, status_value)
                self.write_cell(ws, r, name_col, username)
                self.write_cell(ws, r, date_col, current_date)
                wb.save(checklist_path)
            except PermissionError:
                messagebox.showerror("File Locked", 
                                   "⚠️ Please close the Excel file and try again.",
                                   icon='error')
                return
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update checklist:\n{e}")
                return

            if pos[0] < len(matches) - 1:
                pos[0] += 1
                show_item(pos[0])
            else:
                messagebox.showinfo("Review Complete", 
                                  f"✓ Checklist review finished!\n{len(matches)} items processed.",
                                  icon='info')
                dlg.destroy()

        def on_ok():
            do_action_set_status("OK")

        def on_nok():
            do_action_set_status("NOK")

        def on_na():
            """Handle N/A status with mandatory remark"""
            r, ref_str, desc = matches[pos[0]]
            
            # Mandatory remark dialog for N/A
            remark = simpledialog.askstring(
                "Remark Required",
                "N/A status requires a remark.\nPlease provide a reason:",
                parent=dlg
            )
            
            if not remark or not remark.strip():
                messagebox.showwarning(
                    "Remark Required",
                    "You must provide a remark for N/A status.",
                    parent=dlg
                )
                return
            
            current_date = datetime.now().strftime("%Y-%m-%d")
            
            try:
                # Get username
                try:
                    username = os.getlogin()
                except:
                    username = getpass.getuser()
                
                # Write all columns properly
                self.write_cell(ws, r, status_col, "N/A")
                self.write_cell(ws, r, date_col, current_date)
                self.write_cell(ws, r, name_col, username)
                self.write_cell(ws, r, remark_col, remark)
                wb.save(checklist_path)
                
                messagebox.showinfo("Remark Saved", 
                                  f"N/A status with remark:\n{remark}",
                                  parent=dlg)
            except PermissionError:
                messagebox.showerror("File Locked", 
                                   "⚠️ Please close the Excel file and try again.",
                                   icon='error',
                                   parent=dlg)
                return
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update checklist:\n{e}",
                                   parent=dlg)
                return

            if pos[0] < len(matches) - 1:
                pos[0] += 1
                show_item(pos[0])
            else:
                messagebox.showinfo("Review Complete", 
                                  f"✓ Checklist review finished!\n{len(matches)} items processed.",
                                  icon='info',
                                  parent=dlg)
                dlg.destroy()

        def on_prev():
            if pos[0] > 0:
                pos[0] -= 1
                show_item(pos[0])

        def on_next():
            if pos[0] < len(matches) - 1:
                pos[0] += 1
                show_item(pos[0])

        # Modern button frame
        btn_frame = tk.Frame(dlg, bg='#f8fafc')
        btn_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        btn_style = {
            'font': ('Segoe UI', 10, 'bold'),
            'relief': tk.FLAT,
            'borderwidth': 0,
            'cursor': 'hand2',
            'padx': 20,
            'pady': 12
        }

        tk.Button(btn_frame, text="◀ Previous", command=on_prev, bg='#94a3b8', 
                 fg='white', width=12, **btn_style).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="✓ OK", command=on_ok, bg='#10b981', 
                 fg='white', width=14, **btn_style).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="✗ NOK", command=on_nok, bg='#ef4444', 
                 fg='white', width=14, **btn_style).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="○ Not Applicable", command=on_na, bg='#f59e0b', 
                 fg='white', width=16, **btn_style).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Next ▶", command=on_next, bg='#94a3b8', 
                 fg='white', width=12, **btn_style).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Cancel", command=lambda: dlg.destroy(), 
                 bg='#64748b', fg='white', width=10, **btn_style).pack(side=tk.RIGHT, padx=5)

        dlg.wait_window()


    # ================================================================
    # 4. HELPER: gather_checklist_matches - Returns column info and matches
    # ================================================================

    def gather_checklist_matches(self, checklist_path, refs_set):
        """Returns Interphase rows where Reference No is NOT in refs_set."""
        wb = load_workbook(checklist_path)
        if self.interphase_sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError("Interphase sheet not found")

        ws = wb[self.interphase_sheet_name]
        ref_col = self.interphase_cols['ref_no']
        desc_col = self.interphase_cols['description']
        status_col = self.interphase_cols['status']
        name_col = self.interphase_cols['name']
        date_col = self.interphase_cols['date']
        remark_col = self.interphase_cols['remark']

        matches = []
        max_row = ws.max_row if ws.max_row else 2000

        for r in range(11, max_row + 1):
            ref_val = self.read_cell(ws, r, ref_col)
            if ref_val is None:
                continue

            ref_str = str(ref_val).strip()

            if ref_str in refs_set:
                continue

            status_val = self.read_cell(ws, r, status_col)
            status_str = str(status_val).strip().lower() if status_val is not None else ''

            if status_str in ('ok', 'nok', 'n/a', 'na', 'not applicable'):
                continue

            desc_val = self.read_cell(ws, r, desc_col) or ''
            matches.append((r, ref_str, str(desc_val)))

        wb.close()
        return {
            'ref_col': ref_col, 
            'desc_col': desc_col, 
            'status_col': status_col,
            'name_col': name_col,
            'date_col': date_col,
            'remark_col': remark_col
        }, matches
    # ================================================================
    # EXCEL HELPERS
    # ================================================================

    def save_interphase_excel(self):
        if not self.current_pdf_path:
            messagebox.showwarning("No PDF", "Load a PDF first.")
            return

        if not self.excel_file or not os.path.exists(self.excel_file):
            messagebox.showerror("Missing File", "Working Excel file not found.")
            return

        save_path = os.path.join(
            self.project_dirs["interphase_export"],
            f"{self.cabinet_id.replace(' ', '_')}_Interphase.xlsx"
        )

        try:
            shutil.copy2(self.excel_file, save_path)
            messagebox.showinfo("Saved", f"Final Interphase Excel saved:\n\n{save_path}")
        except PermissionError:
            messagebox.showerror("File Open", "Close the Excel file and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel:\n{e}")

    def open_excel(self):
        if not self.excel_file or not os.path.exists(self.excel_file):
            messagebox.showwarning("No Excel", "No working Excel file found.")
            return

        try:
            if os.name == 'nt':
                os.startfile(self.excel_file)
            else:
                if sys.platform == 'darwin':
                    cmd = f"open {shlex.quote(self.excel_file)}"
                else:
                    cmd = f"xdg-open {shlex.quote(self.excel_file)}"
                subprocess.Popen(cmd, shell=True)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Excel: {e}")

    # ================================================================
    # FUZZY MATCH HELPER
    # ================================================================

    def find_row_by_sr_or_text(self, sr_no, punch_text, min_ratio=0.60):
        try:
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
            row = 8

            while True:
                cell = self.read_cell(ws, row, self.punch_cols['sr_no'])
                if cell is None:
                    if self.read_cell(ws, row, self.punch_cols['desc']) is None:
                        break
                    else:
                        row += 1
                        continue
                try:
                    if int(cell) == int(sr_no):
                        wb.close()
                        return (row, 1.0, 'sr_exact')
                except:
                    if str(cell).strip() == str(sr_no).strip():
                        wb.close()
                        return (row, 1.0, 'sr_exact')
                row += 1
                if row > 2000:
                    break

            best_row = None
            best_ratio = 0.0
            row = 8

            while True:
                txt = self.read_cell(ws, row, self.punch_cols['desc'])
                if txt is None:
                    if row > 2000:
                        break
                    row += 1
                    continue
                try:
                    ratio = SequenceMatcher(None, str(punch_text).strip().lower(), str(txt).strip().lower()).ratio()
                except:
                    ratio = 0.0
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_row = row
                row += 1
                if row > 2000:
                    break

            wb.close()
            if best_row and best_ratio >= min_ratio:
                return (best_row, best_ratio, 'fuzzy_text')
            return (None, best_ratio, None)
        except Exception as e:
            try:
                wb.close()
            except:
                pass
            return (None, 0.0, None)

    # ================================================================
    # HANDOVER TO PRODUCTION
    # ================================================================
    def handover_to_production(self):
        """Handover current cabinet to production"""
        
        if not self.pdf_document or not self.excel_file:
            messagebox.showwarning("Incomplete", 
                                  "Please load a PDF and Excel file first.")
            return
        
        if not self.cabinet_id or not self.project_name:
            messagebox.showwarning("Missing Info", 
                                  "Project details are incomplete.")
            return
        
        # Check if there are any annotations
        if not self.annotations:
            proceed = messagebox.askyesno(
                "No Annotations",
                "No annotations found. Handover anyway?",
                icon='warning'
            )
            if not proceed:
                return
        
        # Count open punches
        open_punches = self.count_open_punches()
        
        if open_punches > 0:
            proceed = messagebox.askyesno(
                "Open Punches Detected",
                f"⚠️ There are {open_punches} open punch(es).\n\n"
                "Are you sure you want to handover to production?",
                icon='warning'
            )
            if not proceed:
                return
        
        # Save session before handover
        self.save_session()
        
        # Get user name
        try:
            username = os.getlogin()
        except:
            username = getpass.getuser()
        
        # Prepare handover data
        session_path = os.path.join(
            self.project_dirs.get("sessions", ""),
            f"{self.cabinet_id}_annotations.json"
        )
        
        handover_data = {
            "cabinet_id": self.cabinet_id,
            "project_name": self.project_name,
            "sales_order_no": self.sales_order_no,
            "pdf_path": self.current_pdf_path,
            "excel_path": self.excel_file,
            "session_path": session_path if os.path.exists(session_path) else None,
            "total_punches": len([a for a in self.annotations if a.get('type') == 'error']),
            "open_punches": open_punches,
            "closed_punches": len([a for a in self.annotations if a.get('type') == 'error']) - open_punches,
            "handed_over_by": username,
            "handed_over_date": datetime.now().isoformat()
        }
        
        # FIXED: Use handover_db instead of db
        success = self.handover_db.add_quality_handover(handover_data)
        
        # FIXED: Use correct method name
        self.update_status_and_sync('handed_to_production')
        
        if success:
            messagebox.showinfo("Handover Complete", 
                              "✓ Successfully handed over to Production\n\n"
                              f"Cabinet: {self.cabinet_id}\n"
                              f"Open Punches: {open_punches}")
        else:
            messagebox.showwarning("Already Handed Over", 
                                 "Cabinet already in production queue")

    def count_open_punches(self) -> int:
        """Count open punches in current Excel"""
        try:
            if not self.excel_file or not os.path.exists(self.excel_file):
                return 0
            
            wb = load_workbook(self.excel_file, data_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
            
            open_count = 0
            row = 8
            
            while row <= ws.max_row + 5:
                sr = self.read_cell(ws, row, self.punch_cols['sr_no'])
                if sr is None:
                    break
                
                closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
                if not closed:
                    open_count += 1
                
                row += 1
            
            wb.close()
            return open_count
            
        except Exception as e:
            print(f"Error counting open punches: {e}")
            return 0

    # ================================================================
    # VIEW PRODUCTION HANDBACK ITEMS
    # ================================================================

    def sync_manager_stats(self):
        """Sync current cabinet statistics to manager database WITHOUT changing status"""
        if not self.pdf_document or not self.cabinet_id:
            return
        
        try:
            # Count pages with annotations
            annotated_pages = len(set(ann['page'] for ann in self.annotations if ann.get('page') is not None))
            total_pages = len(self.pdf_document)
            
            # Count punches by type
            error_anns = [a for a in self.annotations if a.get('type') == 'error']
            total_punches = len(error_anns)
            
            # Count from Excel for accuracy
            open_punches = self.count_open_punches()
            
            # Count implemented (has implemented_name but no closed_name)
            implemented_punches = 0
            closed_punches = 0
            
            if self.excel_file and os.path.exists(self.excel_file):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(self.excel_file, data_only=True)
                    ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
                    
                    row = 8
                    while row <= ws.max_row + 5:
                        sr = self.read_cell(ws, row, self.punch_cols['sr_no'])
                        if sr is None:
                            break
                        
                        implemented = self.read_cell(ws, row, self.punch_cols['implemented_name'])
                        closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
                        
                        if closed:
                            closed_punches += 1
                        elif implemented:
                            implemented_punches += 1
                        
                        row += 1
                    
                    wb.close()
                except:
                    pass
            
            # FIXED: Get existing status from database, don't override it
            existing_status = self.get_current_status_from_db()
            
            # Update manager database with EXISTING status preserved
            self.manager_db.update_cabinet(
                self.cabinet_id,
                self.project_name,
                self.sales_order_no,
                total_pages,
                annotated_pages,
                total_punches,
                open_punches,
                implemented_punches,
                closed_punches,
                existing_status,  # CHANGED: Use existing status instead of hardcoded 'quality_inspection'
                storage_location=getattr(self, 'storage_location', None),
                excel_path=self.excel_file
            )
        except Exception as e:
            print(f"Manager sync error: {e}")
            import traceback
            traceback.print_exc()

            
    # ============================================================================
    # UPDATED: view_production_handbacks - Auto-open punch closing
    # ============================================================================

    def view_production_handbacks(self):
        """View and verify items returned from production"""
        
        pending_items = self.handover_db.get_pending_quality_items()
        
        if not pending_items:
            messagebox.showinfo(
                "No Items",
                "✓ No items pending verification from production.",
                icon='info'
            )
            return
        
        # Create modern dialog
        dlg = tk.Toplevel(self.root)
        dlg.title("Production Handback - Verification")
        dlg.geometry("1000x600")
        dlg.configure(bg='#f8fafc')
        dlg.transient(self.root)
        dlg.grab_set()
        
        # Header
        header_frame = tk.Frame(dlg, bg='#7c3aed', height=60)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="🔍 Production Handback Verification", 
                bg='#7c3aed', fg='white', 
                font=('Segoe UI', 14, 'bold')).pack(pady=15)
        
        # Info bar
        info_frame = tk.Frame(dlg, bg='#eff6ff')
        info_frame.pack(fill=tk.X, padx=20, pady=(15, 5))
        
        tk.Label(info_frame, text=f"Total items pending verification: {len(pending_items)}", 
                bg='#eff6ff', fg='#1e40af', 
                font=('Segoe UI', 10, 'bold')).pack(pady=8)
        
        # Listbox frame
        list_frame = tk.Frame(dlg, bg='white')
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        tk.Label(list_frame, text="Select item to verify:", 
                font=('Segoe UI', 10, 'bold'),
                bg='white', fg='#1e293b').pack(anchor='w', pady=(0, 10))
        
        # Scrollbar and Listbox
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        listbox = tk.Listbox(list_frame, font=('Consolas', 9),
                            yscrollcommand=scrollbar.set,
                            bg='#f8fafc', relief=tk.FLAT,
                            selectmode=tk.SINGLE, height=15)
        listbox.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=listbox.yview)
        
        # Populate listbox
        for item in pending_items:
            display_text = (
                f"{item['cabinet_id']:20} | {item['project_name']:30} | "
                f"Rework by: {item['rework_completed_by']:15} | "
                f"{item['rework_completed_date'][:10]}"
            )
            listbox.insert(tk.END, display_text)
        
        def load_selected():
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select an item first.")
                return
            
            item = pending_items[selection[0]]
            
            # Load the project
            try:
                project_data = self.db.get_project(item['cabinet_id'])
                if not project_data:
                    messagebox.showerror("Error", "Project not found in database")
                    return
                
                self.cabinet_id = item['cabinet_id']
                self.project_name = item['project_name']
                self.sales_order_no = item['sales_order_no']
                self.storage_location = project_data['storage_location']
                
                self.prepare_project_folders()
                
                if not os.path.exists(item['pdf_path']):
                    messagebox.showerror("Error", f"PDF file not found:\n{item['pdf_path']}")
                    return
                
                self.pdf_document = fitz.open(item['pdf_path'])
                self.current_pdf_path = item['pdf_path']
                self.current_page = 0
                self.zoom_level = 1.0
                self.tool_mode = None
                self.root.config(cursor="")
                
                self.excel_file = item['excel_path']
                self.working_excel_path = item['excel_path']
                
                self.current_sr_no = self.get_next_sr_no()
                
                if item.get('session_path') and os.path.exists(item['session_path']):
                    self.load_session_from_path(item['session_path'])
                else:
                    self.annotations = []
                    self.display_page()
                
                # UPDATED: Set status to "Rework being verified"
                self.update_status_and_sync('being_closed_by_quality')
                
                dlg.destroy()
                
                # UPDATED: Auto-open punch closing dialog
                self.verify_production_work_with_punch_closing(item)
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load item:\n{e}")
                import traceback
                traceback.print_exc()
        
        # Buttons
        btn_frame = tk.Frame(dlg, bg='#f8fafc')
        btn_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        btn_style = {
            'font': ('Segoe UI', 10, 'bold'),
            'relief': tk.FLAT,
            'cursor': 'hand2',
            'padx': 20,
            'pady': 12
        }
        
        tk.Button(btn_frame, text="📂 Load & Verify", command=load_selected,
                 bg='#3b82f6', fg='white', **btn_style).pack(side=tk.LEFT, padx=5)
        
        # REMOVED: Quick Verify button as requested
        
        tk.Button(btn_frame, text="Cancel", command=dlg.destroy,
                 bg='#64748b', fg='white', **btn_style).pack(side=tk.RIGHT, padx=5)
        
        listbox.bind('<Double-Button-1>', lambda e: load_selected())


    # ============================================================================
    # NEW: verify_production_work_with_punch_closing - Auto-open punch closing
    # ============================================================================

    def verify_production_work_with_punch_closing(self, item_data):
        """Auto-open punch closing mode after loading production item"""
        
        # Count open punches
        open_count = self.count_open_punches()
        
        # Show initial info
        info_msg = (
            f"Cabinet: {item_data['cabinet_id']}\n"
            f"Project: {item_data['project_name']}\n\n"
            f"Rework by: {item_data['rework_completed_by']}\n"
            f"Date: {item_data['rework_completed_date'][:10]}\n\n"
            f"Open Punches: {open_count}\n\n"
            f"Opening Punch Closing Mode..."
        )
        
        messagebox.showinfo("Production Handback Loaded", info_msg, icon='info')
        
        # Auto-open punch closing mode
        if open_count > 0:
            self.punch_closing_mode_for_verification(item_data)
        else:
            # No punches - directly ask to close
            self.finalize_verification(item_data)


    # ============================================================================
    # NEW: punch_closing_mode_for_verification - Modified punch closing for handback
    # ============================================================================

    def punch_closing_mode_for_verification(self, item_data):
        """Punch closing mode specifically for verification workflow - Converts orange to green"""
        
        punches = self.read_open_punches_from_excel()
        
        if not punches:
            # All closed, proceed to finalization
            self.finalize_verification(item_data)
            return
        
        punches.sort(key=lambda p: (not p['implemented'], p['sr_no']))
        
        # Dialog
        dlg = tk.Toplevel(self.root)
        dlg.title("Punch Verification Mode")
        dlg.geometry("950x600")
        dlg.configure(bg='#f8fafc')
        dlg.transient(self.root)
        dlg.grab_set()
        
        # Header
        header_frame = tk.Frame(dlg, bg='#7c3aed', height=50)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="✓ Punch Verification Mode", 
                bg='#7c3aed', fg='white', 
                font=('Segoe UI', 13, 'bold')).pack(pady=12)
        
        # Progress
        progress_frame = tk.Frame(dlg, bg='#f8fafc')
        progress_frame.pack(fill=tk.X, padx=20, pady=(10, 5))
        
        idx_label = tk.Label(progress_frame, text="", font=('Segoe UI', 10, 'bold'),
                            bg='#f8fafc', fg='#1e293b')
        idx_label.pack()
        
        # Info cards
        info_frame = tk.Frame(dlg, bg='#f8fafc')
        info_frame.pack(fill=tk.X, padx=20, pady=8)
        
        sr_card = tk.Frame(info_frame, bg='#dbeafe', relief=tk.FLAT)
        sr_card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        tk.Label(sr_card, text="SR No.", font=('Segoe UI', 8), 
                bg='#dbeafe', fg='#1e40af').pack(anchor='w', padx=10, pady=(6, 2))
        sr_label = tk.Label(sr_card, text="", font=('Segoe UI', 12, 'bold'),
                           bg='#dbeafe', fg='#1e293b')
        sr_label.pack(anchor='w', padx=10, pady=(0, 6))
        
        ref_card = tk.Frame(info_frame, bg='#e0e7ff', relief=tk.FLAT)
        ref_card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        tk.Label(ref_card, text="Reference", font=('Segoe UI', 8), 
                bg='#e0e7ff', fg='#4338ca').pack(anchor='w', padx=10, pady=(6, 2))
        ref_label = tk.Label(ref_card, text="", font=('Segoe UI', 12, 'bold'),
                            bg='#e0e7ff', fg='#1e293b')
        ref_label.pack(anchor='w', padx=10, pady=(0, 6))
        
        status_card = tk.Frame(info_frame, bg='#fef3c7', relief=tk.FLAT)
        status_card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
        
        tk.Label(status_card, text="Status", font=('Segoe UI', 8), 
                bg='#fef3c7', fg='#92400e').pack(anchor='w', padx=10, pady=(6, 2))
        impl_label = tk.Label(status_card, text="", font=('Segoe UI', 12, 'bold'),
                             bg='#fef3c7', fg='#1e293b')
        impl_label.pack(anchor='w', padx=10, pady=(0, 6))
        
        # Content
        content_frame = tk.Frame(dlg, bg='white', relief=tk.FLAT)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=8)
        
        tk.Label(content_frame, text="Punch Description:", font=('Segoe UI', 9, 'bold'),
                bg='white', fg='#64748b', anchor='w').pack(fill=tk.X, padx=15, pady=(8, 3))
        
        text_widget = tk.Text(content_frame, wrap=tk.WORD, height=9,
                             font=('Segoe UI', 10), bg='#f8fafc',
                             relief=tk.FLAT, padx=10, pady=8)
        text_widget.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 10))
        text_widget.config(state=tk.DISABLED)

        pos = [0]

        def show_item():
            p = punches[pos[0]]
            
            progress_text = f"Item {pos[0]+1} of {len(punches)}"
            progress_pct = f"({int((pos[0]+1)/len(punches)*100)}% complete)"
            idx_label.config(text=f"{progress_text} {progress_pct}")
            
            sr_label.config(text=str(p['sr_no']))
            ref_label.config(text=str(p['ref_no']))
            
            impl_status = "✓ Implemented" if p['implemented'] else "⚠ Not Implemented"
            impl_color = '#10b981' if p['implemented'] else '#f59e0b'
            impl_label.config(text=impl_status, fg=impl_color)
            
            text_widget.config(state=tk.NORMAL)
            text_widget.delete("1.0", tk.END)
            text_widget.insert(tk.END, p['punch_text'])
            text_widget.insert(tk.END, f"\n\n──────────────────\n")
            text_widget.insert(tk.END, f"Category: {p['category']}\n")

            # UPDATED: Find annotation - checks for both SR number and excel row
            ann = next((a for a in self.annotations 
                       if a.get('sr_no') == p['sr_no'] 
                       or (a.get('excel_row') == p['row'])), None)
            
            if ann and ann.get('implementation_remark'):
                text_widget.insert(tk.END, f"\n──────────────────\n")
                text_widget.insert(tk.END, "Implementation Remarks:\n")
                text_widget.insert(tk.END, ann['implementation_remark'])
            
            # Show production remarks
            try:
                handover_data = self.handover_db.get_handover_by_cabinet(self.cabinet_id)
                if handover_data and handover_data.get('production_remarks'):
                    text_widget.insert(tk.END, f"\n\n──────────────────\n")
                    text_widget.insert(tk.END, "🔧 Production Remarks:\n")
                    text_widget.insert(tk.END, handover_data['production_remarks'])
                    text_widget.insert(tk.END, f"\n\nRework by: {handover_data.get('rework_completed_by', 'N/A')}")
                    text_widget.insert(tk.END, f"\nDate: {handover_data.get('rework_completed_date', 'N/A')[:10]}")
            except Exception as e:
                print(f"Could not load production remarks: {e}")

            text_widget.config(state=tk.DISABLED)

        show_item()

        def close_punch():
            p = punches[pos[0]]

            try:
                default_user = os.getlogin()
            except:
                default_user = getpass.getuser()

            name = simpledialog.askstring("Closed By", 
                                         "Enter your name to close this punch:", 
                                         initialvalue=default_user, 
                                         parent=dlg)
            if not name:
                return

            try:
                wb = load_workbook(self.excel_file)
                ws = wb[self.punch_sheet_name]

                self.write_cell(ws, p['row'], self.punch_cols['closed_name'], name)
                self.write_cell(ws, p['row'], self.punch_cols['closed_date'], 
                              datetime.now().strftime("%Y-%m-%d"))

                wb.save(self.excel_file)
                wb.close()

            except PermissionError:
                messagebox.showerror("File Locked", 
                                   "⚠️ Please close the Excel file and try again.")
                return
            except Exception as e:
                messagebox.showerror("Error", f"Failed to close punch:\n{e}")
                return

            # ============================================================
            # UPDATED: Find and convert annotation color from orange to green
            # ============================================================
            ann = next((a for a in self.annotations 
                       if a.get('sr_no') == p['sr_no'] 
                       or (a.get('excel_row') == p['row'])), None)
            
            if ann:
                # ✅ Convert orange highlight to green (KEEP the annotation)
                if ann.get('type') == 'highlight' and ann.get('color') == 'orange':
                    ann['color'] = 'green'  # Change from error to OK
                    print(f"✓ Verification: Converted annotation to green for SR {p['sr_no']}")
                
                # ✅ Also handle old rectangle-style error annotations
                elif ann.get('type') == 'error':
                    ann['type'] = 'ok'  # Convert error rectangle to OK
                    print(f"✓ Verification: Converted error rectangle to OK for SR {p['sr_no']}")
                
                # ✅ Store closure information
                ann['closed_by'] = name
                ann['closed_date'] = datetime.now().strftime("%Y-%m-%d")
            else:
                print(f"⚠️ Warning: No annotation found for SR {p['sr_no']} (Row {p['row']})")

            # ✅ Refresh display to show green highlight
            self.display_page()
            
            # ✅ Update stats after closing
            self.sync_manager_stats_only()

            if pos[0] < len(punches) - 1:
                pos[0] += 1
                show_item()
            else:
                messagebox.showinfo("All Punches Closed", 
                                  f"✓ All punches verified and closed!\n{len(punches)} items processed.",
                                  icon='info')
                dlg.destroy()
                # Proceed to finalization
                self.finalize_verification(item_data)

        def next_item():
            if pos[0] < len(punches) - 1:
                pos[0] += 1
                show_item()

        def prev_item():
            if pos[0] > 0:
                pos[0] -= 1
                show_item()

        # Buttons
        btn_frame = tk.Frame(dlg, bg='#f8fafc', height=80)
        btn_frame.pack(fill=tk.X, padx=20, pady=(10, 25))
        btn_frame.pack_propagate(False)
        
        btn_container = tk.Frame(btn_frame, bg='#f8fafc')
        btn_container.pack(expand=True)
        
        btn_style = {
            'font': ('Segoe UI', 12, 'bold'),
            'relief': tk.FLAT,
            'borderwidth': 0,
            'cursor': 'hand2',
            'padx': 35,
            'pady': 18,
            'width': 15
        }

        tk.Button(btn_container, text="◀  Previous", command=prev_item, 
                 bg='#94a3b8', fg='white', **btn_style).pack(side=tk.LEFT, padx=8)
        
        close_btn_style = btn_style.copy()
        close_btn_style['width'] = 18
        tk.Button(btn_container, text="✓  CLOSE PUNCH", command=close_punch, 
                 bg='#10b981', fg='white', **close_btn_style).pack(side=tk.LEFT, padx=8)
        
        tk.Button(btn_container, text="Next  ▶", command=next_item, 
                 bg='#94a3b8', fg='white', **btn_style).pack(side=tk.LEFT, padx=8)
        
        tk.Button(btn_container, text="Cancel", command=dlg.destroy, 
                 bg='#64748b', fg='white', **btn_style).pack(side=tk.LEFT, padx=8)

        dlg.wait_window()


    # ============================================================================
    # NEW: finalize_verification - Check checklist, save Excel, export PDF
    # ============================================================================

    def finalize_verification(self, item_data):
        """Final step: check checklist completeness, save Excel, export PDF"""
        
        # 1. Check checklist completeness
        checklist_complete, incomplete_refs = self.check_checklist_completeness()
        
        if not checklist_complete:
            warning_msg = (
                f"⚠️ Checklist Incomplete!\n\n"
                f"{len(incomplete_refs)} reference(s) without status:\n"
                f"{', '.join(incomplete_refs[:10])}"
            )
            if len(incomplete_refs) > 10:
                warning_msg += f"\n... and {len(incomplete_refs) - 10} more"
            
            warning_msg += "\n\nDo you want to review the checklist now?"
            
            review = messagebox.askyesno("Incomplete Checklist", warning_msg, icon='warning')
            if review:
                self.review_checklist_now()
                # After review, ask again if they want to finalize
                retry = messagebox.askyesno(
                    "Continue?",
                    "Review complete. Proceed with closing the cabinet?",
                    icon='question'
                )
                if not retry:
                    return
                
                # Re-check completeness after review
                checklist_complete, incomplete_refs = self.check_checklist_completeness()
                if not checklist_complete:
                    messagebox.showwarning(
                        "Still Incomplete",
                        f"Checklist still has {len(incomplete_refs)} incomplete item(s).\n"
                        "Proceeding anyway...",
                        icon='warning'
                    )
            else:
                # User chose not to review - warn but allow to continue
                proceed = messagebox.askyesno(
                    "Continue Anyway?",
                    "Checklist is incomplete. Continue with closing?",
                    icon='warning'
                )
                if not proceed:
                    return
        
        # 2. Confirm closing
        confirm_msg = (
            f"Cabinet: {item_data['cabinet_id']}\n"
            f"Project: {item_data['project_name']}\n\n"
            f"✓ All punches closed\n"
            f"{'✓' if checklist_complete else '⚠'} Checklist {'complete' if checklist_complete else 'reviewed'}\n\n"
            f"This will:\n"
            f"• Save Interphase Excel\n"
            f"• Export Annotated PDF\n"
            f"• Mark cabinet as CLOSED\n\n"
            f"Proceed?"
        )
        
        confirm = messagebox.askyesno("Close Cabinet", confirm_msg, icon='question')
        if not confirm:
            return
        
        # 3. Save current session first
        try:
            self.save_session()
        except Exception as e:
            print(f"Session save warning: {e}")
        
        # 4. Auto-save Interphase Excel
        try:
            if not self.excel_file or not os.path.exists(self.excel_file):
                messagebox.showerror("Error", "Working Excel file not found.")
                return
            
            save_path = os.path.join(
                self.project_dirs["interphase_export"],
                f"{self.cabinet_id.replace(' ', '_')}_Interphase.xlsx"
            )
            
            shutil.copy2(self.excel_file, save_path)
            print(f"✓ Interphase Excel saved: {save_path}")
            
        except PermissionError:
            messagebox.showerror("Error", "Excel file is open. Please close it and try again.")
            return
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel:\n{e}")
            return
        
        # 5. Auto-export annotated PDF
        try:
            if not self.pdf_document:
                messagebox.showerror("Error", "No PDF document loaded.")
                return
            
            export_path = os.path.join(
                self.project_dirs["annotated_drawings"],
                f"{self.cabinet_id.replace(' ', '_')}_Annotated.pdf"
            )
            
            # Create output PDF
            out_doc = fitz.open()
            for pnum in range(len(self.pdf_document)):
                out_doc.insert_pdf(self.pdf_document, from_page=pnum, to_page=pnum)
            
            # Open Excel for SR No lookup
            wb = None
            ws = None
            if self.excel_file and os.path.exists(self.excel_file):
                try:
                    wb = load_workbook(self.excel_file, data_only=True)
                    ws = wb[self.punch_sheet_name]
                except:
                    pass
            
            # Draw annotations
            for ann in self.annotations:
                p = ann.get('page')
                if p is None or p < 0 or p >= len(out_doc):
                    continue
                
                target_page = out_doc[p]
                ann_type = ann.get('type')
                
                # Rectangle annotations
                if ann_type in ('ok', 'error') and 'bbox_page' in ann:
                    x1, y1, x2, y2 = ann['bbox_page']
                    rect = self.transform_bbox_for_rotation((x1, y1, x2, y2), target_page)
                    
                    if ann_type == 'ok':
                        target_page.draw_rect(rect, color=(0, 1, 0), width=2)
                    else:
                        target_page.draw_rect(rect, color=(1, 0.55, 0), width=2)
                    
                    sr_text = None
                    row = ann.get('excel_row')
                    
                    if row and ws:
                        try:
                            sr_val = self.read_cell(ws, row, self.punch_cols['sr_no'])
                            if sr_val is not None:
                                sr_text = f"Sr {sr_val}"
                        except:
                            sr_text = None
                    
                    if sr_text:
                        text_pos = fitz.Point(rect.x0, max(rect.y0 - 12, rect.y0))
                        try:
                            target_page.insert_text(text_pos, sr_text, fontsize=8)
                        except:
                            pass
                
                # Pen strokes
                elif ann_type == 'pen' and 'points' in ann:
                    points = ann['points']
                    if len(points) >= 2:
                        transformed_points = [
                            self.transform_point_for_rotation(pt, target_page) 
                            for pt in points
                        ]
                        
                        for i in range(len(transformed_points) - 1):
                            p1 = transformed_points[i]
                            p2 = transformed_points[i + 1]
                            target_page.draw_line(p1, p2, color=(1, 0, 0), width=2)
                
                # Text annotations
                elif ann_type == 'text' and 'pos_page' in ann:
                    pos = ann['pos_page']
                    text = ann.get('text', '')
                    if text:
                        text_point = self.transform_point_for_rotation(pos, target_page)
                        try:
                            target_page.insert_text(
                                text_point, text,
                                fontsize=10, color=(1, 0, 0),
                                rotate=target_page.rotation
                            )
                        except:
                            pass
            
            if wb:
                wb.close()
            
            out_doc.save(export_path)
            out_doc.close()
            
            print(f"✓ Annotated PDF exported: {export_path}")
            
        except PermissionError:
            messagebox.showerror("Error", "PDF file is open. Please close it and try again.")
            return
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export PDF:\n{e}")
            return
        
        # 6. Update verification status
        try:
            username = os.getlogin()
        except:
            username = getpass.getuser()
        
        success = self.handover_db.update_quality_verification(
            item_data['cabinet_id'],
            status="closed",
            user=username
        )
        
        if success:
            # Update manager status to 'closed'
            self.update_status_and_sync('closed')
            
            messagebox.showinfo(
                "✓ Verification Complete",
                f"Cabinet {item_data['cabinet_id']} successfully closed!\n\n"
                f"✓ Interphase Excel saved\n"
                f"✓ Annotated PDF exported\n"
                f"✓ Status updated to CLOSED\n\n"
                f"Files saved to:\n{self.project_dirs['root']}",
                icon='info'
            )
        else:
            messagebox.showerror("Error", "Failed to update verification status.")


    # ============================================================================
    # NEW: check_checklist_completeness - Helper to check Interphase completion
    # ============================================================================

    def check_checklist_completeness(self):
        """Check if all Interphase items have status (OK/NOK/N/A)
        Returns: (is_complete: bool, incomplete_refs: list)
        """
        if not self.excel_file or not os.path.exists(self.excel_file):
            return False, []
        
        try:
            wb = load_workbook(self.excel_file, data_only=True)
            if self.interphase_sheet_name not in wb.sheetnames:
                wb.close()
                return False, []
            
            ws = wb[self.interphase_sheet_name]
            ref_col = self.interphase_cols['ref_no']
            status_col = self.interphase_cols['status']
            
            incomplete_refs = []
            max_row = ws.max_row if ws.max_row else 2000
            
            for r in range(11, max_row + 1):
                ref_val = self.read_cell(ws, r, ref_col)
                if ref_val is None:
                    continue
                
                ref_str = str(ref_val).strip()
                if not ref_str:
                    continue
                
                status_val = self.read_cell(ws, r, status_col)
                status_str = str(status_val).strip().lower() if status_val is not None else ''
                
                # Check if status is empty or invalid
                if status_str not in ('ok', 'nok', 'n/a', 'na', 'not applicable'):
                    incomplete_refs.append(ref_str)
            
            wb.close()
            
            is_complete = len(incomplete_refs) == 0
            return is_complete, incomplete_refs
            
        except Exception as e:
            print(f"Checklist check error: {e}")
            return False, []


    def auto_save_session(self):
        """Auto-save session every 5 minutes"""
        if self.pdf_document and hasattr(self, 'project_dirs'):
            try:
                self.save_session()
                print(f"✓ Auto-saved at {datetime.now().strftime('%H:%M:%S')}")
            except Exception as e:
                print(f"⚠️ Auto-save failed: {e}")
        
        self.root.after(300000, self.auto_save_session)

    def count_open_punches(self):
        """Count open punches"""
        try:
            if not self.excel_file or not os.path.exists(self.excel_file):
                return 0
            
            wb = load_workbook(self.excel_file, data_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
            
            open_count = 0
            row = 8
            
            while row <= ws.max_row + 5:
                sr = self.read_cell(ws, row, self.punch_cols['sr_no'])
                if sr is None:
                    break
                
                closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
                if not closed:
                    open_count += 1
                
                row += 1
            
            wb.close()
            return open_count
            
        except Exception as e:
            print(f"Error counting open punches: {e}")
            return 0
    def punch_closing_mode(self):
        """Modern dialog for punch closing workflow - Converts orange to green highlights"""
        punches = self.read_open_punches_from_excel()

        if not punches:
            messagebox.showinfo("No Open Punches", 
                              "✓ All punches are closed!\nNo items require attention.",
                              icon='info')
            return

        punches.sort(key=lambda p: (not p['implemented'], p['sr_no']))

        # Modern dialog window
        dlg = tk.Toplevel(self.root)
        dlg.title("Punch Closing Mode")
        dlg.geometry("950x600")
        dlg.configure(bg='#f8fafc')
        dlg.transient(self.root)
        dlg.grab_set()
        
        # Header
        header_frame = tk.Frame(dlg, bg='#1e293b', height=50)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="✓ Punch Closing Mode", 
                bg='#1e293b', fg='white', 
                font=('Segoe UI', 13, 'bold')).pack(pady=12)
        
        # Progress
        progress_frame = tk.Frame(dlg, bg='#f8fafc')
        progress_frame.pack(fill=tk.X, padx=20, pady=(10, 5))
        
        idx_label = tk.Label(progress_frame, text="", font=('Segoe UI', 10, 'bold'),
                            bg='#f8fafc', fg='#1e293b')
        idx_label.pack()
        
        # Info cards
        info_frame = tk.Frame(dlg, bg='#f8fafc')
        info_frame.pack(fill=tk.X, padx=20, pady=8)
        
        # SR Number card
        sr_card = tk.Frame(info_frame, bg='#dbeafe', relief=tk.FLAT)
        sr_card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        tk.Label(sr_card, text="SR No.", font=('Segoe UI', 8), 
                bg='#dbeafe', fg='#1e40af').pack(anchor='w', padx=10, pady=(6, 2))
        sr_label = tk.Label(sr_card, text="", font=('Segoe UI', 12, 'bold'),
                           bg='#dbeafe', fg='#1e293b')
        sr_label.pack(anchor='w', padx=10, pady=(0, 6))
        
        # Reference card
        ref_card = tk.Frame(info_frame, bg='#e0e7ff', relief=tk.FLAT)
        ref_card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        tk.Label(ref_card, text="Reference", font=('Segoe UI', 8), 
                bg='#e0e7ff', fg='#4338ca').pack(anchor='w', padx=10, pady=(6, 2))
        ref_label = tk.Label(ref_card, text="", font=('Segoe UI', 12, 'bold'),
                            bg='#e0e7ff', fg='#1e293b')
        ref_label.pack(anchor='w', padx=10, pady=(0, 6))
        
        # Status card
        status_card = tk.Frame(info_frame, bg='#fef3c7', relief=tk.FLAT)
        status_card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
        
        tk.Label(status_card, text="Status", font=('Segoe UI', 8), 
                bg='#fef3c7', fg='#92400e').pack(anchor='w', padx=10, pady=(6, 2))
        impl_label = tk.Label(status_card, text="", font=('Segoe UI', 12, 'bold'),
                             bg='#fef3c7', fg='#1e293b')
        impl_label.pack(anchor='w', padx=10, pady=(0, 6))
        
        # Content
        content_frame = tk.Frame(dlg, bg='white', relief=tk.FLAT)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=8)
        
        tk.Label(content_frame, text="Punch Description:", font=('Segoe UI', 9, 'bold'),
                bg='white', fg='#64748b', anchor='w').pack(fill=tk.X, padx=15, pady=(8, 3))
        
        text_widget = tk.Text(content_frame, wrap=tk.WORD, height=9,
                             font=('Segoe UI', 10), bg='#f8fafc',
                             relief=tk.FLAT, padx=10, pady=8)
        text_widget.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 10))
        text_widget.config(state=tk.DISABLED)

        pos = [0]

        def show_item():
            p = punches[pos[0]]
            
            # Update progress
            progress_text = f"Item {pos[0]+1} of {len(punches)}"
            progress_pct = f"({int((pos[0]+1)/len(punches)*100)}% complete)"
            idx_label.config(text=f"{progress_text} {progress_pct}")
            
            # Update info cards
            sr_label.config(text=str(p['sr_no']))
            ref_label.config(text=str(p['ref_no']))
            
            impl_status = "✓ Implemented" if p['implemented'] else "⚠ Not Implemented"
            impl_color = '#10b981' if p['implemented'] else '#f59e0b'
            impl_label.config(text=impl_status, fg=impl_color)
            
            # Update description
            text_widget.config(state=tk.NORMAL)
            text_widget.delete("1.0", tk.END)
            text_widget.insert(tk.END, p['punch_text'])
            text_widget.insert(tk.END, f"\n\n──────────────────\n")
            text_widget.insert(tk.END, f"Category: {p['category']}\n")

            # Find annotation - checks for highlight type and old rectangle type
            ann = next((a for a in self.annotations 
                       if a.get('sr_no') == p['sr_no'] 
                       or (a.get('excel_row') == p['row'])), None)
            
            if ann and ann.get('implementation_remark'):
                text_widget.insert(tk.END, f"\n──────────────────\n")
                text_widget.insert(tk.END, "Implementation Remarks:\n")
                text_widget.insert(tk.END, ann['implementation_remark'])
            
            # Check for production remarks
            try:
                handover_data = self.handover_db.get_handover_by_cabinet(self.cabinet_id)
                if handover_data and handover_data.get('production_remarks'):
                    text_widget.insert(tk.END, f"\n\n──────────────────\n")
                    text_widget.insert(tk.END, "🔧 Production Remarks:\n")
                    text_widget.insert(tk.END, handover_data['production_remarks'])
                    text_widget.insert(tk.END, f"\n\nRework by: {handover_data.get('rework_completed_by', 'N/A')}")
                    text_widget.insert(tk.END, f"\nDate: {handover_data.get('rework_completed_date', 'N/A')[:10]}")
            except Exception as e:
                print(f"Could not load production remarks: {e}")

            text_widget.config(state=tk.DISABLED)

        show_item()

        def close_punch():
            p = punches[pos[0]]

            try:
                default_user = os.getlogin()
            except:
                default_user = getpass.getuser()

            name = simpledialog.askstring("Closed By", 
                                         "Enter your name to close this punch:", 
                                         initialvalue=default_user, 
                                         parent=dlg)
            if not name:
                return

            try:
                wb = load_workbook(self.excel_file)
                ws = wb[self.punch_sheet_name]

                self.write_cell(ws, p['row'], self.punch_cols['closed_name'], name)
                self.write_cell(ws, p['row'], self.punch_cols['closed_date'], 
                              datetime.now().strftime("%Y-%m-%d"))

                wb.save(self.excel_file)
                wb.close()

            except PermissionError:
                messagebox.showerror("File Locked", 
                                   "⚠️ Please close the Excel file and try again.")
                return
            except Exception as e:
                messagebox.showerror("Error", f"Failed to close punch:\n{e}")
                return

            # ============================================================
            # UPDATED: Find and convert annotation color from orange to green
            # ============================================================
            ann = next((a for a in self.annotations 
                       if a.get('sr_no') == p['sr_no'] 
                       or (a.get('excel_row') == p['row'])), None)
            
            if ann:
                # ✅ Convert orange highlight to green (KEEP the annotation)
                if ann.get('type') == 'highlight' and ann.get('color') == 'orange':
                    ann['color'] = 'green'  # Change from error to OK
                    print(f"✓ Converted annotation to green for SR {p['sr_no']}")
                
                # ✅ Also handle old rectangle-style error annotations
                elif ann.get('type') == 'error':
                    ann['type'] = 'ok'  # Convert error rectangle to OK
                    print(f"✓ Converted error rectangle to OK for SR {p['sr_no']}")
                
                # ✅ Store closure information
                ann['closed_by'] = name
                ann['closed_date'] = datetime.now().strftime("%Y-%m-%d")
            else:
                print(f"⚠️ Warning: No annotation found for SR {p['sr_no']}")

            # ✅ Refresh display to show green highlight
            self.display_page()
            
            # ✅ Update stats after closing
            self.sync_manager_stats_only()

            if pos[0] < len(punches) - 1:
                pos[0] += 1
                show_item()
            else:
                messagebox.showinfo("Complete", 
                                  f"✓ All punches closed!\n{len(punches)} items processed.",
                                  icon='info')
                dlg.destroy()

        def next_item():
            if pos[0] < len(punches) - 1:
                pos[0] += 1
                show_item()

        def prev_item():
            if pos[0] > 0:
                pos[0] -= 1
                show_item()

        # Buttons
        btn_frame = tk.Frame(dlg, bg='#f8fafc', height=80)
        btn_frame.pack(fill=tk.X, padx=20, pady=(10, 25))
        btn_frame.pack_propagate(False)
        
        btn_container = tk.Frame(btn_frame, bg='#f8fafc')
        btn_container.pack(expand=True)
        
        btn_style = {
            'font': ('Segoe UI', 12, 'bold'),
            'relief': tk.FLAT,
            'borderwidth': 0,
            'cursor': 'hand2',
            'padx': 35,
            'pady': 18,
            'width': 15
        }

        tk.Button(btn_container, text="◀  Previous", command=prev_item, 
                 bg='#94a3b8', fg='white', **btn_style).pack(side=tk.LEFT, padx=8)
        
        close_btn_style = btn_style.copy()
        close_btn_style['width'] = 18
        tk.Button(btn_container, text="✓  CLOSE PUNCH", command=close_punch, 
                 bg='#10b981', fg='white', **close_btn_style).pack(side=tk.LEFT, padx=8)
        
        tk.Button(btn_container, text="Next  ▶", command=next_item, 
                 bg='#94a3b8', fg='white', **btn_style).pack(side=tk.LEFT, padx=8)
        
        tk.Button(btn_container, text="Cancel", command=dlg.destroy, 
                 bg='#64748b', fg='white', **btn_style).pack(side=tk.LEFT, padx=8)

        dlg.wait_window()

    def save_recent_project(self):
        """Save current project to database with storage location - HIGHLIGHTER VERSION"""
        if not self.current_pdf_path or not self.excel_file:
            return
        
        try:
            session_path = os.path.join(
                self.project_dirs.get("sessions", ""),
                f"{self.cabinet_id}_annotations.json"
            ) if hasattr(self, 'project_dirs') else None
            
            project_data = {
                'cabinet_id': self.cabinet_id,
                'project_name': self.project_name,
                'sales_order_no': self.sales_order_no,
                'storage_location': self.storage_location,
                'pdf_path': self.current_pdf_path,
                'excel_path': self.excel_file,
                'session_path': session_path if session_path and os.path.exists(session_path) else None,
                'last_accessed': datetime.now().isoformat()
            }
            
            if self.db.project_exists(self.cabinet_id):
                self.db.update_project(self.cabinet_id, project_data)
            else:
                project_data['created_date'] = datetime.now().isoformat()
                self.db.add_project(project_data)
            
            self.update_recent_dropdown()
            self.sync_manager_stats_only()
            
        except Exception as e:
            print(f"Error saving recent project: {e}")


    def load_recent_projects_ui(self):
        """Load and display recent projects from SQLite - HIGHLIGHTER VERSION"""
        self.update_recent_dropdown()


    def update_recent_dropdown(self):
        """Update the recent projects dropdown from database"""
        try:
            recent_projects = self.db.get_recent_projects(limit=20)
            
            menu = self.recent_dropdown['menu']
            menu.delete(0, 'end')
            
            if not recent_projects:
                menu.add_command(label="No recent projects", command=lambda: None)
                return
            
            for proj in recent_projects:
                label = f"{proj['cabinet_id']} - {proj['project_name']}"
                menu.add_command(
                    label=label,
                    command=lambda p=proj: self.load_recent_project_from_db(p)
                )
                
        except Exception as e:
            print(f"Error updating recent dropdown: {e}")


    def load_recent_project_from_db(self, project_data):
        """Load a recent project from database - HIGHLIGHTER VERSION"""
        try:
            # Set project details
            self.cabinet_id = project_data['cabinet_id']
            self.project_name = project_data['project_name']
            self.sales_order_no = project_data.get('sales_order_no', '')
            self.storage_location = project_data['storage_location']
            
            self.prepare_project_folders()
            
            expected_excel_path = os.path.join(
                self.project_dirs["working_excel"],
                f"{self.cabinet_id.replace(' ', '_')}_Working.xlsx"
            )
            
            expected_session_path = os.path.join(
                self.project_dirs["sessions"],
                f"{self.cabinet_id}_annotations.json"
            )
            
            # Check PDF
            pdf_path = project_data.get('pdf_path')
            if not pdf_path or not os.path.exists(pdf_path):
                messagebox.showerror("Error", 
                                   f"PDF file not found:\n{pdf_path}\n\n"
                                   "The file may have been moved or deleted.")
                return
            
            # Check Excel
            if not os.path.exists(expected_excel_path):
                old_excel_path = project_data.get('excel_path')
                if old_excel_path and os.path.exists(old_excel_path):
                    try:
                        shutil.copy2(old_excel_path, expected_excel_path)
                        messagebox.showinfo("Excel Migrated", 
                                          f"Excel file migrated to new location:\n{expected_excel_path}")
                    except Exception as e:
                        messagebox.showerror("Error", 
                                           f"Excel file not found and couldn't migrate:\n{e}")
                        return
                else:
                    messagebox.showerror("Error", 
                                       f"Excel file not found at:\n{expected_excel_path}\n\n"
                                       "The file may have been moved or deleted.")
                    return
            
            # Load PDF
            self.pdf_document = fitz.open(pdf_path)
            self.current_pdf_path = pdf_path
            self.current_page = 0
            self.annotations = []
            self.zoom_level = 1.0
            self.tool_mode = None
            
            # ADDED: Reset highlighter state
            self.active_highlighter = None
            self.highlight_points = []
            self.update_color_button()
            
            self.root.config(cursor="")
            
            # Set Excel
            self.excel_file = expected_excel_path
            self.working_excel_path = expected_excel_path
            self.current_sr_no = self.get_next_sr_no()
            
            # Load session
            if os.path.exists(expected_session_path):
                self.load_session_from_path(expected_session_path)
            else:
                old_session_path = project_data.get('session_path')
                if old_session_path and os.path.exists(old_session_path):
                    try:
                        shutil.copy2(old_session_path, expected_session_path)
                        self.load_session_from_path(expected_session_path)
                    except:
                        self.display_page()
                else:
                    self.display_page()
            
            # Update database
            self.db.update_project(self.cabinet_id, {
                'pdf_path': self.current_pdf_path,
                'excel_path': expected_excel_path,
                'session_path': expected_session_path if os.path.exists(expected_session_path) else None,
                'last_accessed': datetime.now().isoformat()
            })
            
            messagebox.showinfo(
                "Project Loaded",
                f"✓ Loaded: {self.cabinet_id}\n"
                f"Project: {self.project_name}\n"
                f"Location: {self.storage_location}\n"
                f"Pages: {len(self.pdf_document)}"
            )
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load project:\n{e}")
            import traceback
            traceback.print_exc()


    # ============================================================================
    # UPDATED: count_open_punches - Now counts orange highlights as errors
    # ============================================================================

    def count_open_punches(self):
        """Count open punches in current Excel - HIGHLIGHTER VERSION"""
        try:
            if not self.excel_file or not os.path.exists(self.excel_file):
                return 0
            
            wb = load_workbook(self.excel_file, data_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
            
            open_count = 0
            row = 8
            
            while row <= ws.max_row + 5:
                sr = self.read_cell(ws, row, self.punch_cols['sr_no'])
                if sr is None:
                    break
                
                closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
                if not closed:
                    open_count += 1
                
                row += 1
            
            wb.close()
            return open_count
            
        except Exception as e:
            print(f"Error counting open punches: {e}")
            return 0


    # ============================================================================
    # UPDATED: handover_to_production - Counts orange highlights as punches
    # ============================================================================

    def handover_to_production(self):
        """Handover current cabinet to production - HIGHLIGHTER VERSION"""
        
        if not self.pdf_document or not self.excel_file:
            messagebox.showwarning("Incomplete", 
                                  "Please load a PDF and Excel file first.")
            return
        
        if not self.cabinet_id or not self.project_name:
            messagebox.showwarning("Missing Info", 
                                  "Project details are incomplete.")
            return
        
        # Check if there are any annotations
        if not self.annotations:
            proceed = messagebox.askyesno(
                "No Annotations",
                "No annotations found. Handover anyway?",
                icon='warning'
            )
            if not proceed:
                return
        
        # Count open punches
        open_punches = self.count_open_punches()
        
        if open_punches > 0:
            proceed = messagebox.askyesno(
                "Open Punches Detected",
                f"⚠️ There are {open_punches} open punch(es).\n\n"
                "Are you sure you want to handover to production?",
                icon='warning'
            )
            if not proceed:
                return
        
        # Save session before handover
        self.save_session()
        
        # Get user name
        try:
            username = os.getlogin()
        except:
            username = getpass.getuser()
        
        # Prepare handover data
        session_path = os.path.join(
            self.project_dirs.get("sessions", ""),
            f"{self.cabinet_id}_annotations.json"
        )
        
        # UPDATED: Count orange highlights as error punches
        error_highlights = [a for a in self.annotations 
                           if a.get('type') == 'highlight' and a.get('color') == 'orange']
        total_punches = len(error_highlights)
        
        handover_data = {
            "cabinet_id": self.cabinet_id,
            "project_name": self.project_name,
            "sales_order_no": self.sales_order_no,
            "pdf_path": self.current_pdf_path,
            "excel_path": self.excel_file,
            "session_path": session_path if os.path.exists(session_path) else None,
            "total_punches": total_punches,  # CHANGED: Now counts orange highlights
            "open_punches": open_punches,
            "closed_punches": total_punches - open_punches,
            "handed_over_by": username,
            "handed_over_date": datetime.now().isoformat()
        }
        
        success = self.handover_db.add_quality_handover(handover_data)
        self.update_status_and_sync('handed_to_production')
        
        if success:
            messagebox.showinfo("Handover Complete", 
                              "✓ Successfully handed over to Production\n\n"
                              f"Cabinet: {self.cabinet_id}\n"
                              f"Total Error Highlights: {total_punches}\n"
                              f"Open Punches: {open_punches}")
        else:
            messagebox.showwarning("Already Handed Over", 
                                 "Cabinet already in production queue")

    # ================================================================
    # COMPREHENSIVE STATUS AND STATISTICS MANAGEMENT
    # ================================================================

    def get_current_status_from_db(self):
        """Get the current status of this cabinet from manager database
        
        Returns the workflow status without modifying anything.
        Safe to call anytime to check current state.
        """
        try:
            conn = sqlite3.connect(self.manager_db.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT status FROM cabinets WHERE cabinet_id = ?
            ''', (self.cabinet_id,))
            
            result = cursor.fetchone()
            conn.close()
            
            if result:
                return result[0]
            else:
                # New cabinet - default to quality_inspection
                return 'quality_inspection'
                
        except Exception as e:
            print(f"Error getting status: {e}")
            return 'quality_inspection'  # Safe fallback


    def update_status_and_sync(self, new_status):
        """Update status and sync stats in one atomic operation
        
        This is the RECOMMENDED approach for status changes.
        Use this when you want to change workflow status and update statistics.
        
        Args:
            new_status: One of:
                - 'quality_inspection' (initial state)
                - 'handed_to_production' (after handover)
                - 'in_progress' (production working on it)
                - 'being_closed_by_quality' (quality verifying rework)
                - 'closed' (final state)
        
        Returns:
            bool: True if successful, False otherwise
        """
        if not self.cabinet_id:
            print("⚠️ Cannot update status: cabinet_id not set")
            return False
        
        try:
            # Count pages with annotations
            annotated_pages = 0
            total_pages = 0
            
            if self.pdf_document:
                annotated_pages = len(set(ann['page'] for ann in self.annotations 
                                         if ann.get('page') is not None))
                total_pages = len(self.pdf_document)
            
            # Count punches by type
            error_anns = [a for a in self.annotations if a.get('type') == 'error']
            total_punches = len(error_anns)
            
            # Count from Excel for accuracy
            open_punches = self.count_open_punches()
            
            # Count implemented and closed from Excel
            implemented_punches = 0
            closed_punches = 0
            
            if self.excel_file and os.path.exists(self.excel_file):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(self.excel_file, data_only=True)
                    ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
                    
                    row = 9  # Start from row 9 (matching manager code)
                    while row <= ws.max_row + 5:
                        # Check if this row has a punch (has checked_name)
                        checked = self.read_cell(ws, row, self.punch_cols['checked_name'])
                        
                        if checked:  # This is a logged punch
                            implemented = self.read_cell(ws, row, self.punch_cols['implemented_name'])
                            closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
                            
                            if closed:
                                closed_punches += 1
                            elif implemented:
                                implemented_punches += 1
                        
                        row += 1
                        
                        # Safety limit
                        if row > 2000:
                            break
                    
                    wb.close()
                except Exception as e:
                    print(f"⚠️ Error reading Excel stats: {e}")
            
            # Update manager database with NEW status AND stats
            success = self.manager_db.update_cabinet(
                self.cabinet_id,
                self.project_name,
                self.sales_order_no,
                total_pages,
                annotated_pages,
                total_punches,
                open_punches,
                implemented_punches,
                closed_punches,
                new_status,  # NEW STATUS
                storage_location=getattr(self, 'storage_location', None),
                excel_path=self.excel_file
            )
            
            if success:
                print(f"✓ Updated {self.cabinet_id}: {new_status}")
                print(f"  Stats: {total_punches} punches ({open_punches} open, "
                      f"{implemented_punches} implemented, {closed_punches} closed)")
            
            return success
            
        except Exception as e:
            print(f"❌ Update status and sync error: {e}")
            import traceback
            traceback.print_exc()
            return False


    def sync_manager_stats_only(self):
        """Sync ONLY statistics, don't touch status at all
        
        Use this when you want to update counts without changing workflow status.
        This is safe to call frequently (e.g., after each annotation).
        
        IMPORTANT: If cabinet doesn't exist in database, it will be created with
        status determined from Interphase worksheet (or defaults to 'quality_inspection')
        """
        if not self.pdf_document or not self.cabinet_id:
            return
        
        try:
            # Count pages with annotations
            annotated_pages = len(set(ann['page'] for ann in self.annotations 
                                     if ann.get('page') is not None))
            total_pages = len(self.pdf_document)
            
            # Count punches
            error_anns = [a for a in self.annotations if a.get('type') == 'error']
            total_punches = len(error_anns)
            open_punches = self.count_open_punches()
            
            # Count implemented and closed
            implemented_punches = 0
            closed_punches = 0
            
            if self.excel_file and os.path.exists(self.excel_file):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(self.excel_file, data_only=True)
                    ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
                    
                    row = 9  # Start from row 9 (matching manager code)
                    while row <= ws.max_row + 5:
                        # Check if this row has a punch (has checked_name)
                        checked = self.read_cell(ws, row, self.punch_cols['checked_name'])
                        
                        if checked:  # This is a logged punch
                            implemented = self.read_cell(ws, row, self.punch_cols['implemented_name'])
                            closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
                            
                            if closed:
                                closed_punches += 1
                            elif implemented:
                                implemented_punches += 1
                        
                        row += 1
                        
                        # Safety limit
                        if row > 2000:
                            break
                    
                    wb.close()
                except:
                    pass
            
            # Check if cabinet exists in database
            conn = sqlite3.connect(self.manager_db.db_path)
            cursor = conn.cursor()
            
            cursor.execute('SELECT status FROM cabinets WHERE cabinet_id = ?', (self.cabinet_id,))
            existing = cursor.fetchone()
            
            if existing:
                # Cabinet exists - UPDATE ONLY statistics (preserve status)
                cursor.execute('''
                    UPDATE cabinets 
                    SET total_pages = ?,
                        annotated_pages = ?,
                        total_punches = ?,
                        open_punches = ?,
                        implemented_punches = ?,
                        closed_punches = ?,
                        last_updated = ?,
                        excel_path = ?,
                        storage_location = ?
                    WHERE cabinet_id = ?
                ''', (total_pages, annotated_pages, total_punches, open_punches,
                      implemented_punches, closed_punches, datetime.now().isoformat(),
                      self.excel_file, getattr(self, 'storage_location', None),
                      self.cabinet_id))
                
                print(f"✓ Updated stats for {self.cabinet_id} (status: {existing[0]})")
            else:
                # Cabinet doesn't exist - CREATE with initial status
                # Try to get status from Interphase worksheet first
                initial_status = self.manager_db.get_status_from_interphase(self.excel_file)
                if not initial_status:
                    initial_status = 'quality_inspection'  # Default fallback
                
                cursor.execute('''
                    INSERT INTO cabinets (
                        cabinet_id, project_name, sales_order_no,
                        total_pages, annotated_pages, total_punches,
                        open_punches, implemented_punches, closed_punches,
                        status, created_date, last_updated,
                        storage_location, excel_path
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    self.cabinet_id,
                    self.project_name,
                    self.sales_order_no,
                    total_pages,
                    annotated_pages,
                    total_punches,
                    open_punches,
                    implemented_punches,
                    closed_punches,
                    initial_status,
                    datetime.now().isoformat(),
                    datetime.now().isoformat(),
                    getattr(self, 'storage_location', None),
                    self.excel_file
                ))
                
                print(f"✓ Created {self.cabinet_id} in dashboard with status: {initial_status}")
                print(f"  Stats: {total_punches} punches ({open_punches} open, "
                      f"{implemented_punches} implemented, {closed_punches} closed)")
            
            conn.commit()
            conn.close()
            
        except Exception as e:
            print(f"Stats sync error: {e}")
            import traceback
            traceback.print_exc()


    def sync_manager_stats(self):
        """LEGACY: Sync statistics while preserving existing status
        
        This is kept for backward compatibility.
        Internally calls sync_manager_stats_only().
        
        Will now also CREATE cabinet if it doesn't exist (making it visible on dashboard)
        """
        self.sync_manager_stats_only()


    def ensure_visible_on_dashboard(self):
        """Explicitly ensure this cabinet is visible on the manager dashboard
        
        Call this when:
        - Opening a PDF for the first time
        - Starting quality inspection
        - Any time you want to make sure the cabinet appears on dashboard
        
        This will:
        1. Create the cabinet in database if it doesn't exist
        2. Set initial status from Interphase worksheet (or 'quality_inspection')
        3. Sync all current statistics
        """
        if not self.cabinet_id:
            print("⚠️ Cannot make visible: cabinet_id not set")
            return False
        
        try:
            conn = sqlite3.connect(self.manager_db.db_path)
            cursor = conn.cursor()
            
            # Check if exists
            cursor.execute('SELECT cabinet_id FROM cabinets WHERE cabinet_id = ?', (self.cabinet_id,))
            exists = cursor.fetchone()
            
            conn.close()
            
            if exists:
                # Already exists, just sync stats
                self.sync_manager_stats_only()
                print(f"✓ {self.cabinet_id} already on dashboard - stats synced")
                return True
            else:
                # Doesn't exist, create it
                self.sync_manager_stats_only()  # This will create it now
                print(f"✓ {self.cabinet_id} is now visible on dashboard")
                return True
                
        except Exception as e:
            print(f"❌ Error ensuring visibility: {e}")
            return False


    def count_open_punches(self):
        """Count open punches in current Excel
        
        Returns:
            int: Number of punches that are not closed
        """
        try:
            if not self.excel_file or not os.path.exists(self.excel_file):
                return 0
            
            wb = load_workbook(self.excel_file, data_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
            
            open_count = 0
            row = 8
            
            while row <= ws.max_row + 5:
                sr = self.read_cell(ws, row, self.punch_cols['sr_no'])
                if sr is None:
                    break
                
                closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
                if not closed:
                    open_count += 1
                
                row += 1
            
            wb.close()
            return open_count
            
        except Exception as e:
            print(f"Error counting open punches: {e}")
            return 0




    


# ================================================================
# MAIN ENTRY POINT
# ================================================================

def main():
    root = tk.Tk()
    app = CircuitInspector(root)
    root.mainloop()


if __name__ == "__main__":
    main()
