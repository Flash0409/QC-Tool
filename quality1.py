"""
Add these changes to your quality.py file to integrate with Manager UI

STEP 1: Add this import at the top of your quality.py file
"""
import sqlite3

"""
STEP 2: Add ManagerDB class right after your imports, before CircuitInspector class
"""

class ManagerDB:
    """Simple manager database integration"""
    def __init__(self, db_path):
        self.db_path = db_path
        self.init_database()
    
    def init_database(self):
        """Initialize tables if they don't exist"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS cabinets (
            cabinet_id TEXT PRIMARY KEY,
            project_name TEXT,
            sales_order_no TEXT,
            total_pages INTEGER DEFAULT 0,
            annotated_pages INTEGER DEFAULT 0,
            total_punches INTEGER DEFAULT 0,
            open_punches INTEGER DEFAULT 0,
            implemented_punches INTEGER DEFAULT 0,
            closed_punches INTEGER DEFAULT 0,
            status TEXT DEFAULT 'quality_inspection',
            created_date TEXT,
            last_updated TEXT
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS category_occurrences (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cabinet_id TEXT,
            project_name TEXT,
            category TEXT,
            subcategory TEXT,
            occurrence_date TEXT
        )''')
        
        conn.commit()
        conn.close()
    
    def update_cabinet(self, cabinet_id, project_name, sales_order_no, 
                      total_pages, annotated_pages, total_punches, 
                      open_punches, implemented_punches, closed_punches, status):
        """Update cabinet statistics"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            from datetime import datetime
            cursor.execute('''
                INSERT OR REPLACE INTO cabinets 
                (cabinet_id, project_name, sales_order_no, total_pages, annotated_pages,
                 total_punches, open_punches, implemented_punches, closed_punches, status,
                 created_date, last_updated)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 
                        COALESCE((SELECT created_date FROM cabinets WHERE cabinet_id = ?), ?),
                        ?)
            ''', (cabinet_id, project_name, sales_order_no, total_pages, annotated_pages,
                  total_punches, open_punches, implemented_punches, closed_punches, status,
                  cabinet_id, datetime.now().isoformat(), datetime.now().isoformat()))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Manager DB update error: {e}")
            return False
    
    def log_category_occurrence(self, cabinet_id, project_name, category, subcategory):
        """Log a category occurrence"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            from datetime import datetime
            cursor.execute('''
                INSERT INTO category_occurrences 
                (cabinet_id, project_name, category, subcategory, occurrence_date)
                VALUES (?, ?, ?, ?, ?)
            ''', (cabinet_id, project_name, category, subcategory, datetime.now().isoformat()))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Category logging error: {e}")
            return False
    
    def update_status(self, cabinet_id, status):
        """Update cabinet status"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            from datetime import datetime
            cursor.execute('''
                UPDATE cabinets 
                SET status = ?, last_updated = ?
                WHERE cabinet_id = ?
            ''', (status, datetime.now().isoformat(), cabinet_id))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Status update error: {e}")
            return False


"""
STEP 3: In CircuitInspector.__init__(), add manager_db initialization 
Find this line:
    self.db = DatabaseManager(db_path)

Add RIGHT AFTER it:
"""

        # Initialize Manager DB
        manager_db_path = os.path.join(base, "manager.db")
        self.manager_db = ManagerDB(manager_db_path)


"""
STEP 4: Add this helper method to CircuitInspector class
(add it anywhere in the class, I suggest near count_open_punches)
"""

    def sync_manager_stats(self):
        """Sync current cabinet statistics to manager database"""
        if not self.pdf_document or not self.cabinet_id:
            return
        
        try:
            # Count pages with annotations
            annotated_pages = len(set(ann['page'] for ann in self.annotations if ann.get('page') is not None))
            total_pages = len(self.pdf_document)
            
            # Count punches by type
            error_anns = [a for a in self.annotations if a.get('type') == 'error']
            total_punches = len(error_anns)
            
            # Count from Excel for accuracy
            open_punches = self.count_open_punches()
            
            # Count implemented (has implemented_name but no closed_name)
            implemented_punches = 0
            closed_punches = 0
            
            if self.excel_file and os.path.exists(self.excel_file):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(self.excel_file, data_only=True)
                    ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
                    
                    row = 8
                    while row <= ws.max_row + 5:
                        sr = self.read_cell(ws, row, self.punch_cols['sr_no'])
                        if sr is None:
                            break
                        
                        implemented = self.read_cell(ws, row, self.punch_cols['implemented_name'])
                        closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
                        
                        if closed:
                            closed_punches += 1
                        elif implemented:
                            implemented_punches += 1
                        
                        row += 1
                    
                    wb.close()
                except:
                    pass
            
            # Update manager database
            self.manager_db.update_cabinet(
                self.cabinet_id,
                self.project_name,
                self.sales_order_no,
                total_pages,
                annotated_pages,
                total_punches,
                open_punches,
                implemented_punches,
                closed_punches,
                'quality_inspection'
            )
        except Exception as e:
            print(f"Manager sync error: {e}")


"""
STEP 5: In log_error() method, add category logging
Find the line that says:
    self.annotations.append(ann)

Add RIGHT AFTER it:
"""

            # Log to manager database
            try:
                self.manager_db.log_category_occurrence(
                    self.cabinet_id,
                    self.project_name,
                    component_type,
                    error_name
                )
            except Exception as e:
                print(f"Manager category logging failed: {e}")


"""
STEP 6: In log_custom_error() method, add category logging
Find the line that says:
    self.annotations.append(ann)

Add RIGHT AFTER it:
"""

            # Log to manager database
            try:
                self.manager_db.log_category_occurrence(
                    self.cabinet_id,
                    self.project_name,
                    custom_category,
                    None
                )
            except Exception as e:
                print(f"Manager category logging failed: {e}")


"""
STEP 7: Call sync_manager_stats at strategic points

A) In display_page() method, at the END of the try block, add:
"""
            # Sync stats to manager database (non-blocking)
            try:
                self.sync_manager_stats()
            except:
                pass


"""
B) In save_session() method, AFTER successfully saving, add:
"""
            # Sync to manager database
            try:
                self.sync_manager_stats()
            except:
                pass


"""
C) In handover_to_production() method, RIGHT AFTER self.save_session(), add:
"""

        # Update status in manager database
        try:
            self.sync_manager_stats()
            self.manager_db.update_status(self.cabinet_id, 'handed_to_production')
        except Exception as e:
            print(f"Manager status update failed: {e}")


"""
STEP 8: In punch_closing_mode(), when a punch is closed
Find the code that saves to Excel in the close_punch() function:
    wb.save(self.excel_file)
    wb.close()

Add RIGHT AFTER:
"""

                # Update manager stats
                try:
                    self.sync_manager_stats()
                except:
                    pass


"""
STEP 9: Export annotated PDF - sync after completion
In export_annotated_pdf(), after the success message, add:
"""

            # Sync final stats
            try:
                self.sync_manager_stats()
            except:
                pass


"""
=====================================================================
COMPLETE EXAMPLE - How your __init__ should look:
=====================================================================
"""

def __init__(self, root):
    self.root = root
    self.root.title("Quality Inspection Tool")
    self.root.geometry("1400x900")

    # ... all your existing initialization code ...

    base = get_app_base_dir()
    db_path = os.path.join(base, "inspection_tool.db")
    self.db = DatabaseManager(db_path)
    
    # ADD THIS - Manager DB Integration
    manager_db_path = os.path.join(base, "manager.db")
    self.manager_db = ManagerDB(manager_db_path)


"""
=====================================================================
SUMMARY OF CHANGES:
=====================================================================

1. Added ManagerDB class (simple, no complex dependencies)
2. Initialize manager_db in __init__
3. Added sync_manager_stats() helper method
4. Log category occurrences when punches are created
5. Sync statistics at key points (display, save, export)
6. Update status when handing over to production

This integration is:
- Non-blocking (won't crash if manager DB has issues)
- Automatic (syncs in background)
- Lightweight (only writes, no complex queries)
- Compatible with your existing code

The manager.db file will be created automatically in the same folder
as inspection_tool.db when you first log a punch.
"""
