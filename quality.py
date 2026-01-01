import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, Menu
from PIL import Image, ImageTk, ImageDraw, ImageFont
import fitz  # PyMuPDF
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime
import shutil
import tempfile
import re
import os
import json
import numpy as np
import getpass
import sys
import subprocess
import sqlite3
import shlex
from difflib import SequenceMatcher
from handover_database import HandoverDB
from database_manager import DatabaseManager
from PIL import Image, ImageTk, ImageDraw, ImageFont

def get_app_base_dir():
    """
    Returns the directory where the app is running from.
    Works for both .py and PyInstaller .exe
    """
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

class ManagerDB:
    """Simple manager database integration"""
    def __init__(self, db_path):
        self.db_path = db_path
        self.init_database()
    
    def init_database(self):
        """Initialize tables if they don't exist"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS cabinets (
            cabinet_id TEXT PRIMARY KEY,
            project_name TEXT,
            sales_order_no TEXT,
            total_pages INTEGER DEFAULT 0,
            annotated_pages INTEGER DEFAULT 0,
            total_punches INTEGER DEFAULT 0,
            open_punches INTEGER DEFAULT 0,
            implemented_punches INTEGER DEFAULT 0,
            closed_punches INTEGER DEFAULT 0,
            status TEXT DEFAULT 'quality_inspection',
            created_date TEXT,
            last_updated TEXT
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS category_occurrences (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cabinet_id TEXT,
            project_name TEXT,
            category TEXT,
            subcategory TEXT,
            occurrence_date TEXT
        )''')
        
        conn.commit()
        conn.close()
    
    def update_cabinet(self, cabinet_id, project_name, sales_order_no, 
                      total_pages, annotated_pages, total_punches, 
                      open_punches, implemented_punches, closed_punches, status):
        """Update cabinet statistics"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            from datetime import datetime
            cursor.execute('''
                INSERT OR REPLACE INTO cabinets 
                (cabinet_id, project_name, sales_order_no, total_pages, annotated_pages,
                 total_punches, open_punches, implemented_punches, closed_punches, status,
                 created_date, last_updated)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 
                        COALESCE((SELECT created_date FROM cabinets WHERE cabinet_id = ?), ?),
                        ?)
            ''', (cabinet_id, project_name, sales_order_no, total_pages, annotated_pages,
                  total_punches, open_punches, implemented_punches, closed_punches, status,
                  cabinet_id, datetime.now().isoformat(), datetime.now().isoformat()))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Manager DB update error: {e}")
            return False
    
    def log_category_occurrence(self, cabinet_id, project_name, category, subcategory):
        """Log a category occurrence"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            from datetime import datetime
            cursor.execute('''
                INSERT INTO category_occurrences 
                (cabinet_id, project_name, category, subcategory, occurrence_date)
                VALUES (?, ?, ?, ?, ?)
            ''', (cabinet_id, project_name, category, subcategory, datetime.now().isoformat()))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Category logging error: {e}")
            return False
    
    def update_status(self, cabinet_id, status):
        """Update cabinet status"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            from datetime import datetime
            cursor.execute('''
                UPDATE cabinets 
                SET status = ?, last_updated = ?
                WHERE cabinet_id = ?
            ''', (status, datetime.now().isoformat(), cabinet_id))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Status update error: {e}")
            return False


class CircuitInspector:
    def __init__(self, root):
        self.root = root
        self.root.title("Quality Inspection Tool")
        self.root.geometry("1400x900")

        # Data / files
        self.pdf_document = None
        self.current_pdf_path = None
        self.current_page = 0
        self.project_name = ""
        self.sales_order_no = ""
        self.cabinet_id = ""
        self.annotations = []
        base = get_app_base_dir()
        self.master_excel_file = os.path.join(base, "Emerson.xlsx")

        self.excel_file = None
        self.working_excel_path = None
        self.checklist_file = self.excel_file
        self.zoom_level = 1.0
        self.current_sr_no = 1
        self.current_page_image = None
        self.tool_mode = None  # None, "pen", or "text"
        self.pen_points = []
        self.session_refs = set()
        self.project_dirs = {}

        # Fixed column mapping
        self.punch_sheet_name = 'Punch Sheet'
        self.punch_cols = {
            'sr_no': 'A',
            'ref_no': 'B',
            'desc': 'C',
            'category': 'D',
            'checked_name': 'E',
            'checked_date': 'F',
            'implemented_name': 'G',
            'implemented_date': 'H',
            'closed_name': 'I',
            'closed_date': 'J'
        }
        
        self.interphase_sheet_name = 'Interphase'
        self.interphase_cols = {
            'ref_no': 'B',
            'description': 'C',
            'status': 'D',
            'name':'E',
            'date':'F',
            'remark':'G'
        }

        self.header_cells = {
            "Interphase": {
                "project_name": "C4",
                "sales_order": "C6",
                "cabinet_id": "F6"
            },
            "Punch Sheet": {
                "project_name": "C2",
                "sales_order": "C4",
                "cabinet_id": "H4"
            }
        }

        self.categories = []
        self.category_file = os.path.join(os.path.dirname(get_app_base_dir()), "assets", "categories.json")
        self.load_categories()

        # Drawing / selection state
        self.drawing = False
        self.drawing_type = None  # 'ok', 'error', 'pen', 'text'
        self.rect_start_x = None
        self.rect_start_y = None
        self.temp_rect_id = None
        self.temp_line_ids = []  # Store temporary pen line IDs
        self.selected_annotation = None

        self.setup_ui()
        self.current_sr_no = self.get_next_sr_no()
        
        base = get_app_base_dir()
        db_path = os.path.join(base, "inspection_tool.db")
        self.db = DatabaseManager(db_path)
        manager_db_path = os.path.join(base, "manager.db")
        self.manager_db = ManagerDB(manager_db_path)
        self.handover_db = HandoverDB(os.path.join(base, "handover_db.json"))
        self.load_recent_projects_ui()

    # ================================================================
    # COORDINATE CONVERSION HELPERS
    # ================================================================
    
    def page_to_display_scale(self):
        return 2.0 * self.zoom_level

    def display_to_page_coords(self, pts):
        """
        Convert display-space coordinates to page-space coordinates.
        Handles:
        - Single point: (x, y) -> (x/scale, y/scale)
        - List of points: [(x1,y1), ...] -> [(x1/scale, y1/scale), ...]
        """
        scale = self.page_to_display_scale()
        
        # Handle single point tuple
        if isinstance(pts, tuple) and len(pts) == 2:
            if not isinstance(pts[0], (list, tuple)):
                return (pts[0] / scale, pts[1] / scale)
        
        # Handle list of points
        return [(x / scale, y / scale) for x, y in pts]

    def page_to_display_coords(self, pts):
        """
        Convert page-space coordinates to display-space coordinates.
        Handles:
        - Single point: (x, y) -> (x*scale, y*scale)
        - List of points: [(x1,y1), ...] -> [(x1*scale, y1*scale), ...]
        """
        scale = self.page_to_display_scale()
        
        # Handle single point tuple
        if isinstance(pts, tuple) and len(pts) == 2:
            if not isinstance(pts[0], (list, tuple)):
                return (pts[0] * scale, pts[1] * scale)
        
        # Handle list of points
        return [(x * scale, y * scale) for x, y in pts]

    def bbox_page_to_display(self, bbox_page):
        scale = self.page_to_display_scale()
        x1, y1, x2, y2 = bbox_page
        return (x1 * scale, y1 * scale, x2 * scale, y2 * scale)

    def bbox_display_to_page(self, bbox_display):
        scale = self.page_to_display_scale()
        x1, y1, x2, y2 = bbox_display
        return (x1 / scale, y1 / scale, x2 / scale, y2 / scale)

    # ================================================================
    # TOOL MODE CONTROL
    # ================================================================

    def set_tool_mode(self, mode):
        """Set the current tool mode with visual feedback"""
        if self.tool_mode == mode:
            # Toggle off
            self.tool_mode = None
            self.root.config(cursor="")
            self.pen_btn.config(bg='#334155', relief=tk.FLAT)
            self.text_btn.config(bg='#334155', relief=tk.FLAT)
            self._flash_status("Normal mode - Left drag: OK | Right drag: Error", bg='#64748b')
        else:
            # Reset all buttons
            self.pen_btn.config(bg='#334155', relief=tk.FLAT)
            self.text_btn.config(bg='#334155', relief=tk.FLAT)
            
            self.tool_mode = mode
            if mode == "pen":
                self.root.config(cursor="pencil")
                self.pen_btn.config(bg='#3b82f6', relief=tk.SUNKEN)
                self._flash_status("✏️ Pen Tool Active - Right-click to exit", bg='#3b82f6')
            elif mode == "text":
                self.root.config(cursor="xterm")
                self.text_btn.config(bg='#3b82f6', relief=tk.SUNKEN)
                self._flash_status("🅰️ Text Tool Active - Right-click to exit", bg='#3b82f6')
        
        self.pen_points.clear()
        self.clear_temp_drawings()


    def clear_temp_drawings(self):
        """Clear temporary drawing elements from canvas"""
        for line_id in self.temp_line_ids:
            try:
                self.canvas.delete(line_id)
            except:
                pass
        self.temp_line_ids.clear()
        
        if self.temp_rect_id:
            try:
                self.canvas.delete(self.temp_rect_id)
            except:
                pass
            self.temp_rect_id = None

    # ================================================================
    # MOUSE EVENT HANDLERS - COMPLETELY REWRITTEN
    # ================================================================

    def on_left_press(self, event):
        """Handle left mouse button press"""
        if not self.pdf_document:
            messagebox.showwarning("Warning", "Please load a PDF first")
            return

        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)

        # -------- PEN TOOL --------
        if self.tool_mode == "pen":
            self.drawing = True
            self.drawing_type = "pen"
            self.pen_points.clear()
            self.clear_temp_drawings()
            self.pen_points.append((x, y))
            return

        # -------- TEXT TOOL --------
        if self.tool_mode == "text":
            self.drawing = True
            self.drawing_type = "text"
            # Store position for text placement
            self.text_pos_x = x
            self.text_pos_y = y
            return

        # -------- NORMAL MODE: OK Rectangle --------
        self.drawing = True
        self.drawing_type = 'ok'
        self.rect_start_x = x
        self.rect_start_y = y

    def on_left_drag(self, event):
        """Handle left mouse button drag"""
        if not self.drawing:
            return

        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)

        # -------- PEN TOOL DRAWING --------
        if self.drawing_type == "pen":
            if len(self.pen_points) > 0:
                last_x, last_y = self.pen_points[-1]
                # Draw line segment on canvas
                line_id = self.canvas.create_line(
                    last_x, last_y, x, y,
                    fill="red", width=3,
                    capstyle=tk.ROUND, smooth=True
                )
                self.temp_line_ids.append(line_id)
            self.pen_points.append((x, y))
            return

        # -------- OK RECTANGLE PREVIEW --------
        if self.drawing_type == 'ok':
            if self.temp_rect_id:
                self.canvas.delete(self.temp_rect_id)
            self.temp_rect_id = self.canvas.create_rectangle(
                self.rect_start_x, self.rect_start_y, x, y,
                outline='green', width=3, dash=(5, 5)
            )

    def on_left_release(self, event):
        """Handle left mouse button release"""
        if not self.pdf_document or not self.drawing:
            return

        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)

        # -------- PEN TOOL FINISH --------
        if self.drawing_type == "pen":
            if len(self.pen_points) >= 2:
                # Convert to page coordinates and save
                points_page = self.display_to_page_coords(self.pen_points)
                self.annotations.append({
                    'type': 'pen',
                    'page': self.current_page,
                    'points': points_page,
                    'timestamp': datetime.now().isoformat()
                })
            self.pen_points.clear()
            self.clear_temp_drawings()
            self.drawing = False
            self.drawing_type = None
            self.display_page()
            return

        # -------- TEXT TOOL FINISH --------
        if self.drawing_type == "text":
            txt = simpledialog.askstring("Text", "Enter text:", parent=self.root)
            if txt and txt.strip():
                pos_page = self.display_to_page_coords((self.text_pos_x, self.text_pos_y))
                self.annotations.append({
                    'type': 'text',
                    'page': self.current_page,
                    'pos_page': pos_page,
                    'text': txt.strip(),
                    'timestamp': datetime.now().isoformat()
                })
                self.display_page()
            self.drawing = False
            self.drawing_type = None
            return

        # -------- OK RECTANGLE FINISH --------
        if self.drawing_type == 'ok':
            self.clear_temp_drawings()

            x1 = min(self.rect_start_x, x)
            y1 = min(self.rect_start_y, y)
            x2 = max(self.rect_start_x, x)
            y2 = max(self.rect_start_y, y)

            # Minimum size check
            if abs(x2 - x1) < 10 or abs(y2 - y1) < 10:
                self.drawing = False
                self.drawing_type = None
                return

            bbox_display = (x1, y1, x2, y2)
            bbox_page = self.bbox_display_to_page(bbox_display)

            self.annotations.append({
                'type': 'ok',
                'page': self.current_page,
                'bbox_page': bbox_page,
                'timestamp': datetime.now().isoformat()
            })
            self.display_page()

        self.drawing = False
        self.drawing_type = None

    # ================================================================
    # UPDATED: on_right_press - Right-click exits pen/text mode
    # ================================================================

    def on_right_press(self, event):
        """Handle right mouse button press.
        - If in pen/text mode: exit to normal mode
        - If in normal mode: start error rectangle
        """
        if not self.pdf_document:
            messagebox.showwarning("Warning", "Please load a PDF first")
            return

        # -------- EXIT TOOL MODE ON RIGHT CLICK --------
        if self.tool_mode in ("pen", "text"):
            self.tool_mode = None
            self.root.config(cursor="")
            self.pen_points.clear()
            self.clear_temp_drawings()
            self.drawing = False
            self.drawing_type = None
            
            # Visual feedback
            self.root.after(50, lambda: self._flash_status("Normal mode - Right drag to mark errors"))
            return

        # -------- NORMAL MODE: Start Error Rectangle --------
        self.drawing = True
        self.drawing_type = 'error'
        self.rect_start_x = self.canvas.canvasx(event.x)
        self.rect_start_y = self.canvas.canvasy(event.y)


    def _flash_status(self, message, bg='#10b981'):
        """Show a modern temporary status message"""
        status_label = tk.Label(
            self.root, 
            text=message, 
            bg=bg, 
            fg='white', 
            font=('Segoe UI', 10, 'bold'),
            padx=25, 
            pady=12,
            relief=tk.FLAT,
            borderwidth=0
        )
        status_label.place(relx=0.5, rely=0.08, anchor='center')
        
        # Fade out effect
        def fade_out(alpha=1.0):
            if alpha > 0:
                # Simulate fade by destroying after delay
                self.root.after(50, lambda: fade_out(alpha - 0.1))
            else:
                status_label.destroy()
        
        self.root.after(1500, lambda: fade_out())


    def on_right_drag(self, event):
        """Handle right mouse button drag"""
        if not self.drawing or self.drawing_type != 'error':
            return

        if self.temp_rect_id:
            try:
                self.canvas.delete(self.temp_rect_id)
            except:
                pass

        current_x = self.canvas.canvasx(event.x)
        current_y = self.canvas.canvasy(event.y)
        self.temp_rect_id = self.canvas.create_rectangle(
            self.rect_start_x, self.rect_start_y,
            current_x, current_y,
            outline='orange', width=3, dash=(5, 5)
        )

    def on_right_release(self, event):
        """Handle right mouse button release"""
        if not self.drawing or self.drawing_type != 'error':
            return

        self.clear_temp_drawings()

        end_x = self.canvas.canvasx(event.x)
        end_y = self.canvas.canvasy(event.y)

        x1 = min(self.rect_start_x, end_x)
        y1 = min(self.rect_start_y, end_y)
        x2 = max(self.rect_start_x, end_x)
        y2 = max(self.rect_start_y, end_y)

        if abs(x2 - x1) < 10 or abs(y2 - y1) < 10:
            self.drawing = False
            self.drawing_type = None
            return

        bbox_display = (x1, y1, x2, y2)
        bbox_page = self.bbox_display_to_page(bbox_display)

        # Show category menu
        menu = Menu(self.root, tearoff=0)

        for cat in self.categories:
            if cat.get("mode") == "template":
                menu.add_command(
                    label=f"🔧 {cat['name']}",
                    command=lambda c=cat, bp=bbox_page: self.handle_template_category(c, bp)
                )
            elif cat.get("mode") == "parent":
                cat_menu = Menu(menu, tearoff=0)
                for sub in cat.get("subcategories", []):
                    cat_menu.add_command(
                        label=sub["name"],
                        command=lambda c=cat, s=sub, bp=bbox_page: self.handle_subcategory(c, s, bp)
                    )
                menu.add_cascade(label=f"🔧 {cat['name']}", menu=cat_menu)

        menu.add_separator()
        menu.add_command(
            label="📝 Custom Action Point",
            command=lambda bp=bbox_page: self.log_custom_error(bp, None)
        )

        menu.tk_popup(event.x_root, event.y_root)
        self.drawing = False
        self.drawing_type = None

    # ================================================================
    # DISPLAY PAGE - WITH PEN AND TEXT RENDERING
    # ================================================================

    def display_page(self):
        """Render the current PDF page with all annotations"""
        if not self.pdf_document:
            self.canvas.delete("all")
            self.page_label.config(text="Page: 0/0")
            return

        try:
            page = self.pdf_document[self.current_page]
            mat = fitz.Matrix(self.page_to_display_scale(), self.page_to_display_scale())
            pix = page.get_pixmap(matrix=mat)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            self.current_page_image = np.array(img)
            draw = ImageDraw.Draw(img, 'RGBA')

            # Try to load a font for text
            try:
                font_size = max(12, int(14 * self.zoom_level))
                font = ImageFont.truetype("arial.ttf", font_size)
            except:
                font = ImageFont.load_default()

            for ann in self.annotations:
                if ann.get('page') != self.current_page:
                    continue

                ann_type = ann.get('type')

                # -------- RECTANGLE ANNOTATIONS (ok/error) --------
                if ann_type in ('ok', 'error') and 'bbox_page' in ann:
                    x1d, y1d, x2d, y2d = self.bbox_page_to_display(ann['bbox_page'])
                    is_selected = (self.selected_annotation is ann)
                    w = int(5 * self.zoom_level) if is_selected else int(3 * self.zoom_level)

                    if ann_type == 'ok':
                        draw.rectangle([x1d, y1d, x2d, y2d],
                                       fill=(0, 255, 0, 80),
                                       outline='blue' if is_selected else 'green',
                                       width=w)
                    else:
                        draw.rectangle([x1d, y1d, x2d, y2d],
                                       fill=(255, 165, 0, 120),
                                       outline='blue' if is_selected else 'orange',
                                       width=w)

                    if ann.get('closed_by'):
                        cx = x1d + 8
                        cy = y1d + 8
                        draw.ellipse([cx - 6, cy - 6, cx + 6, cy + 6], fill=(0, 128, 0, 200))

                # -------- PEN STROKES --------
                elif ann_type == 'pen' and 'points' in ann:
                    points_page = ann['points']
                    if len(points_page) >= 2:
                        points_display = self.page_to_display_coords(points_page)
                        stroke_width = max(2, int(3 * self.zoom_level))
                        for i in range(len(points_display) - 1):
                            x1, y1 = points_display[i]
                            x2, y2 = points_display[i + 1]
                            draw.line([x1, y1, x2, y2], fill='red', width=stroke_width)

                # -------- TEXT ANNOTATIONS --------
                elif ann_type == 'text' and 'pos_page' in ann:
                    pos_page = ann['pos_page']
                    pos_display = self.page_to_display_coords(pos_page)
                    text = ann.get('text', '')
                    if text:
                        # Draw text background for visibility
                        try:
                            bbox = draw.textbbox(pos_display, text, font=font)
                            padding = 2
                            draw.rectangle(
                                [bbox[0] - padding, bbox[1] - padding,
                                 bbox[2] + padding, bbox[3] + padding],
                                fill=(255, 255, 200, 200)
                            )
                        except:
                            pass
                        draw.text(pos_display, text, fill='red', font=font)

            self.photo = ImageTk.PhotoImage(img)
            self.canvas.delete("all")
            self.canvas.create_image(0, 0, anchor=tk.NW, image=self.photo)
            self.canvas.config(scrollregion=self.canvas.bbox(tk.ALL))
            self.page_label.config(text=f"Page: {self.current_page + 1}/{len(self.pdf_document)}")
            self.sync_manager_stats()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to display page: {e}")

    # ================================================================
    # SAVE SESSION - WITH PROPER SERIALIZATION
    # ================================================================

    def save_session(self):
        """Save current session to JSON file"""
        if not self.pdf_document:
            messagebox.showwarning("No PDF", "Load a PDF first before saving a session.")
            return

        if not hasattr(self, 'project_dirs') or not self.project_dirs.get("sessions"):
            messagebox.showerror("Error", "Project directories not set up. Load a PDF first.")
            return

        save_path = os.path.join(
            self.project_dirs["sessions"],
            f"{self.cabinet_id}_annotations.json"
        )

        data = {
            'project_name': self.project_name,
            'sales_order_no': self.sales_order_no,
            'cabinet_id': getattr(self, 'cabinet_id', ''),
            'pdf_path': self.current_pdf_path,
            'current_page': self.current_page,
            'zoom_level': self.zoom_level,
            'current_sr_no': self.current_sr_no,
            'session_refs': list(self.session_refs),
            'annotations': []
        }

        for ann in self.annotations:
            entry = ann.copy()

            # Serialize bbox_page (rectangles)
            if 'bbox_page' in entry:
                entry['bbox_page'] = [float(x) for x in entry['bbox_page']]

            # Serialize points (pen strokes) - convert tuples to lists
            if 'points' in entry:
                entry['points'] = [[float(x), float(y)] for x, y in entry['points']]

            # Serialize pos_page (text position)
            if 'pos_page' in entry:
                pos = entry['pos_page']
                entry['pos_page'] = [float(pos[0]), float(pos[1])]

            data['annotations'].append(entry)
        self.sync_manager_stats()

        try:
            with open(save_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2)
            messagebox.showinfo("Saved", f"Session saved to:\n{save_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save session: {e}")

    # ================================================================
    # LOAD SESSION - WITH PROPER DESERIALIZATION
    # ================================================================

    def load_session(self):
        """Load session from JSON file via file dialog"""
        path = filedialog.askopenfilename(
            title="Load Session JSON",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if not path:
            return

        self.load_session_from_path(path)
        self.sync_manager_stats()

    def load_session_from_path(self, path):
        """Load session from a specific JSON file path"""
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror("Session Load Error", f"Failed to load session:\n{e}")
            return

        # Restore basic state
        self.project_name = data.get('project_name', self.project_name)
        self.sales_order_no = data.get('sales_order_no', self.sales_order_no)
        self.cabinet_id = data.get('cabinet_id', getattr(self, "cabinet_id", ""))
        self.current_page = data.get('current_page', 0)
        self.zoom_level = data.get('zoom_level', 1.0)
        self.current_sr_no = data.get('current_sr_no', self.current_sr_no)

        # Restore session refs
        self.session_refs = set(data.get('session_refs', []))

        # Restore annotations with proper type conversion
        self.annotations = []
        for entry in data.get('annotations', []):
            ann = entry.copy()

            # Deserialize bbox_page (convert list back to tuple)
            if 'bbox_page' in ann:
                ann['bbox_page'] = tuple(float(x) for x in ann['bbox_page'])

            # Deserialize points (pen strokes - convert lists to tuples)
            if 'points' in ann:
                ann['points'] = [(float(p[0]), float(p[1])) for p in ann['points']]

            # Deserialize pos_page (text position - convert list to tuple)
            if 'pos_page' in ann:
                pos = ann['pos_page']
                ann['pos_page'] = (float(pos[0]), float(pos[1]))

            self.annotations.append(ann)

            # Add ref_no to session refs
            if ann.get('ref_no'):
                self.session_refs.add(str(ann['ref_no']).strip())

        self.display_page()
        messagebox.showinfo("Loaded", f"Session loaded with {len(self.annotations)} annotations.\nMake sure the same PDF is open.")

    # ================================================================
    # EXPORT ANNOTATED PDF - WITH PEN AND TEXT
    # ================================================================

    def export_annotated_pdf(self):
        """Export PDF with all annotations including pen strokes and text"""
        if not self.pdf_document:
            messagebox.showwarning("Warning", "Please load a PDF first")
            return

        if not hasattr(self, 'project_dirs') or not self.project_dirs.get("annotated_drawings"):
            messagebox.showerror("Error", "Project directories not set up.")
            return

        try:
            save_path = os.path.join(
                self.project_dirs["annotated_drawings"],
                f"{self.cabinet_id.replace(' ', '_')}_Annotated.pdf"
            )

            # Create output PDF
            out_doc = fitz.open()
            for pnum in range(len(self.pdf_document)):
                out_doc.insert_pdf(self.pdf_document, from_page=pnum, to_page=pnum)

            # Open Excel for SR No lookup
            wb = None
            ws = None
            if self.excel_file and os.path.exists(self.excel_file):
                try:
                    wb = load_workbook(self.excel_file, data_only=True)
                    ws = wb[self.punch_sheet_name]
                except:
                    pass

            # Draw annotations
            for ann in self.annotations:
                p = ann.get('page')
                if p is None or p < 0 or p >= len(out_doc):
                    continue

                target_page = out_doc[p]
                ann_type = ann.get('type')

                # -------- RECTANGLE ANNOTATIONS --------
                if ann_type in ('ok', 'error') and 'bbox_page' in ann:
                    x1, y1, x2, y2 = ann['bbox_page']
                    rect = self.transform_bbox_for_rotation((x1, y1, x2, y2), target_page)

                    if ann_type == 'ok':
                        target_page.draw_rect(rect, color=(0, 1, 0), width=2)
                    else:
                        target_page.draw_rect(rect, color=(1, 0.55, 0), width=2)

                    sr_text = None
                    row = ann.get('excel_row')

                    if row:
                        try:
                            sr_val = self.read_cell(ws, row, self.punch_cols['sr_no'])
                            if sr_val is not None:
                                sr_text = f"Sr {sr_val}"
                        except:
                            sr_text = None

                    if sr_text:
                        text_pos = fitz.Point(rect.x0, max(rect.y0 - 12, rect.y0))
                        try:
                            target_page.insert_text(text_pos, sr_text, fontsize=8)
                        except:
                            pass

                # -------- PEN STROKES --------
                elif ann_type == 'pen' and 'points' in ann:
                    points = ann['points']
                    if len(points) >= 2:
                        # Transform all points for rotation
                        transformed_points = [
                            self.transform_point_for_rotation(pt, target_page) 
                            for pt in points
                        ]
                        
                        for i in range(len(transformed_points) - 1):
                            p1 = transformed_points[i]
                            p2 = transformed_points[i + 1]
                            target_page.draw_line(p1, p2, color=(1, 0, 0), width=2)

                # -------- TEXT ANNOTATIONS --------
                elif ann_type == 'text' and 'pos_page' in ann:
                    pos = ann['pos_page']
                    text = ann.get('text', '')
                    if text:
                        # Transform text position for rotation
                        text_point = self.transform_point_for_rotation(pos, target_page)
                        try:
                            target_page.insert_text(
                                text_point, text,
                                fontsize=10, color=(1, 0, 0),
                                rotate=target_page.rotation
                            )
                        except:
                            pass

            if wb:
                wb.close()

            out_doc.save(save_path)
            out_doc.close()

            messagebox.showinfo("Success", f"Annotated PDF saved to:\n{save_path}")
            self.sync_manager_stats()

        except PermissionError:
            messagebox.showerror("Error", "Close the target file (if open) and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export annotated PDF:\n{e}")


    def transform_bbox_for_rotation(self, rect, page):
        """Transform bbox for page rotation"""
        r = page.rotation
        w = page.rect.width
        h = page.rect.height
        x1, y1, x2, y2 = rect

        if r == 0:
            return fitz.Rect(x1, y1, x2, y2)
        if r == 90:
            return fitz.Rect(y1, w - x2, y2, w - x1)
        if r == 180:
            return fitz.Rect(w - x2, h - y2, w - x1, h - y1)
        if r == 270:
            return fitz.Rect(h - y2, x1, h - y1, x2)

        return fitz.Rect(x1, y1, x2, y2)


    def transform_point_for_rotation(self, point, page):
        """Transform a single point (x, y) or tuple for page rotation
        
        Args:
            point: tuple (x, y) representing a point coordinate
            page: fitz page object with rotation info
            
        Returns:
            fitz.Point object with transformed coordinates
        """
        r = page.rotation
        w = page.rect.width
        h = page.rect.height
        x, y = point

        if r == 0:
            return fitz.Point(x, y)
        elif r == 90:
            return fitz.Point(y, w - x)
        elif r == 180:
            return fitz.Point(w - x, h - y)
        elif r == 270:
            return fitz.Point(h - y, x)
        
        return fitz.Point(x, y)


    def get_text_position_above_rect(self, rect, page):
        """Get the correct position for text above a rectangle based on page rotation
        
        Args:
            rect: fitz.Rect object (already transformed)
            page: fitz page object with rotation info
            
        Returns:
            fitz.Point object for text position
        """
        r = page.rotation
        offset = 12  # Distance from rectangle edge
        
        if r == 0:
            # Normal orientation - text above top edge
            return fitz.Point(rect.x0, max(rect.y0 - offset, 5))
        elif r == 90:
            # 90° rotation - text to the left of left edge
            return fitz.Point(max(rect.x0 - offset, 5), rect.y0)
        elif r == 180:
            # 180° rotation - text below bottom edge
            return fitz.Point(rect.x0, min(rect.y1 + offset, page.rect.height - 5))
        elif r == 270:
            # 270° rotation - text to the right of right edge
            return fitz.Point(min(rect.x1 + offset, page.rect.width - 5), rect.y0)
        
        # Default fallback
        return fitz.Point(rect.x0, max(rect.y0 - offset, 5))

    # ================================================================
    # UI SETUP
    # ================================================================

    """
    Correct placement for the Handover to Production button in setup_ui()
    Find your existing right_frame section and add the button there
    """

    # ============================================================================
    # UPDATED FUNCTIONS FOR SETUP_UI AND HANDOVER VERIFICATION
    # ============================================================================

    def setup_ui(self):
        """Setup modern professional UI with grouped menu items"""
        
        # Main toolbar with modern styling
        toolbar = tk.Frame(self.root, bg='#1e293b', height=70)
        toolbar.pack(side=tk.TOP, fill=tk.X)
        
        # Enhanced Menu Bar
        menubar = Menu(self.root, bg='#1e293b', fg='white', activebackground='#3b82f6')
        self.root.config(menu=menubar)
        
        # File Menu
        file_menu = Menu(menubar, tearoff=0, bg='#1e293b', fg='white', activebackground='#3b82f6')
        menubar.add_cascade(label="📁 File", menu=file_menu)
        file_menu.add_command(label="Open PDF", command=self.load_pdf, accelerator="Ctrl+O")
        file_menu.add_separator()
        file_menu.add_command(label="Load Session", command=self.load_session, accelerator="Ctrl+L")
        file_menu.add_command(label="Save Session", command=self.save_session, accelerator="Ctrl+S")
        file_menu.add_separator()
        file_menu.add_command(label="Export Annotated PDF", command=self.export_annotated_pdf, accelerator="Ctrl+E")
        file_menu.add_command(label="Save Interphase Excel", command=self.save_interphase_excel)
        file_menu.add_command(label="Open Excel", command=self.open_excel, accelerator="Ctrl+Shift+E")
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)
        
        # Tools Menu
        tools_menu = Menu(menubar, tearoff=0, bg='#1e293b', fg='white', activebackground='#3b82f6')
        menubar.add_cascade(label="🛠️ Tools", menu=tools_menu)
        tools_menu.add_command(label="Review Checklist", command=self.review_checklist_now, accelerator="Ctrl+R")
        tools_menu.add_command(label="Punch Closing Mode", command=self.punch_closing_mode, accelerator="Ctrl+Shift+P")
        tools_menu.add_separator()
        tools_menu.add_command(label="🔍 View Production Handbacks", command=self.view_production_handbacks, accelerator="Ctrl+Shift+V")
        
        # View Menu
        view_menu = Menu(menubar, tearoff=0, bg='#1e293b', fg='white', activebackground='#3b82f6')
        menubar.add_cascade(label="👁️ View", menu=view_menu)
        view_menu.add_command(label="Zoom In", command=self.zoom_in, accelerator="Ctrl++")
        view_menu.add_command(label="Zoom Out", command=self.zoom_out, accelerator="Ctrl+-")
        view_menu.add_command(label="Reset Zoom", command=lambda: setattr(self, 'zoom_level', 1.0) or self.display_page())
        
        # Keyboard shortcuts
        self.root.bind_all("<Control-o>", lambda e: self.load_pdf())
        self.root.bind_all("<Control-s>", lambda e: self.save_session())
        self.root.bind_all("<Control-l>", lambda e: self.load_session())
        self.root.bind_all("<Control-e>", lambda e: self.export_annotated_pdf())
        self.root.bind_all("<Control-r>", lambda e: self.review_checklist_now())
        self.root.bind_all("<Control-Shift-p>", lambda e: self.punch_closing_mode())
        self.root.bind_all("<Control-Shift-e>", lambda e: self.open_excel())
        self.root.bind_all("<Control-Shift-v>", lambda e: self.view_production_handbacks())
        self.root.bind_all("<Control-plus>", lambda e: self.zoom_in())
        self.root.bind_all("<Control-minus>", lambda e: self.zoom_out())
        
        # Modern button style
        btn_style = {
            'bg': '#3b82f6',
            'fg': 'white',
            'padx': 12,
            'pady': 10,
            'font': ('Segoe UI', 9, 'bold'),
            'relief': tk.FLAT,
            'borderwidth': 0,
            'cursor': 'hand2'
        }
        
        # Left section - File operations
        left_frame = tk.Frame(toolbar, bg='#1e293b')
        left_frame.pack(side=tk.LEFT, padx=10, pady=10)
        
        tk.Button(left_frame, text="📁 Open PDF", command=self.load_pdf, **btn_style).pack(side=tk.LEFT, padx=3)
        
        # Recent Projects Dropdown with modern styling
        recent_frame = tk.Frame(left_frame, bg='#1e293b')
        recent_frame.pack(side=tk.LEFT, padx=8)
        
        tk.Label(recent_frame, text="Recent:", bg='#1e293b', fg='#94a3b8', 
                 font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=(0, 5))
        
        self.recent_var = tk.StringVar(value="Select Project...")
        self.recent_dropdown = tk.OptionMenu(recent_frame, self.recent_var, "Select Project...", 
                                            command=self.load_recent_projects_ui)
        self.recent_dropdown.config(bg='#334155', fg='white', font=('Segoe UI', 9), 
                                   width=22, relief=tk.FLAT, borderwidth=0)
        self.recent_dropdown.pack(side=tk.LEFT)
        
        # Center section - Navigation
        center_frame = tk.Frame(toolbar, bg='#1e293b')
        center_frame.pack(side=tk.LEFT, padx=20)
        
        self.page_label = tk.Label(center_frame, text="Page: 0/0", bg='#1e293b', 
                                  fg='white', font=('Segoe UI', 10, 'bold'))
        self.page_label.pack(side=tk.LEFT, padx=10)
        
        nav_btn_style = btn_style.copy()
        nav_btn_style['bg'] = '#64748b'
        
        tk.Button(center_frame, text="◀", command=self.prev_page, width=3, **nav_btn_style).pack(side=tk.LEFT, padx=2)
        tk.Button(center_frame, text="▶", command=self.next_page, width=3, **nav_btn_style).pack(side=tk.LEFT, padx=2)
        
        # Zoom controls
        zoom_frame = tk.Frame(center_frame, bg='#1e293b')
        zoom_frame.pack(side=tk.LEFT, padx=15)
        
        zoom_btn_style = btn_style.copy()
        zoom_btn_style['bg'] = '#10b981'
        
        tk.Button(zoom_frame, text="🔍+", command=self.zoom_in, width=4, **zoom_btn_style).pack(side=tk.LEFT, padx=2)
        tk.Button(zoom_frame, text="🔍−", command=self.zoom_out, width=4, **zoom_btn_style).pack(side=tk.LEFT, padx=2)
        
        # Tool section - Annotation tools
        tool_frame = tk.Frame(toolbar, bg='#1e293b')
        tool_frame.pack(side=tk.LEFT, padx=10)

        tk.Label(tool_frame, text="Tools:", bg='#1e293b', fg='#94a3b8', 
                 font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=(0, 8))

        # UPDATED: Load icons with proper sizing to fill button
        try:
            assets_dir = os.path.join(os.path.dirname(get_app_base_dir()), "assets")
            
            # FIXED: Larger icon size to fill button (44x44 for 40x40 button with padding)
            icon_size = (44, 44)
            
            # Pen icon
            pen_icon_path = os.path.join(assets_dir, "pen_icon.png")
            pen_img = Image.open(pen_icon_path).resize(icon_size, Image.Resampling.LANCZOS)
            self.pen_icon = ImageTk.PhotoImage(pen_img)
            
            # Text icon
            text_icon_path = os.path.join(assets_dir, "text_icon.png")
            text_img = Image.open(text_icon_path).resize(icon_size, Image.Resampling.LANCZOS)
            self.text_icon = ImageTk.PhotoImage(text_img)
            
            # FIXED: Buttons with compound='center' and no padding to fill completely
            self.pen_btn = tk.Button(tool_frame, image=self.pen_icon, 
                                     command=lambda: self.set_tool_mode("pen"),
                                     bg='#334155', width=48, height=48,
                                     relief=tk.FLAT, cursor='hand2',
                                     borderwidth=0, compound='center',
                                     padx=0, pady=0)  # No padding
            self.pen_btn.pack(side=tk.LEFT, padx=2)
            
            self.text_btn = tk.Button(tool_frame, image=self.text_icon,
                                      command=lambda: self.set_tool_mode("text"),
                                      bg='#334155', width=48, height=48,
                                      relief=tk.FLAT, cursor='hand2',
                                      borderwidth=0, compound='center',
                                      padx=0, pady=0)  # No padding
            self.text_btn.pack(side=tk.LEFT, padx=2)
            
        except Exception as e:
            print(f"Could not load tool icons: {e}")
            # Fallback to text-based buttons
            self.pen_btn = tk.Button(tool_frame, text="✏️ Pen", 
                                     command=lambda: self.set_tool_mode("pen"),
                                     bg='#334155', fg='white', width=8, height=1,
                                     font=('Segoe UI', 9, 'bold'), relief=tk.FLAT, 
                                     cursor='hand2', borderwidth=2)
            self.pen_btn.pack(side=tk.LEFT, padx=2)
            
            self.text_btn = tk.Button(tool_frame, text="🅰️ Text", 
                                      command=lambda: self.set_tool_mode("text"),
                                      bg='#334155', fg='white', width=8, height=1,
                                      font=('Segoe UI', 9, 'bold'), relief=tk.FLAT, 
                                      cursor='hand2', borderwidth=2)
            self.text_btn.pack(side=tk.LEFT, padx=2)

        # Add tooltips
        self.create_tooltip(self.pen_btn, "Pen Tool - Draw freehand annotations")
        self.create_tooltip(self.text_btn, "Text Tool - Add text annotations")
        
        # Right section - Action buttons
        right_frame = tk.Frame(toolbar, bg='#1e293b')
        right_frame.pack(side=tk.RIGHT, padx=10, pady=10)
        
        export_btn_style = btn_style.copy()
        export_btn_style['bg'] = '#f59e0b'
        
        handover_btn_style = btn_style.copy()
        handover_btn_style['bg'] = '#8b5cf6'
        
        tk.Button(right_frame, text="🚀 Handover to Production", 
                 command=self.handover_to_production, 
                 **handover_btn_style).pack(side=tk.RIGHT, padx=3)
        
        tk.Button(right_frame, text="📥 Export", 
                 command=self.export_annotated_pdf, 
                 **export_btn_style).pack(side=tk.RIGHT, padx=3)
        
        # Canvas with scrollbars
        canvas_frame = tk.Frame(self.root, bg='#f1f5f9')
        canvas_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        v_scrollbar = tk.Scrollbar(canvas_frame, orient=tk.VERTICAL)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar = tk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.canvas = tk.Canvas(canvas_frame, bg='#f8fafc',
                               yscrollcommand=v_scrollbar.set,
                               xscrollcommand=h_scrollbar.set,
                               highlightthickness=0)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        v_scrollbar.config(command=self.canvas.yview)
        h_scrollbar.config(command=self.canvas.xview)
        
        # Bind mouse events
        self.canvas.bind("<ButtonPress-1>", self.on_left_press)
        self.canvas.bind("<B1-Motion>", self.on_left_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_left_release)
        
        self.canvas.bind("<ButtonPress-3>", self.on_right_press)
        self.canvas.bind("<B3-Motion>", self.on_right_drag)
        self.canvas.bind("<ButtonRelease-3>", self.on_right_release)
        
        self.canvas.bind("<Control-Button-1>", self.on_ctrl_click)
        self.root.bind("<Delete>", self.delete_selected_annotation)
        
        self.canvas.bind("<Double-Button-1>", self.on_double_left_zoom)
        self.canvas.bind("<Double-Button-3>", self.on_double_right_zoom)
        self.root.bind("<Escape>", self.exit_tool_mode)
        
        # Modern status bar
        status_bar = tk.Frame(self.root, bg='#334155', height=40)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        instructions_text = "🖱️ Left Click: OK (Green) | Right Click: Error (Orange) | Ctrl+Click: Select | Delete: Remove"
        tk.Label(status_bar, text=instructions_text, bg='#334155', fg='#e2e8f0', 
                 font=('Segoe UI', 9), pady=10).pack()


    def create_tooltip(self, widget, text):
        """Create a simple tooltip for a widget"""
        def on_enter(event):
            tooltip = tk.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
            label = tk.Label(tooltip, text=text, background="#1e293b", foreground="white",
                            relief=tk.SOLID, borderwidth=1, font=("Segoe UI", 8), padx=5, pady=3)
            label.pack()
            widget.tooltip = tooltip
            
        def on_leave(event):
            if hasattr(widget, 'tooltip'):
                widget.tooltip.destroy()
                del widget.tooltip
            
        widget.bind("<Enter>", on_enter)

        widget.bind("<Leave>", on_leave)



    # ================================================================
    # RECENT PROJECTS MANAGEMENT
    # ================================================================
    def save_recent_project(self):
        """Save current project to database with storage location"""
        if not self.current_pdf_path or not self.excel_file:
            return
        
        try:
            # Create session path
            session_path = os.path.join(
                self.project_dirs.get("sessions", ""),
                f"{self.cabinet_id}_annotations.json"
            ) if hasattr(self, 'project_dirs') else None
            
            # Update or add project in database
            project_data = {
                'cabinet_id': self.cabinet_id,
                'project_name': self.project_name,
                'sales_order_no': self.sales_order_no,
                'storage_location': self.storage_location,  # IMPORTANT: Store location
                'pdf_path': self.current_pdf_path,
                'excel_path': self.excel_file,
                'session_path': session_path if session_path and os.path.exists(session_path) else None,
                'last_accessed': datetime.now().isoformat()
            }
            
            if self.db.project_exists(self.cabinet_id):
                # Update existing
                self.db.update_project(self.cabinet_id, project_data)
            else:
                # Add new
                project_data['created_date'] = datetime.now().isoformat()
                self.db.add_project(project_data)
            
            # Update dropdown
            self.update_recent_dropdown()
            self.sync_manager_stats()
            
        except Exception as e:
            print(f"Error saving recent project: {e}")
    
    def load_recent_projects_ui(self):
        """Load and display recent projects from SQLite"""
        self.update_recent_dropdown()

    def update_recent_dropdown(self):
        """Update the recent projects dropdown from database"""
        try:
            recent_projects = self.db.get_recent_projects(limit=20)
            
            menu = self.recent_dropdown['menu']
            menu.delete(0, 'end')
            
            if not recent_projects:
                menu.add_command(label="No recent projects", command=lambda: None)
                return
            
            for proj in recent_projects:
                label = f"{proj['cabinet_id']} - {proj['project_name']}"
                menu.add_command(
                    label=label,
                    command=lambda p=proj: self.load_recent_project_from_db(p)
                )
                
        except Exception as e:
            print(f"Error updating recent dropdown: {e}")

    def load_recent_project_from_db(self, project_data):
        """Load a recent project from database"""
        try:
            # Set project details FIRST
            self.cabinet_id = project_data['cabinet_id']
            self.project_name = project_data['project_name']
            self.sales_order_no = project_data.get('sales_order_no', '')
            self.storage_location = project_data['storage_location']
            
            # Prepare folders (this creates the proper folder structure)
            self.prepare_project_folders()
            
            # Now construct the CORRECT paths based on folder structure
            expected_excel_path = os.path.join(
                self.project_dirs["working_excel"],
                f"{self.cabinet_id.replace(' ', '_')}_Working.xlsx"
            )
            
            expected_session_path = os.path.join(
                self.project_dirs["sessions"],
                f"{self.cabinet_id}_annotations.json"
            )
            
            # Check PDF path from database
            pdf_path = project_data.get('pdf_path')
            if not pdf_path or not os.path.exists(pdf_path):
                messagebox.showerror("Error", 
                                   f"PDF file not found:\n{pdf_path}\n\n"
                                   "The file may have been moved or deleted.")
                return
            
            # Use expected Excel path, not database path
            if not os.path.exists(expected_excel_path):
                # Try to find it in the old database path first
                old_excel_path = project_data.get('excel_path')
                if old_excel_path and os.path.exists(old_excel_path):
                    # Copy from old location to new location
                    try:
                        import shutil
                        shutil.copy2(old_excel_path, expected_excel_path)
                        messagebox.showinfo("Excel Migrated", 
                                          f"Excel file migrated to new location:\n{expected_excel_path}")
                    except Exception as e:
                        messagebox.showerror("Error", 
                                           f"Excel file not found and couldn't migrate:\n{e}")
                        return
                else:
                    messagebox.showerror("Error", 
                                       f"Excel file not found at:\n{expected_excel_path}\n\n"
                                       "The file may have been moved or deleted.")
                    return
            
            # Load PDF
            self.pdf_document = fitz.open(pdf_path)
            self.current_pdf_path = pdf_path
            self.current_page = 0
            self.annotations = []
            self.zoom_level = 1.0
            self.tool_mode = None
            self.root.config(cursor="")
            
            # Set Excel file to expected path
            self.excel_file = expected_excel_path
            self.working_excel_path = expected_excel_path
            
            # Get next SR number
            self.current_sr_no = self.get_next_sr_no()
            
            # Load session if available at expected location
            if os.path.exists(expected_session_path):
                self.load_session_from_path(expected_session_path)
            else:
                # Try old session path from database
                old_session_path = project_data.get('session_path')
                if old_session_path and os.path.exists(old_session_path):
                    # Copy to new location
                    try:
                        import shutil
                        shutil.copy2(old_session_path, expected_session_path)
                        self.load_session_from_path(expected_session_path)
                    except:
                        self.display_page()
                else:
                    self.display_page()
            
            # Update database with correct paths
            self.db.update_project(self.cabinet_id, {
                'pdf_path': self.current_pdf_path,
                'excel_path': expected_excel_path,
                'session_path': expected_session_path if os.path.exists(expected_session_path) else None,
                'last_accessed': datetime.now().isoformat()
            })
            
            messagebox.showinfo(
                "Project Loaded",
                f"✓ Loaded: {self.cabinet_id}\n"
                f"Project: {self.project_name}\n"
                f"Location: {self.storage_location}\n"
                f"Pages: {len(self.pdf_document)}"
            )
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load project:\n{e}")
            import traceback
            traceback.print_exc()

    # ================================================================
    # CELL HELPERS
    # ================================================================

    def split_cell(self, cell_ref):
        """Splits 'F6' -> (6, 'F')"""
        m = re.match(r"([A-Z]+)(\d+)", cell_ref)
        if not m:
            raise ValueError(f"Invalid cell reference: {cell_ref}")
        col, row = m.groups()
        return int(row), col

    def _resolve_merged_target(self, ws, row, col_idx):
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col_idx <= merged.max_col:
                return merged.min_row, merged.min_col
        return row, col_idx

    def write_cell(self, ws, row, col, value):
        if isinstance(col, str):
            col_idx = column_index_from_string(col)
        else:
            col_idx = int(col)
        target_row, target_col = self._resolve_merged_target(ws, int(row), col_idx)
        ws.cell(row=target_row, column=target_col).value = value

    def read_cell(self, ws, row, col):
        if isinstance(col, str):
            col_idx = column_index_from_string(col)
        else:
            col_idx = int(col)
        target_row, target_col = self._resolve_merged_target(ws, int(row), col_idx)
        return ws.cell(row=target_row, column=target_col).value
    
    # ================================================================
    # Add these methods to CircuitInspector class for exit tool mode
    # ================================================================

    def exit_tool_mode(self, event=None):
        """Exit pen/text tool mode and return to normal with visual feedback"""
        if self.tool_mode:
            self.tool_mode = None
            self.root.config(cursor="")
            self.pen_btn.config(bg='#334155', relief=tk.FLAT)
            self.text_btn.config(bg='#334155', relief=tk.FLAT)
            self.pen_points.clear()
            self.clear_temp_drawings()
            self.drawing = False
            self.drawing_type = None
            self._flash_status("Normal mode", bg='#64748b')

    # ================================================================
    # PDF LOADING
    # ================================================================

    def load_pdf(self):
        file_path = filedialog.askopenfilename(
            title="Select Circuit Diagram PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if file_path:
            try:
                self.pdf_document = fitz.open(file_path)
                self.current_pdf_path = file_path
                self.current_page = 0
                self.annotations = []
                self.zoom_level = 1.0
                self.tool_mode = None
                self.root.config(cursor="")
                self.current_sr_no = self.get_next_sr_no()
                self.display_page()
                messagebox.showinfo("Success", f"Loaded PDF with {len(self.pdf_document)} pages")
                
                # Ask project details (will auto-fill if project exists)
                self.ask_project_details()
                self.prepare_project_folders()

                # Session working Excel
                try:
                    self.working_excel_path = os.path.join(
                        self.project_dirs["working_excel"],
                        f"{self.cabinet_id.replace(' ', '_')}_Working.xlsx"
                    )

                    if os.path.exists(self.working_excel_path):
                        resume = messagebox.askyesno(
                            "Resume Inspection",
                            f"Existing working Excel found:\n\n{os.path.basename(self.working_excel_path)}\n\nResume previous inspection?"
                        )
                        if not resume:
                            shutil.copy2(self.master_excel_file, self.working_excel_path)
                    else:
                        shutil.copy2(self.master_excel_file, self.working_excel_path)

                    self.excel_file = self.working_excel_path

                except Exception as e:
                    messagebox.showerror("Excel Error", f"Failed to prepare working Excel:\n{e}")
                    return

                self.write_project_details_to_excel()

                # Get expected session path
                expected_session_path = os.path.join(
                    self.project_dirs["sessions"],
                    f"{self.cabinet_id}_annotations.json"
                )

                # Update database with CORRECT paths immediately
                self.db.update_project(self.cabinet_id, {
                    'pdf_path': self.current_pdf_path,
                    'excel_path': self.excel_file,
                    'session_path': expected_session_path if os.path.exists(expected_session_path) else None,
                    'storage_location': self.storage_location,
                    'last_accessed': datetime.now().isoformat()
                })

                # Auto load session if exists
                if os.path.exists(expected_session_path):
                    resume = messagebox.askyesno(
                        "Resume Session",
                        f"Existing session found for this drawing:\n\n"
                        f"{os.path.basename(expected_session_path)}\n\n"
                        "Do you want to resume it?"
                    )
                    if resume:
                        self.load_session_from_path(expected_session_path)
                
                # Save to recent projects
                self.save_recent_project()

            except Exception as e:
                messagebox.showerror("Error", f"Failed to load PDF: {str(e)}")

    def get_next_sr_no(self):
        try:
            if not self.excel_file or not os.path.exists(self.excel_file):
                return 1
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
            last_sr_no = 0
            row_num = 8
            while row_num <= ws.max_row + 5:
                val = self.read_cell(ws, row_num, self.punch_cols['sr_no'])
                if val is None:
                    break
                try:
                    last_sr_no = int(val)
                except:
                    pass
                row_num += 1
            wb.close()
            return last_sr_no + 1
        except Exception:
            return 1

    # ================================================================
    # ZOOM FUNCTIONS
    # ================================================================

    def zoom_in(self):
        if self.zoom_level < 3.0:
            self.zoom_level += 0.25
            self.display_page()

    def zoom_out(self):
        if self.zoom_level > 0.5:
            self.zoom_level -= 0.25
            self.display_page()

    def zoom_at_point(self, canvas_x, canvas_y, zoom_delta):
        if not self.pdf_document:
            return

        old_zoom = self.zoom_level
        new_zoom = max(0.5, min(3.0, old_zoom + zoom_delta))

        if new_zoom == old_zoom:
            return

        self.zoom_level = new_zoom
        self.display_page()

        scale = new_zoom / old_zoom
        bbox = self.canvas.bbox("all")
        if not bbox:
            return

        self.canvas.xview_moveto((canvas_x * scale) / max(1, bbox[2]))
        self.canvas.yview_moveto((canvas_y * scale) / max(1, bbox[3]))

    def on_double_left_zoom(self, event):
        self.drawing = False
        self.temp_rect_id = None
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        self.zoom_at_point(x, y, +0.25)

    def on_double_right_zoom(self, event):
        self.drawing = False
        self.temp_rect_id = None
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        self.zoom_at_point(x, y, -0.25)

    # ================================================================
    # PAGE NAVIGATION
    # ================================================================

    def prev_page(self):
        if self.pdf_document and self.current_page > 0:
            self.current_page -= 1
            self.display_page()

    def next_page(self):
        if self.pdf_document and self.current_page < len(self.pdf_document) - 1:
            self.current_page += 1
            self.display_page()

    # ================================================================
    # SELECTION / DELETE
    # ================================================================

    def on_ctrl_click(self, event):
        if not self.pdf_document:
            return
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)

        for ann in reversed(self.annotations):
            if ann.get('page') != self.current_page:
                continue

            ann_type = ann.get('type')

            # Check rectangles
            if 'bbox_page' in ann:
                x1d, y1d, x2d, y2d = self.bbox_page_to_display(ann['bbox_page'])
                if ann_type == 'ok':
                    cx = (x1d + x2d) / 2
                    cy = (y1d + y2d) / 2
                    radius = max(abs(x2d - x1d), abs(y2d - y1d)) / 2
                    inside = ((x - cx) ** 2 + (y - cy) ** 2) ** 0.5 <= radius
                else:
                    inside = (x1d <= x <= x2d) and (y1d <= y <= y2d)

                if inside:
                    self.selected_annotation = None if self.selected_annotation is ann else ann
                    self.display_page()
                    return

            # Check pen strokes (bounding box)
            elif ann_type == 'pen' and 'points' in ann:
                points_display = self.page_to_display_coords(ann['points'])
                if len(points_display) >= 2:
                    xs = [p[0] for p in points_display]
                    ys = [p[1] for p in points_display]
                    if min(xs) - 10 <= x <= max(xs) + 10 and min(ys) - 10 <= y <= max(ys) + 10:
                        self.selected_annotation = None if self.selected_annotation is ann else ann
                        self.display_page()
                        return

            # Check text annotations
            elif ann_type == 'text' and 'pos_page' in ann:
                pos_display = self.page_to_display_coords(ann['pos_page'])
                if abs(x - pos_display[0]) < 50 and abs(y - pos_display[1]) < 20:
                    self.selected_annotation = None if self.selected_annotation is ann else ann
                    self.display_page()
                    return

        self.selected_annotation = None
        self.display_page()

    def delete_selected_annotation(self, event=None):
        if not self.selected_annotation:
            messagebox.showinfo("No Selection", "Please select an annotation first by Ctrl+Clicking on it")
            return
        if messagebox.askyesno("Delete Annotation", "Are you sure you want to delete this annotation?"):
            try:
                self.annotations.remove(self.selected_annotation)
            except:
                pass
            self.selected_annotation = None
            self.display_page()
            messagebox.showinfo("Deleted", "Annotation removed successfully")

    # ================================================================
    # INTERPHASE UPDATE
    # ================================================================

    def update_interphase_status_for_ref(self, ref_no, status='NOK'):
        """Update Interphase status, name, and date for a reference number"""
        try:
            wb = load_workbook(self.excel_file)
            if self.interphase_sheet_name not in wb.sheetnames:
                wb.close()
                return False
            ws = wb[self.interphase_sheet_name]
            ref_col = self.interphase_cols['ref_no']
            status_col = self.interphase_cols['status']
            name_col = self.interphase_cols['name']
            date_col = self.interphase_cols['date']
            
            updated_any = False
            max_r = ws.max_row if ws.max_row else 2000
            current_date = datetime.now().strftime("%Y-%m-%d")
            
            # Get username
            try:
                username = os.getlogin()
            except:
                username = getpass.getuser()
            
            for r in range(1, max_r + 1):
                cell_val = self.read_cell(ws, r, ref_col)
                if cell_val is None:
                    continue
                if str(cell_val).strip() == str(ref_no).strip():
                    self.write_cell(ws, r, status_col, status)
                    self.write_cell(ws, r, name_col, username)
                    self.write_cell(ws, r, date_col, current_date)
                    updated_any = True
            
            if updated_any:
                wb.save(self.excel_file)
            wb.close()
            return updated_any
        except Exception as e:
            try:
                wb.close()
            except:
                pass
            print("update_interphase_status_for_ref error:", e)
            return False

    # ================================================================
    # ERROR LOGGING
    # ================================================================

    def log_error(self, component_type, error_name, error_template, bbox_page, tag_name):
        """Logs an error punch into Excel and stores annotation."""
        punch_text = error_template

        if not punch_text:
            messagebox.showerror("Error", "Punch description is empty.")
            return

        ref_no = simpledialog.askstring("Reference No", "Enter Reference Number:", parent=self.root)
        if not ref_no:
            return

        ref_no = str(ref_no).strip()
        self.session_refs.add(ref_no)

        try:
            wb = load_workbook(self.excel_file)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active

            row_num = 8
            while True:
                val = self.read_cell(ws, row_num, self.punch_cols['sr_no'])
                if val is None:
                    break
                row_num += 1

            prev_sr = None
            if row_num > 8:
                prev_sr = self.read_cell(ws, row_num - 1, self.punch_cols['sr_no'])

            try:
                sr_no_assigned = int(prev_sr) + 1 if prev_sr is not None else 1
            except:
                sr_no_assigned = 1

            self.write_cell(ws, row_num, self.punch_cols['sr_no'], sr_no_assigned)
            self.write_cell(ws, row_num, self.punch_cols['ref_no'], ref_no)
            self.write_cell(ws, row_num, self.punch_cols['desc'], punch_text)
            self.write_cell(ws, row_num, self.punch_cols['category'], component_type)

            try:
                uname = os.getlogin()
            except:
                uname = getpass.getuser()

            self.write_cell(ws, row_num, self.punch_cols['checked_name'], uname)
            self.write_cell(ws, row_num, self.punch_cols['checked_date'], datetime.now().strftime("%Y-%m-%d"))

            wb.save(self.excel_file)
            wb.close()

            updated = self.update_interphase_status_for_ref(ref_no, status='NOK')
            if updated:
                print(f"Interphase: marked ref {ref_no} as NOK")

            ann = {
                'type': 'error',
                'page': self.current_page,
                'bbox_page': bbox_page,
                'component': component_type,
                'subcategory': error_name,
                'punch_text': punch_text,
                'ref_no': ref_no,
                'excel_row': row_num,
                'sr_no': sr_no_assigned,
                'implemented': False,
                'implemented_name': None,
                'implemented_date': None,
                'implementation_remark': None,
                'timestamp': datetime.now().isoformat()
            }

            self.annotations.append(ann)
            self.current_sr_no = self.get_next_sr_no()
            self.display_page()

            self.root.after(100, lambda: messagebox.showinfo("Logged", f"Punch logged:\n{punch_text}"))
            try:
                self.manager_db.log_category_occurrence(
                    self.cabinet_id,
                    self.project_name,
                    component_type,
                    error_name
                )
                self.sync_manager_stats()
            except Exception as e:
                print(f"Manager category logging failed: {e}")

        except PermissionError:
            messagebox.showerror("Error", "Close the Excel file before writing to it.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to log punch:\n{e}")

    def log_custom_error(self, bbox_page, tag_name):
        try:
            custom_action = simpledialog.askstring(
                "Custom Action Point",
                "Enter the action point / punch description:",
                parent=self.root
            )
            if not custom_action:
                return

            custom_category = simpledialog.askstring(
                "Custom Category",
                "Enter the category:",
                parent=self.root
            )
            if not custom_category:
                return

            ref_no = simpledialog.askstring("Reference No", "Enter Reference No:", parent=self.root)
            if not ref_no:
                messagebox.showwarning("Reference Required", "Reference No is required for custom action points.")
                return

            ref_no = str(ref_no).strip()
            self.session_refs.add(ref_no)

            wb = load_workbook(self.excel_file)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active

            row_num = 8
            while True:
                val = self.read_cell(ws, row_num, self.punch_cols['sr_no'])
                if val is None:
                    break
                row_num += 1

            prev_sr = None
            if row_num > 8:
                prev_sr = self.read_cell(ws, row_num - 1, self.punch_cols['sr_no'])

            try:
                sr_no_assigned = int(prev_sr) + 1 if prev_sr is not None else 1
            except:
                sr_no_assigned = 1

            self.write_cell(ws, row_num, self.punch_cols['sr_no'], sr_no_assigned)
            self.write_cell(ws, row_num, self.punch_cols['ref_no'], ref_no)
            self.write_cell(ws, row_num, self.punch_cols['desc'], custom_action)
            self.write_cell(ws, row_num, self.punch_cols['category'], custom_category)

            try:
                uname = os.getlogin()
            except:
                uname = getpass.getuser()

            self.write_cell(ws, row_num, self.punch_cols['checked_name'], uname)
            self.write_cell(ws, row_num, self.punch_cols['checked_date'], datetime.now().strftime("%Y-%m-%d"))

            wb.save(self.excel_file)
            wb.close()

            ann = {
                'type': 'error',
                'page': self.current_page,
                'bbox_page': bbox_page,
                'component': custom_category,
                'tag_name': tag_name,
                'error': 'Custom',
                'punch_text': custom_action,
                'ref_no': ref_no,
                'excel_row': row_num,
                'sr_no': sr_no_assigned,
                'timestamp': datetime.now().isoformat()
            }
            self.annotations.append(ann)

            self.current_sr_no = self.get_next_sr_no()
            self.display_page()

            self.root.after(100, lambda: messagebox.showinfo("Logged", f"Custom punch logged:\n{custom_action}"))

            try:
                self.manager_db.log_category_occurrence(
                    self.cabinet_id,
                    self.project_name,
                    custom_category,
                    None
                )
                self.sync_manager_stats()
            except Exception as e:
                print(f"Manager category logging failed: {e}")

            updated = self.update_interphase_status_for_ref(ref_no, status='NOK')
            if updated:
                print(f"Interphase: marked ref {ref_no} as NOK")

        except PermissionError:
            messagebox.showerror("Error", "Close the Excel file before writing to it.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to log custom error:\n{e}")

    # ================================================================
    # CATEGORY HANDLERS
    # ================================================================

    def handle_template_category(self, category, bbox_page):
        punch_text = self.run_template(category, tag_name=None)
        if not punch_text:
            return

        self.log_error(
            component_type=category["name"],
            error_name=None,
            error_template=punch_text,
            bbox_page=bbox_page,
            tag_name=None
        )

    def handle_subcategory(self, category, subcategory, bbox_page):
        punch_text = self.run_template(subcategory, tag_name=None)
        if not punch_text:
            return

        self.log_error(
            component_type=category["name"],
            error_name=subcategory["name"],
            error_template=punch_text,
            bbox_page=bbox_page,
            tag_name=None
        )

    def run_template(self, template_def, tag_name=None):
        """Execute a template definition at runtime."""
        values = {}

        if tag_name:
            values["tag"] = tag_name

        for inp in template_def.get("inputs", []):
            val = simpledialog.askstring("Input Required", inp["label"], parent=self.root)
            if not val:
                return None
            values[inp["name"]] = val.strip()

        try:
            return template_def["template"].format(**values)
        except KeyError as e:
            messagebox.showerror("Template Error", f"Missing placeholder: {e}")
            return None

    def load_categories(self):
        try:
            if os.path.exists(self.category_file):
                with open(self.category_file, "r", encoding="utf-8") as f:
                    self.categories = json.load(f)
                    print("Categories loaded:", self.categories)
            else:
                print(f"Warning: categories.json not found at {self.category_file}")
                self.categories = []
        except Exception as e:
            print(f"Error loading categories: {e}")
            self.categories = []


    # ================================================================
    # UPDATED: punch_closing_mode - Modern Dialog UI
    # ================================================================

    def punch_closing_mode(self):
        """Modern dialog for punch closing workflow"""
        punches = self.read_open_punches_from_excel()

        if not punches:
            messagebox.showinfo("No Open Punches", 
                              "✓ All punches are closed!\nNo items require attention.",
                              icon='info')
            return

        punches.sort(key=lambda p: (not p['implemented'], p['sr_no']))

        # Modern dialog window
        dlg = tk.Toplevel(self.root)
        dlg.title("Punch Closing Mode")
        dlg.geometry("950x600")  # Slightly taller
        dlg.configure(bg='#f8fafc')
        dlg.transient(self.root)
        dlg.grab_set()
        
        # Header - REDUCED HEIGHT
        header_frame = tk.Frame(dlg, bg='#1e293b', height=50)  # Reduced from 60
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="✓ Punch Closing Mode", 
                bg='#1e293b', fg='white', 
                font=('Segoe UI', 13, 'bold')).pack(pady=12)  # Reduced padding
        
        # Progress - REDUCED PADDING
        progress_frame = tk.Frame(dlg, bg='#f8fafc')
        progress_frame.pack(fill=tk.X, padx=20, pady=(10, 5))  # Reduced top padding
        
        idx_label = tk.Label(progress_frame, text="", font=('Segoe UI', 10, 'bold'),
                            bg='#f8fafc', fg='#1e293b')
        idx_label.pack()
        
        # ORIGINAL Info cards - KEPT AS BEFORE
        info_frame = tk.Frame(dlg, bg='#f8fafc')
        info_frame.pack(fill=tk.X, padx=20, pady=8)  # Slightly reduced padding
        
        # SR Number card
        sr_card = tk.Frame(info_frame, bg='#dbeafe', relief=tk.FLAT)
        sr_card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        tk.Label(sr_card, text="SR No.", font=('Segoe UI', 8), 
                bg='#dbeafe', fg='#1e40af').pack(anchor='w', padx=10, pady=(6, 2))  # Reduced padding
        sr_label = tk.Label(sr_card, text="", font=('Segoe UI', 12, 'bold'),
                           bg='#dbeafe', fg='#1e293b')
        sr_label.pack(anchor='w', padx=10, pady=(0, 6))  # Reduced padding
        
        # Reference card
        ref_card = tk.Frame(info_frame, bg='#e0e7ff', relief=tk.FLAT)
        ref_card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        tk.Label(ref_card, text="Reference", font=('Segoe UI', 8), 
                bg='#e0e7ff', fg='#4338ca').pack(anchor='w', padx=10, pady=(6, 2))
        ref_label = tk.Label(ref_card, text="", font=('Segoe UI', 12, 'bold'),
                            bg='#e0e7ff', fg='#1e293b')
        ref_label.pack(anchor='w', padx=10, pady=(0, 6))
        
        # Status card
        status_card = tk.Frame(info_frame, bg='#fef3c7', relief=tk.FLAT)
        status_card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
        
        tk.Label(status_card, text="Status", font=('Segoe UI', 8), 
                bg='#fef3c7', fg='#92400e').pack(anchor='w', padx=10, pady=(6, 2))
        impl_label = tk.Label(status_card, text="", font=('Segoe UI', 12, 'bold'),
                             bg='#fef3c7', fg='#1e293b')
        impl_label.pack(anchor='w', padx=10, pady=(0, 6))
        
        # Content - REDUCED HEIGHT
        content_frame = tk.Frame(dlg, bg='white', relief=tk.FLAT)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=8)  # Reduced padding
        
        tk.Label(content_frame, text="Punch Description:", font=('Segoe UI', 9, 'bold'),
                bg='white', fg='#64748b', anchor='w').pack(fill=tk.X, padx=15, pady=(8, 3))
        
        # REDUCED text widget height
        text_widget = tk.Text(content_frame, wrap=tk.WORD, height=9,  # Reduced from 14 to 9
                             font=('Segoe UI', 10), bg='#f8fafc',
                             relief=tk.FLAT, padx=10, pady=8)
        text_widget.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 10))
        text_widget.config(state=tk.DISABLED)

        pos = [0]

        def show_item():
            p = punches[pos[0]]
            
            # Update progress
            progress_text = f"Item {pos[0]+1} of {len(punches)}"
            progress_pct = f"({int((pos[0]+1)/len(punches)*100)}% complete)"
            idx_label.config(text=f"{progress_text} {progress_pct}")
            
            # Update info cards
            sr_label.config(text=str(p['sr_no']))
            ref_label.config(text=str(p['ref_no']))
            
            impl_status = "✓ Implemented" if p['implemented'] else "⚠ Not Implemented"
            impl_color = '#10b981' if p['implemented'] else '#f59e0b'
            impl_label.config(text=impl_status, fg=impl_color)
            
            # Update description
            text_widget.config(state=tk.NORMAL)
            text_widget.delete("1.0", tk.END)
            text_widget.insert(tk.END, p['punch_text'])
            text_widget.insert(tk.END, f"\n\n──────────────────\n")
            text_widget.insert(tk.END, f"Category: {p['category']}\n")

            # Find annotation for implementation remarks
            ann = next((a for a in self.annotations if a.get('sr_no') == p['sr_no']), None)
            if ann and ann.get('implementation_remark'):
                text_widget.insert(tk.END, f"\n──────────────────\n")
                text_widget.insert(tk.END, "Implementation Remarks:\n")
                text_widget.insert(tk.END, ann['implementation_remark'])
            
            # NEW: Check for production remarks from handover database
            try:
                # Try to find production remarks for this cabinet
                handover_data = self.handover_db.get_handover_by_cabinet(self.cabinet_id)
                if handover_data and handover_data.get('production_remarks'):
                    text_widget.insert(tk.END, f"\n\n──────────────────\n")
                    text_widget.insert(tk.END, "🔧 Production Remarks:\n")
                    text_widget.insert(tk.END, handover_data['production_remarks'])
                    text_widget.insert(tk.END, f"\n\nRework by: {handover_data.get('rework_completed_by', 'N/A')}")
                    text_widget.insert(tk.END, f"\nDate: {handover_data.get('rework_completed_date', 'N/A')[:10]}")
            except Exception as e:
                print(f"Could not load production remarks: {e}")

            text_widget.config(state=tk.DISABLED)

        show_item()

        def close_punch():
            p = punches[pos[0]]

            try:
                default_user = os.getlogin()
            except:
                default_user = getpass.getuser()

            name = simpledialog.askstring("Closed By", 
                                         "Enter your name to close this punch:", 
                                         initialvalue=default_user, 
                                         parent=dlg)
            if not name:
                return

            try:
                wb = load_workbook(self.excel_file)
                ws = wb[self.punch_sheet_name]

                self.write_cell(ws, p['row'], self.punch_cols['closed_name'], name)
                self.write_cell(ws, p['row'], self.punch_cols['closed_date'], 
                              datetime.now().strftime("%Y-%m-%d"))

                wb.save(self.excel_file)
                wb.close()

            except PermissionError:
                messagebox.showerror("File Locked", 
                                   "⚠️ Please close the Excel file and try again.")
                return
            except Exception as e:
                messagebox.showerror("Error", f"Failed to close punch:\n{e}")
                return

            ann = next((a for a in self.annotations if a.get('excel_row') == p['row']), None)
            if ann:
                ann['type'] = 'ok'
                ann['closed_by'] = name
                ann['closed_date'] = datetime.now().strftime("%Y-%m-%d")

            self.display_page()

            if pos[0] < len(punches) - 1:
                pos[0] += 1
                show_item()
            else:
                messagebox.showinfo("Complete", 
                                  f"✓ All punches closed!\n{len(punches)} items processed.",
                                  icon='info')
                dlg.destroy()

        def next_item():
            if pos[0] < len(punches) - 1:
                pos[0] += 1
                show_item()

        def prev_item():
            if pos[0] > 0:
                pos[0] -= 1
                show_item()

        # EXPANDED: Button frame with MORE SPACE
        btn_frame = tk.Frame(dlg, bg='#f8fafc', height=80)  # Fixed height
        btn_frame.pack(fill=tk.X, padx=20, pady=(10, 25))  # More bottom padding
        btn_frame.pack_propagate(False)  # Prevent shrinking
        
        # Button container centered vertically
        btn_container = tk.Frame(btn_frame, bg='#f8fafc')
        btn_container.pack(expand=True)
        
        # LARGER button style
        btn_style = {
            'font': ('Segoe UI', 12, 'bold'),  # Bigger font
            'relief': tk.FLAT,
            'borderwidth': 0,
            'cursor': 'hand2',
            'padx': 35,   # More horizontal padding
            'pady': 18,   # More vertical padding
            'width': 15   # Fixed width in characters
        }

        # Create buttons with consistent sizing
        tk.Button(btn_container, text="◀  Previous", command=prev_item, 
                 bg='#94a3b8', fg='white', **btn_style).pack(side=tk.LEFT, padx=8)
        
        # Main action button - slightly larger
        close_btn_style = btn_style.copy()
        close_btn_style['width'] = 18
        tk.Button(btn_container, text="✓  CLOSE PUNCH", command=close_punch, 
                 bg='#10b981', fg='white', **close_btn_style).pack(side=tk.LEFT, padx=8)
        
        tk.Button(btn_container, text="Next  ▶", command=next_item, 
                 bg='#94a3b8', fg='white', **btn_style).pack(side=tk.LEFT, padx=8)
        
        tk.Button(btn_container, text="Cancel", command=dlg.destroy, 
                 bg='#64748b', fg='white', **btn_style).pack(side=tk.LEFT, padx=8)

        dlg.wait_window()
    def read_open_punches_from_excel(self):
        """Reads punch sheet and returns list of open punches."""
        punches = []

        if not self.excel_file or not os.path.exists(self.excel_file):
            return punches

        wb = load_workbook(self.excel_file, data_only=True)
        ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active

        row = 8
        while True:
            sr = self.read_cell(ws, row, self.punch_cols['sr_no'])
            if sr is None:
                break

            closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
            if closed:
                row += 1
                continue

            implemented = bool(self.read_cell(ws, row, self.punch_cols['implemented_name']))

            punches.append({
                'sr_no': sr,
                'row': row,
                'ref_no': self.read_cell(ws, row, self.punch_cols['ref_no']),
                'punch_text': self.read_cell(ws, row, self.punch_cols['desc']),
                'category': self.read_cell(ws, row, self.punch_cols['category']),
                'implemented': implemented
            })

            row += 1

        wb.close()
        self.sync_manager_stats()
        return punches

    # ================================================================
    # PROJECT DETAILS
    # ================================================================

    def ask_project_details(self):
        """Ask for project details including storage location"""
        dlg = tk.Toplevel(self.root)
        dlg.title("Project Details")
        dlg.geometry("500x380")
        dlg.transient(self.root)
        dlg.grab_set()

        tk.Label(dlg, text="Cabinet ID", font=('Segoe UI', 10, 'bold')).pack(anchor="w", padx=20, pady=(15, 0))
        cabinet_var = tk.StringVar(value=getattr(self, "cabinet_id", ""))
        tk.Entry(dlg, textvariable=cabinet_var, font=('Segoe UI', 10)).pack(fill="x", padx=20)

        tk.Label(dlg, text="Project Name", font=('Segoe UI', 10, 'bold')).pack(anchor="w", padx=20, pady=(10, 0))
        project_var = tk.StringVar(value=self.project_name)
        project_entry = tk.Entry(dlg, textvariable=project_var, font=('Segoe UI', 10))
        project_entry.pack(fill="x", padx=20)

        tk.Label(dlg, text="Sales Order Number", font=('Segoe UI', 10, 'bold')).pack(anchor="w", padx=20, pady=(10, 0))
        so_var = tk.StringVar(value=self.sales_order_no)
        tk.Entry(dlg, textvariable=so_var, font=('Segoe UI', 10)).pack(fill="x", padx=20)

        # Storage Location Frame
        tk.Label(dlg, text="Storage Location", font=('Segoe UI', 10, 'bold')).pack(anchor="w", padx=20, pady=(15, 0))
        
        location_frame = tk.Frame(dlg)
        location_frame.pack(fill="x", padx=20, pady=5)
        
        location_var = tk.StringVar(value=getattr(self, "storage_location", ""))
        location_entry = tk.Entry(location_frame, textvariable=location_var, 
                                 font=('Segoe UI', 9), state='readonly')
        location_entry.pack(side=tk.LEFT, fill="x", expand=True, padx=(0, 5))
        
        def browse_location():
            folder = filedialog.askdirectory(
                title="Select Project Storage Location",
                mustexist=True
            )
            if folder:
                location_var.set(folder)
        
        browse_btn = tk.Button(location_frame, text="Browse...", command=browse_location,
                              bg='#3b82f6', fg='white', font=('Segoe UI', 9, 'bold'),
                              relief=tk.FLAT, padx=15, pady=5)
        browse_btn.pack(side=tk.RIGHT)

        # Auto-load location when project name changes
        def on_project_name_change(*args):
            project_name = project_var.get().strip()
            if project_name:
                # Check database for existing project location
                existing_location = self.db.get_project_location(project_name)
                if existing_location:
                    location_var.set(existing_location)
                    # Show visual feedback
                    location_entry.config(bg='#dcfce7')  # Light green
                    dlg.after(1000, lambda: location_entry.config(bg='white'))
                else:
                    location_var.set("")
        
        # Bind the trace to project name entry
        project_var.trace('w', on_project_name_change)

        def on_ok():
            cabinet = cabinet_var.get().strip()
            project = project_var.get().strip()
            so = so_var.get().strip()
            location = location_var.get().strip()
            
            if not cabinet or not project:
                messagebox.showerror("Missing Information", 
                                   "Please fill in Cabinet ID and Project Name.")
                return
            
            # Check if this cabinet already exists in database
            if self.db.project_exists(cabinet):
                existing = self.db.get_project(cabinet)
                
                # If it's the same project, use existing location
                if existing['project_name'] == project:
                    location = existing['storage_location']
                    messagebox.showinfo("Existing Cabinet", 
                                      f"Cabinet '{cabinet}' found in project '{project}'.\n"
                                      f"Using existing location:\n{location}")
                else:
                    messagebox.showerror("Error", 
                                       f"Cabinet ID '{cabinet}' already exists in different project:\n"
                                       f"{existing['project_name']}")
                    return
            else:
                # New cabinet - check if project exists with different cabinet
                existing_project_location = self.db.get_project_location(project)
                
                if existing_project_location:
                    # Project exists, use its location
                    location = existing_project_location
                    messagebox.showinfo("Existing Project", 
                                      f"Project '{project}' found.\n"
                                      f"Using existing location:\n{location}")
                else:
                    # Brand new project - must have location
                    if not location:
                        messagebox.showerror("Missing Location", 
                                           "This is a new project. Please select a storage location.")
                        return
            
            self.cabinet_id = cabinet
            self.project_name = project
            self.sales_order_no = so
            self.storage_location = location
            
            # Save to database with all paths
            self.db.add_project({
                'cabinet_id': self.cabinet_id,
                'project_name': self.project_name,
                'sales_order_no': self.sales_order_no,
                'storage_location': self.storage_location,
                'pdf_path': self.current_pdf_path if hasattr(self, 'current_pdf_path') else None,
                'excel_path': self.excel_file if hasattr(self, 'excel_file') else None,
                'created_date': datetime.now().isoformat(),
                'last_accessed': datetime.now().isoformat()
            })
            
            dlg.destroy()

        tk.Button(dlg, text="OK", command=on_ok, 
                 bg="#10b981", fg="white", font=('Segoe UI', 10, 'bold'),
                 relief=tk.FLAT, padx=30, pady=10).pack(pady=20)
        dlg.wait_window()

    def write_project_details_to_excel(self):
        if not self.excel_file or not os.path.exists(self.excel_file):
            return

        try:
            wb = load_workbook(self.excel_file)

            for sheet_name, cells in self.header_cells.items():
                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]

                if getattr(self, "project_name", ""):
                    r, c = self.split_cell(cells["project_name"])
                    self.write_cell(ws, r, c, self.project_name)

                if getattr(self, "sales_order_no", ""):
                    r, c = self.split_cell(cells["sales_order"])
                    self.write_cell(ws, r, c, self.sales_order_no)

                if getattr(self, "cabinet_id", ""):
                    r, c = self.split_cell(cells["cabinet_id"])
                    self.write_cell(ws, r, c, self.cabinet_id)

            wb.save(self.excel_file)
            wb.close()

        except PermissionError:
            messagebox.showerror("Excel Locked", "Please close the Excel file before entering project details.")
        except Exception as e:
            messagebox.showerror("Excel Error", f"Failed to write project details:\n{e}")

    def prepare_project_folders(self):
        """Prepare project folders at user-selected location
        Structure: storage_location/project_name/cabinet_id/...
        """
        if not hasattr(self, 'storage_location') or not self.storage_location:
            messagebox.showerror("Error", "Storage location not set")
            return False
        
        if not self.project_name or not self.cabinet_id:
            messagebox.showerror("Error", "Project name and Cabinet ID required")
            return False
        
        # Create structure: storage_location/project_name/cabinet_id/
        project_folder = os.path.join(
            self.storage_location,
            self.project_name.replace(' ', '_')
        )
        
        cabinet_root = os.path.join(
            project_folder,
            self.cabinet_id.replace(' ', '_')
        )
        
        folders = {
            "root": cabinet_root,
            "working_excel": os.path.join(cabinet_root, "Working_Excel"),
            "interphase_export": os.path.join(cabinet_root, "Interphase_Export"),
            "annotated_drawings": os.path.join(cabinet_root, "Annotated_Drawings"),
            "sessions": os.path.join(cabinet_root, "Sessions")
        }
        
        for p in folders.values():
            os.makedirs(p, exist_ok=True)
        
        self.project_dirs = folders
        return True

    def get_session_path_for_pdf(self):
        if not self.current_pdf_path:
            return None

        session_path = os.path.join(
            self.project_dirs.get("sessions", ""),
            f"{self.cabinet_id}_annotations.json"
        )

        return session_path if os.path.exists(session_path) else None

    # ================================================================
    # CHECKLIST FUNCTIONS
    # ================================================================

    def review_checklist_now(self):
        if not self.excel_file or not os.path.exists(self.excel_file):
            messagebox.showerror("Excel Missing", "Working Excel file not found.")
            return

        self.checklist_file = self.excel_file

        try:
            self.review_checklist_before_save(self.checklist_file, self.session_refs)
        except Exception as e:
            messagebox.showerror("Checklist Error", f"Checklist review failed:\n{e}")

    # ================================================================
    # UPDATED: gather_checklist_matches - Updated for new column structure
    # ================================================================
    def review_checklist_before_save(self, checklist_path, refs_set):
        """Modern dialog for reviewing and marking checklist items with date"""
        try:
            cols, matches = self.gather_checklist_matches(checklist_path, refs_set)
        except Exception as e:
            raise

        if not matches:
            messagebox.showinfo("Checklist Complete", 
                              "✓ No items requiring review.\nAll Interphase items are up to date.",
                              icon='info')
            return

        wb = load_workbook(checklist_path)
        ws = wb[self.interphase_sheet_name]
        
        # EXTRACT ALL COLUMNS
        status_col = cols['status_col']
        date_col = cols['date_col']
        name_col = cols['name_col']      # ADD THIS
        remark_col = cols['remark_col']

        # Modern dialog window
        dlg = tk.Toplevel(self.root)
        dlg.title("Interphase Checklist Review")
        dlg.geometry("950x520")
        dlg.configure(bg='#f8fafc')
        dlg.transient(self.root)
        dlg.grab_set()
        
        # Header frame
        header_frame = tk.Frame(dlg, bg='#1e293b', height=60)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="📋 Interphase Checklist Review", 
                bg='#1e293b', fg='white', 
                font=('Segoe UI', 14, 'bold')).pack(pady=15)
        
        # Progress bar
        progress_frame = tk.Frame(dlg, bg='#f8fafc')
        progress_frame.pack(fill=tk.X, padx=20, pady=(15, 5))
        
        idx_label = tk.Label(progress_frame, text="", font=('Segoe UI', 11, 'bold'),
                            bg='#f8fafc', fg='#1e293b')
        idx_label.pack()
        
        # Content frame with modern styling
        content_frame = tk.Frame(dlg, bg='white', relief=tk.FLAT, borderwidth=0)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Reference info frame
        ref_frame = tk.Frame(content_frame, bg='#eff6ff', relief=tk.FLAT)
        ref_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        ref_label = tk.Label(ref_frame, text="", font=('Segoe UI', 10),
                            bg='#eff6ff', fg='#1e40af', anchor='w')
        ref_label.pack(fill=tk.X, padx=15, pady=10)
        
        # Description text
        tk.Label(content_frame, text="Description:", font=('Segoe UI', 9, 'bold'),
                bg='white', fg='#64748b', anchor='w').pack(fill=tk.X, padx=15, pady=(5, 2))
        
        text_widget = tk.Text(content_frame, wrap=tk.WORD, height=12, 
                             font=('Segoe UI', 10), bg='#f8fafc',
                             relief=tk.FLAT, borderwidth=1, padx=10, pady=10)
        text_widget.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))
        text_widget.config(state=tk.DISABLED)

        pos = [0]

        def show_item(p):
            r, ref_str, desc = matches[p]
            
            # Update progress
            progress_text = f"Item {p+1} of {len(matches)}"
            progress_pct = f"({int((p+1)/len(matches)*100)}% complete)"
            idx_label.config(text=f"{progress_text} {progress_pct}")
            
            # Update reference info
            ref_label.config(text=f"📌 Reference: {ref_str}  |  Row: {r}")
            
            # Update description
            text_widget.config(state=tk.NORMAL)
            text_widget.delete('1.0', tk.END)
            text_widget.insert(tk.END, desc)
            text_widget.config(state=tk.DISABLED)

        show_item(pos[0])

        def do_action_set_status(status_value):
            r, ref_str, desc = matches[pos[0]]
            current_date = datetime.now().strftime("%Y-%m-%d")  # NEW

            try:
                self.write_cell(ws, r, status_col, status_value)
                self.write_cell(ws, r, date_col, current_date)  # NEW: Update date
                wb.save(checklist_path)
            except PermissionError:
                messagebox.showerror("File Locked", 
                                   "⚠️ Please close the Excel file and try again.",
                                   icon='error')
                return
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update checklist:\n{e}")
                return

            if pos[0] < len(matches) - 1:
                pos[0] += 1
                show_item(pos[0])
            else:
                messagebox.showinfo("Review Complete", 
                                  f"✓ Checklist review finished!\n{len(matches)} items processed.",
                                  icon='info')
                dlg.destroy()

        def on_ok():
            do_action_set_status("OK")

        def on_nok():
            do_action_set_status("NOK")

        def on_na():
            """Handle N/A status with mandatory remark"""
            r, ref_str, desc = matches[pos[0]]
            
            # Mandatory remark dialog for N/A
            remark = simpledialog.askstring(
                "Remark Required",
                "N/A status requires a remark.\nPlease provide a reason:",
                parent=dlg
            )
            
            if not remark or not remark.strip():
                messagebox.showwarning(
                    "Remark Required",
                    "You must provide a remark for N/A status.",
                    parent=dlg
                )
                return
            
            current_date = datetime.now().strftime("%Y-%m-%d")
            
            try:
                # Get username
                try:
                    username = os.getlogin()
                except:
                    username = getpass.getuser()
                
                # Write all columns properly
                self.write_cell(ws, r, status_col, "N/A")
                self.write_cell(ws, r, date_col, current_date)
                self.write_cell(ws, r, name_col, username)
                self.write_cell(ws, r, remark_col, remark)
                wb.save(checklist_path)
                
                messagebox.showinfo("Remark Saved", 
                                  f"N/A status with remark:\n{remark}",
                                  parent=dlg)
            except PermissionError:
                messagebox.showerror("File Locked", 
                                   "⚠️ Please close the Excel file and try again.",
                                   icon='error',
                                   parent=dlg)
                return
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update checklist:\n{e}",
                                   parent=dlg)
                return

            if pos[0] < len(matches) - 1:
                pos[0] += 1
                show_item(pos[0])
            else:
                messagebox.showinfo("Review Complete", 
                                  f"✓ Checklist review finished!\n{len(matches)} items processed.",
                                  icon='info',
                                  parent=dlg)
                dlg.destroy()


        def on_prev():
            if pos[0] > 0:
                pos[0] -= 1
                show_item(pos[0])

        def on_next():
            if pos[0] < len(matches) - 1:
                pos[0] += 1
                show_item(pos[0])

        # Modern button frame
        btn_frame = tk.Frame(dlg, bg='#f8fafc')
        btn_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        btn_style = {
            'font': ('Segoe UI', 10, 'bold'),
            'relief': tk.FLAT,
            'borderwidth': 0,
            'cursor': 'hand2',
            'padx': 20,
            'pady': 12
        }

        tk.Button(btn_frame, text="◀ Previous", command=on_prev, bg='#94a3b8', 
                 fg='white', width=12, **btn_style).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="✓ OK", command=on_ok, bg='#10b981', 
                 fg='white', width=14, **btn_style).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="✗ NOK", command=on_nok, bg='#ef4444', 
                 fg='white', width=14, **btn_style).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="○ Not Applicable", command=on_na, bg='#f59e0b', 
                 fg='white', width=16, **btn_style).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Next ▶", command=on_next, bg='#94a3b8', 
                 fg='white', width=12, **btn_style).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Cancel", command=lambda: dlg.destroy(), 
                 bg='#64748b', fg='white', width=10, **btn_style).pack(side=tk.RIGHT, padx=5)

        dlg.wait_window()


    # ================================================================
    # UPDATED: gather_checklist_matches - Updated for new column structure
    # ================================================================

    def gather_checklist_matches(self, checklist_path, refs_set):
        """Returns Interphase rows where Reference No is NOT in refs_set."""
        wb = load_workbook(checklist_path)
        if self.interphase_sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError("Interphase sheet not found")

        ws = wb[self.interphase_sheet_name]
        ref_col = self.interphase_cols['ref_no']
        desc_col = self.interphase_cols['description']
        status_col = self.interphase_cols['status']
        name_col = self.interphase_cols['name']
        date_col = self.interphase_cols['date']
        remark_col = self.interphase_cols['remark']

        matches = []
        max_row = ws.max_row if ws.max_row else 2000

        for r in range(11, max_row + 1):
            ref_val = self.read_cell(ws, r, ref_col)
            if ref_val is None:
                continue

            ref_str = str(ref_val).strip()

            if ref_str in refs_set:
                continue

            status_val = self.read_cell(ws, r, status_col)
            status_str = str(status_val).strip().lower() if status_val is not None else ''

            if status_str in ('ok', 'nok', 'n/a', 'na', 'not applicable'):
                continue

            desc_val = self.read_cell(ws, r, desc_col) or ''
            matches.append((r, ref_str, str(desc_val)))

        wb.close()
        return {
            'ref_col': ref_col, 
            'desc_col': desc_col, 
            'status_col': status_col,
            'name_col': name_col,
            'date_col': date_col,
            'remark_col': remark_col
        }, matches    

    # ================================================================
    # EXCEL HELPERS
    # ================================================================

    def save_interphase_excel(self):
        if not self.current_pdf_path:
            messagebox.showwarning("No PDF", "Load a PDF first.")
            return

        if not self.excel_file or not os.path.exists(self.excel_file):
            messagebox.showerror("Missing File", "Working Excel file not found.")
            return

        save_path = os.path.join(
            self.project_dirs["interphase_export"],
            f"{self.cabinet_id.replace(' ', '_')}_Interphase.xlsx"
        )

        try:
            shutil.copy2(self.excel_file, save_path)
            messagebox.showinfo("Saved", f"Final Interphase Excel saved:\n\n{save_path}")
        except PermissionError:
            messagebox.showerror("File Open", "Close the Excel file and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel:\n{e}")

    def open_excel(self):
        if not self.excel_file or not os.path.exists(self.excel_file):
            messagebox.showwarning("No Excel", "No working Excel file found.")
            return

        try:
            if os.name == 'nt':
                os.startfile(self.excel_file)
            else:
                if sys.platform == 'darwin':
                    cmd = f"open {shlex.quote(self.excel_file)}"
                else:
                    cmd = f"xdg-open {shlex.quote(self.excel_file)}"
                subprocess.Popen(cmd, shell=True)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Excel: {e}")

    # ================================================================
    # FUZZY MATCH HELPER
    # ================================================================

    def find_row_by_sr_or_text(self, sr_no, punch_text, min_ratio=0.60):
        try:
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
            row = 8

            while True:
                cell = self.read_cell(ws, row, self.punch_cols['sr_no'])
                if cell is None:
                    if self.read_cell(ws, row, self.punch_cols['desc']) is None:
                        break
                    else:
                        row += 1
                        continue
                try:
                    if int(cell) == int(sr_no):
                        wb.close()
                        return (row, 1.0, 'sr_exact')
                except:
                    if str(cell).strip() == str(sr_no).strip():
                        wb.close()
                        return (row, 1.0, 'sr_exact')
                row += 1
                if row > 2000:
                    break

            best_row = None
            best_ratio = 0.0
            row = 8

            while True:
                txt = self.read_cell(ws, row, self.punch_cols['desc'])
                if txt is None:
                    if row > 2000:
                        break
                    row += 1
                    continue
                try:
                    ratio = SequenceMatcher(None, str(punch_text).strip().lower(), str(txt).strip().lower()).ratio()
                except:
                    ratio = 0.0
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_row = row
                row += 1
                if row > 2000:
                    break

            wb.close()
            if best_row and best_ratio >= min_ratio:
                return (best_row, best_ratio, 'fuzzy_text')
            return (None, best_ratio, None)
        except Exception as e:
            try:
                wb.close()
            except:
                pass
            return (None, 0.0, None)

    # ================================================================
    # HANDOVER TO PRODUCTION
    # ================================================================
    def handover_to_production(self):
        """Handover current cabinet to production"""
        
        if not self.pdf_document or not self.excel_file:
            messagebox.showwarning("Incomplete", 
                                  "Please load a PDF and Excel file first.")
            return
        
        if not self.cabinet_id or not self.project_name:
            messagebox.showwarning("Missing Info", 
                                  "Project details are incomplete.")
            return
        
        # Check if there are any annotations
        if not self.annotations:
            proceed = messagebox.askyesno(
                "No Annotations",
                "No annotations found. Handover anyway?",
                icon='warning'
            )
            if not proceed:
                return
        
        # Count open punches
        open_punches = self.count_open_punches()
        
        if open_punches > 0:
            proceed = messagebox.askyesno(
                "Open Punches Detected",
                f"⚠️ There are {open_punches} open punch(es).\n\n"
                "Are you sure you want to handover to production?",
                icon='warning'
            )
            if not proceed:
                return
        
        # Save session before handover
        self.save_session()
        
        # Get user name
        try:
            username = os.getlogin()
        except:
            username = getpass.getuser()
        
        # Prepare handover data
        session_path = os.path.join(
            self.project_dirs.get("sessions", ""),
            f"{self.cabinet_id}_annotations.json"
        )
        
        handover_data = {
            "cabinet_id": self.cabinet_id,
            "project_name": self.project_name,
            "sales_order_no": self.sales_order_no,
            "pdf_path": self.current_pdf_path,
            "excel_path": self.excel_file,
            "session_path": session_path if os.path.exists(session_path) else None,
            "total_punches": len([a for a in self.annotations if a.get('type') == 'error']),
            "open_punches": open_punches,
            "closed_punches": len([a for a in self.annotations if a.get('type') == 'error']) - open_punches,
            "handed_over_by": username,
            "handed_over_date": datetime.now().isoformat()
        }
        
        # FIXED: Use handover_db instead of db
        success = self.handover_db.add_quality_handover(handover_data)
        
        # FIXED: Use correct method name
        self.manager_db.update_status(self.cabinet_id, 'handed_to_production')
        
        if success:
            messagebox.showinfo("Handover Complete", 
                              "✓ Successfully handed over to Production\n\n"
                              f"Cabinet: {self.cabinet_id}\n"
                              f"Open Punches: {open_punches}")
        else:
            messagebox.showwarning("Already Handed Over", 
                                 "Cabinet already in production queue")

    def count_open_punches(self) -> int:
        """Count open punches in current Excel"""
        try:
            if not self.excel_file or not os.path.exists(self.excel_file):
                return 0
            
            wb = load_workbook(self.excel_file, data_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
            
            open_count = 0
            row = 8
            
            while row <= ws.max_row + 5:
                sr = self.read_cell(ws, row, self.punch_cols['sr_no'])
                if sr is None:
                    break
                
                closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
                if not closed:
                    open_count += 1
                
                row += 1
            
            wb.close()
            return open_count
            
        except Exception as e:
            print(f"Error counting open punches: {e}")
            return 0

    # ================================================================
    # VIEW PRODUCTION HANDBACK ITEMS
    # ================================================================

    def sync_manager_stats(self):
        """Sync current cabinet statistics to manager database"""
        if not self.pdf_document or not self.cabinet_id:
            return
        
        try:
            # Count pages with annotations
            annotated_pages = len(set(ann['page'] for ann in self.annotations if ann.get('page') is not None))
            total_pages = len(self.pdf_document)
            
            # Count punches by type
            error_anns = [a for a in self.annotations if a.get('type') == 'error']
            total_punches = len(error_anns)
            
            # Count from Excel for accuracy
            open_punches = self.count_open_punches()
            
            # Count implemented (has implemented_name but no closed_name)
            implemented_punches = 0
            closed_punches = 0
            
            if self.excel_file and os.path.exists(self.excel_file):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(self.excel_file, data_only=True)
                    ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
                    
                    row = 8
                    while row <= ws.max_row + 5:
                        sr = self.read_cell(ws, row, self.punch_cols['sr_no'])
                        if sr is None:
                            break
                        
                        implemented = self.read_cell(ws, row, self.punch_cols['implemented_name'])
                        closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
                        
                        if closed:
                            closed_punches += 1
                        elif implemented:
                            implemented_punches += 1
                        
                        row += 1
                    
                    wb.close()
                except:
                    pass
            
            # Update manager database
            self.manager_db.update_cabinet(
                self.cabinet_id,
                self.project_name,
                self.sales_order_no,
                total_pages,
                annotated_pages,
                total_punches,
                open_punches,
                implemented_punches,
                closed_punches,
                'quality_inspection'
            )
        except Exception as e:
            print(f"Manager sync error: {e}")

            
    # ============================================================================
    # UPDATED: view_production_handbacks - Auto-open punch closing
    # ============================================================================

    def view_production_handbacks(self):
        """View and verify items returned from production"""
        
        pending_items = self.handover_db.get_pending_quality_items()
        
        if not pending_items:
            messagebox.showinfo(
                "No Items",
                "✓ No items pending verification from production.",
                icon='info'
            )
            return
        
        # Create modern dialog
        dlg = tk.Toplevel(self.root)
        dlg.title("Production Handback - Verification")
        dlg.geometry("1000x600")
        dlg.configure(bg='#f8fafc')
        dlg.transient(self.root)
        dlg.grab_set()
        
        # Header
        header_frame = tk.Frame(dlg, bg='#7c3aed', height=60)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="🔍 Production Handback Verification", 
                bg='#7c3aed', fg='white', 
                font=('Segoe UI', 14, 'bold')).pack(pady=15)
        
        # Info bar
        info_frame = tk.Frame(dlg, bg='#eff6ff')
        info_frame.pack(fill=tk.X, padx=20, pady=(15, 5))
        
        tk.Label(info_frame, text=f"Total items pending verification: {len(pending_items)}", 
                bg='#eff6ff', fg='#1e40af', 
                font=('Segoe UI', 10, 'bold')).pack(pady=8)
        
        # Listbox frame
        list_frame = tk.Frame(dlg, bg='white')
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        tk.Label(list_frame, text="Select item to verify:", 
                font=('Segoe UI', 10, 'bold'),
                bg='white', fg='#1e293b').pack(anchor='w', pady=(0, 10))
        
        # Scrollbar and Listbox
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        listbox = tk.Listbox(list_frame, font=('Consolas', 9),
                            yscrollcommand=scrollbar.set,
                            bg='#f8fafc', relief=tk.FLAT,
                            selectmode=tk.SINGLE, height=15)
        listbox.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=listbox.yview)
        
        # Populate listbox
        for item in pending_items:
            display_text = (
                f"{item['cabinet_id']:20} | {item['project_name']:30} | "
                f"Rework by: {item['rework_completed_by']:15} | "
                f"{item['rework_completed_date'][:10]}"
            )
            listbox.insert(tk.END, display_text)
        
        def load_selected():
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select an item first.")
                return
            
            item = pending_items[selection[0]]
            
            # Load the project
            try:
                project_data = self.db.get_project(item['cabinet_id'])
                if not project_data:
                    messagebox.showerror("Error", "Project not found in database")
                    return
                
                self.cabinet_id = item['cabinet_id']
                self.project_name = item['project_name']
                self.sales_order_no = item['sales_order_no']
                self.storage_location = project_data['storage_location']
                
                self.prepare_project_folders()
                
                if not os.path.exists(item['pdf_path']):
                    messagebox.showerror("Error", f"PDF file not found:\n{item['pdf_path']}")
                    return
                
                self.pdf_document = fitz.open(item['pdf_path'])
                self.current_pdf_path = item['pdf_path']
                self.current_page = 0
                self.zoom_level = 1.0
                self.tool_mode = None
                self.root.config(cursor="")
                
                self.excel_file = item['excel_path']
                self.working_excel_path = item['excel_path']
                
                self.current_sr_no = self.get_next_sr_no()
                
                if item.get('session_path') and os.path.exists(item['session_path']):
                    self.load_session_from_path(item['session_path'])
                else:
                    self.annotations = []
                    self.display_page()
                
                # UPDATED: Set status to "Rework being verified"
                self.manager_db.update_status(self.cabinet_id, 'being_closed_by_quality')
                
                dlg.destroy()
                
                # UPDATED: Auto-open punch closing dialog
                self.verify_production_work_with_punch_closing(item)
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load item:\n{e}")
                import traceback
                traceback.print_exc()
        
        # Buttons
        btn_frame = tk.Frame(dlg, bg='#f8fafc')
        btn_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        btn_style = {
            'font': ('Segoe UI', 10, 'bold'),
            'relief': tk.FLAT,
            'cursor': 'hand2',
            'padx': 20,
            'pady': 12
        }
        
        tk.Button(btn_frame, text="📂 Load & Verify", command=load_selected,
                 bg='#3b82f6', fg='white', **btn_style).pack(side=tk.LEFT, padx=5)
        
        # REMOVED: Quick Verify button as requested
        
        tk.Button(btn_frame, text="Cancel", command=dlg.destroy,
                 bg='#64748b', fg='white', **btn_style).pack(side=tk.RIGHT, padx=5)
        
        listbox.bind('<Double-Button-1>', lambda e: load_selected())


    # ============================================================================
    # NEW: verify_production_work_with_punch_closing - Auto-open punch closing
    # ============================================================================

    def verify_production_work_with_punch_closing(self, item_data):
        """Auto-open punch closing mode after loading production item"""
        
        # Count open punches
        open_count = self.count_open_punches()
        
        # Show initial info
        info_msg = (
            f"Cabinet: {item_data['cabinet_id']}\n"
            f"Project: {item_data['project_name']}\n\n"
            f"Rework by: {item_data['rework_completed_by']}\n"
            f"Date: {item_data['rework_completed_date'][:10]}\n\n"
            f"Open Punches: {open_count}\n\n"
            f"Opening Punch Closing Mode..."
        )
        
        messagebox.showinfo("Production Handback Loaded", info_msg, icon='info')
        
        # Auto-open punch closing mode
        if open_count > 0:
            self.punch_closing_mode_for_verification(item_data)
        else:
            # No punches - directly ask to close
            self.finalize_verification(item_data)


    # ============================================================================
    # NEW: punch_closing_mode_for_verification - Modified punch closing for handback
    # ============================================================================

    def punch_closing_mode_for_verification(self, item_data):
        """Punch closing mode specifically for verification workflow"""
        
        punches = self.read_open_punches_from_excel()
        
        if not punches:
            # All closed, proceed to finalization
            self.finalize_verification(item_data)
            return
        
        punches.sort(key=lambda p: (not p['implemented'], p['sr_no']))
        
        # Dialog
        dlg = tk.Toplevel(self.root)
        dlg.title("Punch Verification Mode")
        dlg.geometry("950x600")
        dlg.configure(bg='#f8fafc')
        dlg.transient(self.root)
        dlg.grab_set()
        
        # Header
        header_frame = tk.Frame(dlg, bg='#7c3aed', height=50)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="✓ Punch Verification Mode", 
                bg='#7c3aed', fg='white', 
                font=('Segoe UI', 13, 'bold')).pack(pady=12)
        
        # Progress
        progress_frame = tk.Frame(dlg, bg='#f8fafc')
        progress_frame.pack(fill=tk.X, padx=20, pady=(10, 5))
        
        idx_label = tk.Label(progress_frame, text="", font=('Segoe UI', 10, 'bold'),
                            bg='#f8fafc', fg='#1e293b')
        idx_label.pack()
        
        # Info cards
        info_frame = tk.Frame(dlg, bg='#f8fafc')
        info_frame.pack(fill=tk.X, padx=20, pady=8)
        
        sr_card = tk.Frame(info_frame, bg='#dbeafe', relief=tk.FLAT)
        sr_card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        tk.Label(sr_card, text="SR No.", font=('Segoe UI', 8), 
                bg='#dbeafe', fg='#1e40af').pack(anchor='w', padx=10, pady=(6, 2))
        sr_label = tk.Label(sr_card, text="", font=('Segoe UI', 12, 'bold'),
                           bg='#dbeafe', fg='#1e293b')
        sr_label.pack(anchor='w', padx=10, pady=(0, 6))
        
        ref_card = tk.Frame(info_frame, bg='#e0e7ff', relief=tk.FLAT)
        ref_card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        tk.Label(ref_card, text="Reference", font=('Segoe UI', 8), 
                bg='#e0e7ff', fg='#4338ca').pack(anchor='w', padx=10, pady=(6, 2))
        ref_label = tk.Label(ref_card, text="", font=('Segoe UI', 12, 'bold'),
                            bg='#e0e7ff', fg='#1e293b')
        ref_label.pack(anchor='w', padx=10, pady=(0, 6))
        
        status_card = tk.Frame(info_frame, bg='#fef3c7', relief=tk.FLAT)
        status_card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
        
        tk.Label(status_card, text="Status", font=('Segoe UI', 8), 
                bg='#fef3c7', fg='#92400e').pack(anchor='w', padx=10, pady=(6, 2))
        impl_label = tk.Label(status_card, text="", font=('Segoe UI', 12, 'bold'),
                             bg='#fef3c7', fg='#1e293b')
        impl_label.pack(anchor='w', padx=10, pady=(0, 6))
        
        # Content
        content_frame = tk.Frame(dlg, bg='white', relief=tk.FLAT)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=8)
        
        tk.Label(content_frame, text="Punch Description:", font=('Segoe UI', 9, 'bold'),
                bg='white', fg='#64748b', anchor='w').pack(fill=tk.X, padx=15, pady=(8, 3))
        
        text_widget = tk.Text(content_frame, wrap=tk.WORD, height=9,
                             font=('Segoe UI', 10), bg='#f8fafc',
                             relief=tk.FLAT, padx=10, pady=8)
        text_widget.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 10))
        text_widget.config(state=tk.DISABLED)

        pos = [0]

        def show_item():
            p = punches[pos[0]]
            
            progress_text = f"Item {pos[0]+1} of {len(punches)}"
            progress_pct = f"({int((pos[0]+1)/len(punches)*100)}% complete)"
            idx_label.config(text=f"{progress_text} {progress_pct}")
            
            sr_label.config(text=str(p['sr_no']))
            ref_label.config(text=str(p['ref_no']))
            
            impl_status = "✓ Implemented" if p['implemented'] else "⚠ Not Implemented"
            impl_color = '#10b981' if p['implemented'] else '#f59e0b'
            impl_label.config(text=impl_status, fg=impl_color)
            
            text_widget.config(state=tk.NORMAL)
            text_widget.delete("1.0", tk.END)
            text_widget.insert(tk.END, p['punch_text'])
            text_widget.insert(tk.END, f"\n\n──────────────────\n")
            text_widget.insert(tk.END, f"Category: {p['category']}\n")

            ann = next((a for a in self.annotations if a.get('sr_no') == p['sr_no']), None)
            if ann and ann.get('implementation_remark'):
                text_widget.insert(tk.END, f"\n──────────────────\n")
                text_widget.insert(tk.END, "Implementation Remarks:\n")
                text_widget.insert(tk.END, ann['implementation_remark'])
            
            # Show production remarks
            try:
                handover_data = self.handover_db.get_handover_by_cabinet(self.cabinet_id)
                if handover_data and handover_data.get('production_remarks'):
                    text_widget.insert(tk.END, f"\n\n──────────────────\n")
                    text_widget.insert(tk.END, "🔧 Production Remarks:\n")
                    text_widget.insert(tk.END, handover_data['production_remarks'])
                    text_widget.insert(tk.END, f"\n\nRework by: {handover_data.get('rework_completed_by', 'N/A')}")
                    text_widget.insert(tk.END, f"\nDate: {handover_data.get('rework_completed_date', 'N/A')[:10]}")
            except Exception as e:
                print(f"Could not load production remarks: {e}")

            text_widget.config(state=tk.DISABLED)

        show_item()

        def close_punch():
            p = punches[pos[0]]

            try:
                default_user = os.getlogin()
            except:
                default_user = getpass.getuser()

            name = simpledialog.askstring("Closed By", 
                                         "Enter your name to close this punch:", 
                                         initialvalue=default_user, 
                                         parent=dlg)
            if not name:
                return

            try:
                wb = load_workbook(self.excel_file)
                ws = wb[self.punch_sheet_name]

                self.write_cell(ws, p['row'], self.punch_cols['closed_name'], name)
                self.write_cell(ws, p['row'], self.punch_cols['closed_date'], 
                              datetime.now().strftime("%Y-%m-%d"))

                wb.save(self.excel_file)
                wb.close()

            except PermissionError:
                messagebox.showerror("File Locked", 
                                   "⚠️ Please close the Excel file and try again.")
                return
            except Exception as e:
                messagebox.showerror("Error", f"Failed to close punch:\n{e}")
                return

            ann = next((a for a in self.annotations if a.get('excel_row') == p['row']), None)
            if ann:
                ann['type'] = 'ok'
                ann['closed_by'] = name
                ann['closed_date'] = datetime.now().strftime("%Y-%m-%d")

            self.display_page()

            if pos[0] < len(punches) - 1:
                pos[0] += 1
                show_item()
            else:
                messagebox.showinfo("All Punches Closed", 
                                  f"✓ All punches verified and closed!\n{len(punches)} items processed.",
                                  icon='info')
                dlg.destroy()
                # Proceed to finalization
                self.finalize_verification(item_data)

        def next_item():
            if pos[0] < len(punches) - 1:
                pos[0] += 1
                show_item()

        def prev_item():
            if pos[0] > 0:
                pos[0] -= 1
                show_item()

        # Buttons
        btn_frame = tk.Frame(dlg, bg='#f8fafc', height=80)
        btn_frame.pack(fill=tk.X, padx=20, pady=(10, 25))
        btn_frame.pack_propagate(False)
        
        btn_container = tk.Frame(btn_frame, bg='#f8fafc')
        btn_container.pack(expand=True)
        
        btn_style = {
            'font': ('Segoe UI', 12, 'bold'),
            'relief': tk.FLAT,
            'borderwidth': 0,
            'cursor': 'hand2',
            'padx': 35,
            'pady': 18,
            'width': 15
        }

        tk.Button(btn_container, text="◀  Previous", command=prev_item, 
                 bg='#94a3b8', fg='white', **btn_style).pack(side=tk.LEFT, padx=8)
        
        close_btn_style = btn_style.copy()
        close_btn_style['width'] = 18
        tk.Button(btn_container, text="✓  CLOSE PUNCH", command=close_punch, 
                 bg='#10b981', fg='white', **close_btn_style).pack(side=tk.LEFT, padx=8)
        
        tk.Button(btn_container, text="Next  ▶", command=next_item, 
                 bg='#94a3b8', fg='white', **btn_style).pack(side=tk.LEFT, padx=8)
        
        tk.Button(btn_container, text="Cancel", command=dlg.destroy, 
                 bg='#64748b', fg='white', **btn_style).pack(side=tk.LEFT, padx=8)

        dlg.wait_window()


    # ============================================================================
    # NEW: finalize_verification - Check checklist, save Excel, export PDF
    # ============================================================================

    def finalize_verification(self, item_data):
        """Final step: check checklist completeness, save Excel, export PDF"""
        
        # 1. Check checklist completeness
        checklist_complete, incomplete_refs = self.check_checklist_completeness()
        
        if not checklist_complete:
            warning_msg = (
                f"⚠️ Checklist Incomplete!\n\n"
                f"{len(incomplete_refs)} reference(s) without status:\n"
                f"{', '.join(incomplete_refs[:10])}"
            )
            if len(incomplete_refs) > 10:
                warning_msg += f"\n... and {len(incomplete_refs) - 10} more"
            
            warning_msg += "\n\nDo you want to review the checklist now?"
            
            review = messagebox.askyesno("Incomplete Checklist", warning_msg, icon='warning')
            if review:
                self.review_checklist_now()
                # After review, ask again if they want to finalize
                retry = messagebox.askyesno(
                    "Continue?",
                    "Review complete. Proceed with closing the cabinet?",
                    icon='question'
                )
                if not retry:
                    return
                
                # Re-check completeness after review
                checklist_complete, incomplete_refs = self.check_checklist_completeness()
                if not checklist_complete:
                    messagebox.showwarning(
                        "Still Incomplete",
                        f"Checklist still has {len(incomplete_refs)} incomplete item(s).\n"
                        "Proceeding anyway...",
                        icon='warning'
                    )
            else:
                # User chose not to review - warn but allow to continue
                proceed = messagebox.askyesno(
                    "Continue Anyway?",
                    "Checklist is incomplete. Continue with closing?",
                    icon='warning'
                )
                if not proceed:
                    return
        
        # 2. Confirm closing
        confirm_msg = (
            f"Cabinet: {item_data['cabinet_id']}\n"
            f"Project: {item_data['project_name']}\n\n"
            f"✓ All punches closed\n"
            f"{'✓' if checklist_complete else '⚠'} Checklist {'complete' if checklist_complete else 'reviewed'}\n\n"
            f"This will:\n"
            f"• Save Interphase Excel\n"
            f"• Export Annotated PDF\n"
            f"• Mark cabinet as CLOSED\n\n"
            f"Proceed?"
        )
        
        confirm = messagebox.askyesno("Close Cabinet", confirm_msg, icon='question')
        if not confirm:
            return
        
        # 3. Save current session first
        try:
            self.save_session()
        except Exception as e:
            print(f"Session save warning: {e}")
        
        # 4. Auto-save Interphase Excel
        try:
            if not self.excel_file or not os.path.exists(self.excel_file):
                messagebox.showerror("Error", "Working Excel file not found.")
                return
            
            save_path = os.path.join(
                self.project_dirs["interphase_export"],
                f"{self.cabinet_id.replace(' ', '_')}_Interphase.xlsx"
            )
            
            shutil.copy2(self.excel_file, save_path)
            print(f"✓ Interphase Excel saved: {save_path}")
            
        except PermissionError:
            messagebox.showerror("Error", "Excel file is open. Please close it and try again.")
            return
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel:\n{e}")
            return
        
        # 5. Auto-export annotated PDF
        try:
            if not self.pdf_document:
                messagebox.showerror("Error", "No PDF document loaded.")
                return
            
            export_path = os.path.join(
                self.project_dirs["annotated_drawings"],
                f"{self.cabinet_id.replace(' ', '_')}_Annotated.pdf"
            )
            
            # Create output PDF
            out_doc = fitz.open()
            for pnum in range(len(self.pdf_document)):
                out_doc.insert_pdf(self.pdf_document, from_page=pnum, to_page=pnum)
            
            # Open Excel for SR No lookup
            wb = None
            ws = None
            if self.excel_file and os.path.exists(self.excel_file):
                try:
                    wb = load_workbook(self.excel_file, data_only=True)
                    ws = wb[self.punch_sheet_name]
                except:
                    pass
            
            # Draw annotations
            for ann in self.annotations:
                p = ann.get('page')
                if p is None or p < 0 or p >= len(out_doc):
                    continue
                
                target_page = out_doc[p]
                ann_type = ann.get('type')
                
                # Rectangle annotations
                if ann_type in ('ok', 'error') and 'bbox_page' in ann:
                    x1, y1, x2, y2 = ann['bbox_page']
                    rect = self.transform_bbox_for_rotation((x1, y1, x2, y2), target_page)
                    
                    if ann_type == 'ok':
                        target_page.draw_rect(rect, color=(0, 1, 0), width=2)
                    else:
                        target_page.draw_rect(rect, color=(1, 0.55, 0), width=2)
                    
                    sr_text = None
                    row = ann.get('excel_row')
                    
                    if row and ws:
                        try:
                            sr_val = self.read_cell(ws, row, self.punch_cols['sr_no'])
                            if sr_val is not None:
                                sr_text = f"Sr {sr_val}"
                        except:
                            sr_text = None
                    
                    if sr_text:
                        text_pos = fitz.Point(rect.x0, max(rect.y0 - 12, rect.y0))
                        try:
                            target_page.insert_text(text_pos, sr_text, fontsize=8)
                        except:
                            pass
                
                # Pen strokes
                elif ann_type == 'pen' and 'points' in ann:
                    points = ann['points']
                    if len(points) >= 2:
                        transformed_points = [
                            self.transform_point_for_rotation(pt, target_page) 
                            for pt in points
                        ]
                        
                        for i in range(len(transformed_points) - 1):
                            p1 = transformed_points[i]
                            p2 = transformed_points[i + 1]
                            target_page.draw_line(p1, p2, color=(1, 0, 0), width=2)
                
                # Text annotations
                elif ann_type == 'text' and 'pos_page' in ann:
                    pos = ann['pos_page']
                    text = ann.get('text', '')
                    if text:
                        text_point = self.transform_point_for_rotation(pos, target_page)
                        try:
                            target_page.insert_text(
                                text_point, text,
                                fontsize=10, color=(1, 0, 0),
                                rotate=target_page.rotation
                            )
                        except:
                            pass
            
            if wb:
                wb.close()
            
            out_doc.save(export_path)
            out_doc.close()
            
            print(f"✓ Annotated PDF exported: {export_path}")
            
        except PermissionError:
            messagebox.showerror("Error", "PDF file is open. Please close it and try again.")
            return
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export PDF:\n{e}")
            return
        
        # 6. Update verification status
        try:
            username = os.getlogin()
        except:
            username = getpass.getuser()
        
        success = self.handover_db.update_quality_verification(
            item_data['cabinet_id'],
            status="closed",
            user=username
        )
        
        if success:
            # Update manager status to 'closed'
            self.manager_db.update_status(item_data['cabinet_id'], 'closed')
            
            # Sync final stats
            self.sync_manager_stats()
            
            messagebox.showinfo(
                "✓ Verification Complete",
                f"Cabinet {item_data['cabinet_id']} successfully closed!\n\n"
                f"✓ Interphase Excel saved\n"
                f"✓ Annotated PDF exported\n"
                f"✓ Status updated to CLOSED\n\n"
                f"Files saved to:\n{self.project_dirs['root']}",
                icon='info'
            )
        else:
            messagebox.showerror("Error", "Failed to update verification status.")


    # ============================================================================
    # NEW: check_checklist_completeness - Helper to check Interphase completion
    # ============================================================================

    def check_checklist_completeness(self):
        """Check if all Interphase items have status (OK/NOK/N/A)
        Returns: (is_complete: bool, incomplete_refs: list)
        """
        if not self.excel_file or not os.path.exists(self.excel_file):
            return False, []
        
        try:
            wb = load_workbook(self.excel_file, data_only=True)
            if self.interphase_sheet_name not in wb.sheetnames:
                wb.close()
                return False, []
            
            ws = wb[self.interphase_sheet_name]
            ref_col = self.interphase_cols['ref_no']
            status_col = self.interphase_cols['status']
            
            incomplete_refs = []
            max_row = ws.max_row if ws.max_row else 2000
            
            for r in range(11, max_row + 1):
                ref_val = self.read_cell(ws, r, ref_col)
                if ref_val is None:
                    continue
                
                ref_str = str(ref_val).strip()
                if not ref_str:
                    continue
                
                status_val = self.read_cell(ws, r, status_col)
                status_str = str(status_val).strip().lower() if status_val is not None else ''
                
                # Check if status is empty or invalid
                if status_str not in ('ok', 'nok', 'n/a', 'na', 'not applicable'):
                    incomplete_refs.append(ref_str)
            
            wb.close()
            
            is_complete = len(incomplete_refs) == 0
            return is_complete, incomplete_refs
            
        except Exception as e:
            print(f"Checklist check error: {e}")
            return False, []




# ================================================================
# MAIN ENTRY POINT
# ================================================================

def main():
    root = tk.Tk()
    app = CircuitInspector(root)
    root.mainloop()


if __name__ == "__main__":
    main()
